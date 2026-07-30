from __future__ import annotations

import base64
import hashlib
import hmac
import json
import time
import zipfile
import io
from pathlib import Path

import openpyxl
import pytest
import xlrd
from fastapi.testclient import TestClient

from app.main import app, clear_reconstruction_rate_limits
from app.master_data_client import ConversionContextError
from app.reconstruction_profile_client import ReconstructionProfileClientError


client = TestClient(app)


@pytest.fixture(autouse=True)
def _internal_service_auth(monkeypatch):
    token = "reconstruction-service-token"
    monkeypatch.setenv("CONVERTER_SERVICE_TOKEN", token)
    client.headers["x-converter-service-token"] = token
    yield
    client.headers.pop("x-converter-service-token", None)


def test_converter_cors_exposes_download_filename_header():
    response = client.get(
        "/healthz",
        headers={"Origin": "http://localhost:5173"},
    )

    assert response.status_code == 200
    assert "content-disposition" in response.headers[
        "access-control-expose-headers"
    ].lower()
    assert response.headers["x-request-id"]


def test_reconstruction_rejects_external_links_before_persisting_state(
    tmp_path,
    monkeypatch,
):
    monkeypatch.setenv("VOUCHER_RECONSTRUCTION_ENABLED", "true")
    monkeypatch.setenv("CONVERSION_CONTEXT_SECRET", "reconstruction-test-secret")
    monkeypatch.setenv("RECONSTRUCTION_STORE_PROVIDER", "filesystem")
    monkeypatch.setenv("RECONSTRUCTION_STORE_DIR", str(tmp_path / "store"))
    monkeypatch.setenv("RECONSTRUCTION_NOTIFY_NODE", "false")
    monkeypatch.setattr(
        "app.reconstruction_workflow.inspect_workbook_structure",
        lambda _path: {"has_external_links": True},
    )
    raw = _write_purchase(tmp_path / "external.xlsx")

    with raw.open("rb") as handle:
        response = client.post(
            "/api/v1/reconstructions/analyze",
            data={"context_token": _token("external-run"), "mode": "purchase"},
            files={
                "file": (
                    raw.name,
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 400
    assert "external links" in response.json()["detail"]
    assert not (tmp_path / "store" / "external-run.json").exists()


def test_reconstruction_analyze_is_rate_limited_per_run(tmp_path, monkeypatch):
    clear_reconstruction_rate_limits()
    monkeypatch.setenv("VOUCHER_RECONSTRUCTION_ENABLED", "true")
    monkeypatch.setenv("CONVERSION_CONTEXT_SECRET", "reconstruction-test-secret")
    monkeypatch.setenv("RECONSTRUCTION_STORE_PROVIDER", "filesystem")
    monkeypatch.setenv("RECONSTRUCTION_STORE_DIR", str(tmp_path / "store"))
    monkeypatch.setenv("RECONSTRUCTION_NOTIFY_NODE", "false")
    monkeypatch.setenv("RECONSTRUCTION_ANALYZE_LIMIT_PER_15_MINUTES", "1")
    token = _token("rate-limit-run")
    raw = _write_purchase(tmp_path / "rate-limit.xlsx")

    with raw.open("rb") as handle:
        first = client.post(
            "/api/v1/reconstructions/analyze",
            data={"context_token": token, "mode": "purchase"},
            files={
                "file": (
                    raw.name,
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert first.status_code == 200, first.text
    with raw.open("rb") as handle:
        second = client.post(
            "/api/v1/reconstructions/analyze",
            data={"context_token": token, "mode": "purchase"},
            files={
                "file": (
                    raw.name,
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert second.status_code == 429
    clear_reconstruction_rate_limits()


def _encode(value: dict) -> str:
    raw = json.dumps(value, separators=(",", ":")).encode("utf-8")
    return base64.urlsafe_b64encode(raw).rstrip(b"=").decode("ascii")


def _token(
    run_id: str,
    secret: str = "reconstruction-test-secret",
    *,
    workspace_id: str = "",
) -> str:
    header = _encode({"alg": "HS256", "typ": "JWT"})
    payload = _encode(
        {
            "purpose": "misa_reconstruction",
            "user_id": "user-1",
            "run_id": run_id,
            "workspace_id": workspace_id,
            "snapshot_set_hash": "snapshot-hash" if workspace_id else "",
            "snapshot_ids": ["snapshot-1"] if workspace_id else [],
            "master_data_revision": 1 if workspace_id else 0,
            "scopes": ["analyze", "review", "approve", "export"],
            "iat": int(time.time()),
            "exp": int(time.time()) + 3600,
        }
    )
    signature = hmac.new(
        secret.encode("utf-8"),
        f"{header}.{payload}".encode("ascii"),
        hashlib.sha256,
    ).digest()
    encoded_signature = base64.urlsafe_b64encode(signature).rstrip(b"=").decode("ascii")
    return f"{header}.{payload}.{encoded_signature}"


def _write_purchase(path: Path) -> Path:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "ChiTietHoaDon"
    sheet.append(
        [
            "Số hóa đơn",
            "Ký hiệu hóa đơn",
            "Ngày hóa đơn",
            "Ngày chứng từ",
            "Mã số thuế NCC",
            "Tên nhà cung cấp",
            "Phân loại",
            "Mã hàng",
            "Tên hàng hóa dịch vụ",
            "Đơn vị tính",
            "Số lượng",
            "Đơn giá",
            "Thành tiền chưa thuế",
            "Thuế suất GTGT",
            "Tiền thuế GTGT",
        ]
    )
    sheet.append(
        [
            "000123",
            "1C26TAA",
            "01/07/2026",
            "01/07/2026",
            "0311111111",
            "Nhà cung cấp A",
            "Hàng hóa",
            "HH01",
            "Hàng 1",
            "Cái",
            2,
            100_000,
            200_000,
            10,
            20_000,
        ]
    )
    sheet.append(
        [
            "",
            "",
            "",
            "",
            "",
            "",
            "Hàng hóa",
            "HH02",
            "Hàng 2",
            "Cái",
            1,
            50_000,
            50_000,
            10,
            5_000,
        ]
    )
    workbook.save(path)
    return path


def test_reconstruction_ai_is_not_called_without_explicit_opt_in(tmp_path, monkeypatch):
    clear_reconstruction_rate_limits()
    monkeypatch.setenv("VOUCHER_RECONSTRUCTION_ENABLED", "true")
    monkeypatch.setenv("CONVERSION_CONTEXT_SECRET", "reconstruction-test-secret")
    monkeypatch.setenv("RECONSTRUCTION_STORE_PROVIDER", "filesystem")
    monkeypatch.setenv("RECONSTRUCTION_STORE_DIR", str(tmp_path / "store"))
    monkeypatch.setenv("RECONSTRUCTION_NOTIFY_NODE", "false")
    monkeypatch.setenv("AI_PROVIDER", "remote_http")
    calls = []
    monkeypatch.setattr(
        "app.reconstruction_workflow.request_reconstruction_suggestion",
        lambda *_args, **_kwargs: calls.append(True)
        or {
            "field_roles": {},
            "grouping_keys": [],
            "direction": "unknown",
            "nature": "unknown",
            "confidence": 0,
            "notes": [],
        },
        raising=False,
    )
    raw = tmp_path / "no-auto-ai.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Cột A", "Cột B"])
    sheet.append(["A1", "Nội dung chưa phân loại"])
    workbook.save(raw)
    token = _token("no-auto-ai")

    with raw.open("rb") as handle:
        response = client.post(
            "/api/v1/reconstructions/analyze",
            data={"context_token": token, "mode": "purchase"},
            files={
                "file": (
                    raw.name,
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200, response.text
    assert response.json()["ai"]["used"] is False
    assert calls == []


def test_reconstruction_analyze_review_approve_and_export(tmp_path, monkeypatch):
    monkeypatch.setenv("VOUCHER_RECONSTRUCTION_ENABLED", "true")
    monkeypatch.setenv("CONVERSION_CONTEXT_SECRET", "reconstruction-test-secret")
    monkeypatch.setenv("RECONSTRUCTION_STORE_PROVIDER", "filesystem")
    monkeypatch.setenv("RECONSTRUCTION_STORE_DIR", str(tmp_path / "store"))
    monkeypatch.setenv("RECONSTRUCTION_NOTIFY_NODE", "false")
    run_id = "run-api-1"
    token = _token(run_id)
    raw = _write_purchase(tmp_path / "purchase.xlsx")

    with raw.open("rb") as handle:
        analyze = client.post(
            "/api/v1/reconstructions/analyze",
            data={
                "context_token": token,
                "mode": "purchase",
                "target_template_id": "misa_purchase_domestic",
            },
            files={
                "file": (
                    raw.name,
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert analyze.status_code == 200, analyze.text
    payload = analyze.json()
    assert payload["reconstruction_id"] == run_id
    assert payload["summary"]["draft_count"] == 1
    assert payload["row_conservation"]["assigned_rows"] == 2
    draft = payload["drafts"][0]
    assert draft["header"]["invoice_number"]["value"] == "000123"

    fetched = client.get(
        f"/api/v1/reconstructions/{run_id}?page=1&limit=1",
        headers={"x-reconstruction-context": token},
    )
    assert fetched.status_code == 200
    assert fetched.json()["pagination"] == {
        "page": 1,
        "limit": 1,
        "total": 1,
        "total_pages": 1,
    }
    revision = fetched.json()["drafts"][0]["revision"]
    fetched_draft = client.get(
        f"/api/v1/reconstructions/{run_id}/drafts/{draft['id']}",
        headers={"x-reconstruction-context": token},
    )
    assert fetched_draft.status_code == 200
    assert fetched_draft.json()["id"] == draft["id"]

    edited = client.patch(
        f"/api/v1/reconstructions/{run_id}/drafts/{draft['id']}",
        headers={"x-reconstruction-context": token},
        json={
            "expected_revision": revision,
            "operations": [
                {
                    "op": "set_field",
                    "path": "header.payment_method",
                    "value": "Chưa thanh toán",
                }
            ],
        },
    )
    assert edited.status_code == 200, edited.text
    assert edited.json()["revision"] == revision + 1

    stale = client.patch(
        f"/api/v1/reconstructions/{run_id}/drafts/{draft['id']}",
        headers={"x-reconstruction-context": token},
        json={
            "expected_revision": revision,
            "operations": [{"op": "set_type", "value": "purchase_goods"}],
        },
    )
    assert stale.status_code == 409

    validation = client.post(
        f"/api/v1/reconstructions/{run_id}/validate",
        headers={"x-reconstruction-context": token},
        json={},
    )
    assert validation.status_code == 200, validation.text
    assert validation.json()["summary"]["blocker"] == 0

    approved = client.post(
        f"/api/v1/reconstructions/{run_id}/approve",
        headers={"x-reconstruction-context": token},
        json={"acknowledge_warnings": True},
    )
    assert approved.status_code == 200, approved.text

    exported = client.post(
        f"/api/v1/reconstructions/{run_id}/export",
        headers={"x-reconstruction-context": token, "idempotency-key": "export-1"},
        json={"acknowledge_warnings": True},
    )
    assert exported.status_code == 200, exported.text
    assert exported.headers["content-type"].startswith("application/vnd.ms-excel")
    workbook = xlrd.open_workbook(file_contents=exported.content, formatting_info=True)
    sheet = workbook.sheet_by_index(0)
    headers = sheet.row_values(7)
    assert sheet.cell_value(8, headers.index("Số hóa đơn")) == "000123"
    assert sheet.cell_value(8, headers.index("Mã hàng (*)")) == "HH01"
    assert sheet.cell_value(9, headers.index("Mã hàng (*)")) == "HH02"

    repeated = client.post(
        f"/api/v1/reconstructions/{run_id}/export",
        headers={"x-reconstruction-context": token, "idempotency-key": "export-2"},
        json={"acknowledge_warnings": True},
    )
    assert repeated.status_code == 200, repeated.text
    repeated_book = xlrd.open_workbook(
        file_contents=repeated.content,
        formatting_info=True,
    )
    assert repeated_book.sheet_by_index(0).cell_value(
        8,
        headers.index("Số hóa đơn"),
    ) == "000123"


def test_reconstruction_split_and_merge_preserve_source_rows(tmp_path, monkeypatch):
    monkeypatch.setenv("VOUCHER_RECONSTRUCTION_ENABLED", "true")
    monkeypatch.setenv("CONVERSION_CONTEXT_SECRET", "reconstruction-test-secret")
    monkeypatch.setenv("RECONSTRUCTION_STORE_PROVIDER", "filesystem")
    monkeypatch.setenv("RECONSTRUCTION_STORE_DIR", str(tmp_path / "store"))
    monkeypatch.setenv("RECONSTRUCTION_NOTIFY_NODE", "false")
    notifications = []
    monkeypatch.setattr(
        "app.reconstruction_workflow._notify_node",
        lambda *_args, **kwargs: notifications.append(kwargs),
    )
    run_id = "run-api-2"
    token = _token(run_id)
    raw = _write_purchase(tmp_path / "purchase.xlsx")
    with raw.open("rb") as handle:
        analyze = client.post(
            "/api/v1/reconstructions/analyze",
            data={"context_token": token, "mode": "purchase"},
            files={"file": (raw.name, handle, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    draft = analyze.json()["drafts"][0]
    second_row = draft["lines"][1]["source_rows"][0]

    split = client.post(
        f"/api/v1/reconstructions/{run_id}/split",
        headers={"x-reconstruction-context": token},
        json={
            "draft_id": draft["id"],
            "expected_revision": draft["revision"],
            "source_rows": [second_row],
        },
    )
    assert split.status_code == 200, split.text
    assert len(split.json()["drafts"]) == 2
    assert sorted(len(item["source_rows"]) for item in split.json()["drafts"]) == [1, 1]

    draft_ids = [item["id"] for item in split.json()["drafts"]]
    revisions = {item["id"]: item["revision"] for item in split.json()["drafts"]}
    stale = client.post(
        f"/api/v1/reconstructions/{run_id}/merge",
        headers={"x-reconstruction-context": token},
        json={
            "draft_ids": draft_ids,
            "expected_revisions": {**revisions, draft_ids[0]: 0},
        },
    )
    assert stale.status_code == 409
    merge = client.post(
        f"/api/v1/reconstructions/{run_id}/merge",
        headers={"x-reconstruction-context": token},
        json={"draft_ids": draft_ids, "expected_revisions": revisions},
    )
    assert merge.status_code == 200, merge.text
    assert len(merge.json()["drafts"]) == 1
    assert len(merge.json()["drafts"][0]["source_rows"]) == 2
    decision_notifications = [item for item in notifications if item.get("decisions")]
    assert [
        item["decisions"][0]["operationType"] for item in decision_notifications
    ] == [
        "split",
        "merge",
    ]
    assert "ai" in merge.json()


def test_mixed_purchase_upload_exports_one_file_per_template(tmp_path, monkeypatch):
    monkeypatch.setenv("VOUCHER_RECONSTRUCTION_ENABLED", "true")
    monkeypatch.setenv("CONVERSION_CONTEXT_SECRET", "reconstruction-test-secret")
    monkeypatch.setenv("RECONSTRUCTION_STORE_PROVIDER", "filesystem")
    monkeypatch.setenv("RECONSTRUCTION_STORE_DIR", str(tmp_path / "store"))
    monkeypatch.setenv("RECONSTRUCTION_NOTIFY_NODE", "false")
    run_id = "run-api-zip"
    token = _token(run_id)
    raw = _write_purchase(tmp_path / "mixed.xlsx")
    workbook = openpyxl.load_workbook(raw)
    sheet = workbook.active
    service_row = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    values = list(service_row)
    values[0] = "000124"
    values[1] = "1C26TBB"
    values[2] = "02/07/2026"
    values[3] = "02/07/2026"
    values[4] = "0312222222"
    values[5] = "Nhà cung cấp B"
    values[6] = "Dịch vụ"
    values[7] = "DV01"
    values[8] = "Dịch vụ tư vấn"
    for column, value in enumerate(values, 1):
        sheet.cell(3, column, value)
    workbook.save(raw)

    with raw.open("rb") as handle:
        analyze = client.post(
            "/api/v1/reconstructions/analyze",
            data={"context_token": token, "mode": "purchase"},
            files={"file": (raw.name, handle, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert analyze.status_code == 200
    assert analyze.json()["summary"]["draft_count"] == 2
    approved = client.post(
        f"/api/v1/reconstructions/{run_id}/approve",
        headers={"x-reconstruction-context": token},
        json={"acknowledge_warnings": True},
    )
    assert approved.status_code == 200, approved.text
    exported = client.post(
        f"/api/v1/reconstructions/{run_id}/export",
        headers={"x-reconstruction-context": token, "idempotency-key": "zip-1"},
        json={"acknowledge_warnings": True},
    )
    assert exported.status_code == 200, exported.text
    assert exported.headers["content-type"].startswith("application/zip")
    with zipfile.ZipFile(io.BytesIO(exported.content)) as archive:
        names = archive.namelist()
        assert "manifest.json" in names
        assert len([name for name in names if name.endswith(".xls")]) == 2
        manifest = json.loads(archive.read("manifest.json"))
        assert manifest["acknowledged_warnings"] is True
        assert manifest["summary"]["blocker"] == 0
        assert {item["template_id"] for item in manifest["files"]} == {
            "misa_purchase_domestic",
            "purchase_service",
        }
        assert all("reconciliation" in item for item in manifest["files"])


def test_shadow_mode_allows_review_but_blocks_export(tmp_path, monkeypatch):
    monkeypatch.setenv("VOUCHER_RECONSTRUCTION_ENABLED", "true")
    monkeypatch.setenv("CONVERSION_CONTEXT_SECRET", "reconstruction-test-secret")
    monkeypatch.setenv("RECONSTRUCTION_STORE_PROVIDER", "filesystem")
    monkeypatch.setenv("RECONSTRUCTION_STORE_DIR", str(tmp_path / "store"))
    monkeypatch.setenv("RECONSTRUCTION_NOTIFY_NODE", "false")
    monkeypatch.setenv("RECONSTRUCTION_SHADOW_MODE", "true")
    run_id = "run-shadow-mode"
    token = _token(run_id)
    raw = _write_purchase(tmp_path / "shadow.xlsx")

    with raw.open("rb") as handle:
        analyzed = client.post(
            "/api/v1/reconstructions/analyze",
            data={"context_token": token, "mode": "purchase"},
            files={
                "file": (
                    raw.name,
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert analyzed.status_code == 200, analyzed.text

    approved = client.post(
        f"/api/v1/reconstructions/{run_id}/approve",
        headers={"x-reconstruction-context": token},
        json={"acknowledge_warnings": True},
    )
    assert approved.status_code == 200, approved.text

    exported = client.post(
        f"/api/v1/reconstructions/{run_id}/export",
        headers={"x-reconstruction-context": token},
        json={"acknowledge_warnings": True},
    )
    assert exported.status_code == 400
    assert "Shadow mode" in exported.json()["detail"]


def test_reconstruction_context_cannot_read_another_run(tmp_path, monkeypatch):
    monkeypatch.setenv("VOUCHER_RECONSTRUCTION_ENABLED", "true")
    monkeypatch.setenv("CONVERSION_CONTEXT_SECRET", "reconstruction-test-secret")
    monkeypatch.setenv("RECONSTRUCTION_STORE_PROVIDER", "filesystem")
    monkeypatch.setenv("RECONSTRUCTION_STORE_DIR", str(tmp_path / "store"))
    monkeypatch.setenv("RECONSTRUCTION_NOTIFY_NODE", "false")
    token = _token("run-owner")
    raw = _write_purchase(tmp_path / "purchase.xlsx")
    with raw.open("rb") as handle:
        client.post(
            "/api/v1/reconstructions/analyze",
            data={"context_token": token, "mode": "purchase"},
            files={"file": (raw.name, handle, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    response = client.get(
        "/api/v1/reconstructions/run-owner",
        headers={"x-reconstruction-context": _token("different-run")},
    )
    assert response.status_code == 409


def test_reconstruction_validation_rejects_stale_workspace_context(
    tmp_path,
    monkeypatch,
):
    monkeypatch.setenv("VOUCHER_RECONSTRUCTION_ENABLED", "true")
    monkeypatch.setenv("CONVERSION_CONTEXT_SECRET", "reconstruction-test-secret")
    monkeypatch.setenv("RECONSTRUCTION_STORE_PROVIDER", "filesystem")
    monkeypatch.setenv("RECONSTRUCTION_STORE_DIR", str(tmp_path / "store"))
    monkeypatch.setenv("RECONSTRUCTION_NOTIFY_NODE", "false")
    monkeypatch.setattr(
        "app.reconstruction_workflow.find_reconstruction_profile",
        lambda *_args, **_kwargs: None,
    )
    monkeypatch.setattr(
        "app.reconstruction_workflow.fetch_master_data_context",
        lambda _token: {
            "workspace": {"id": "workspace-1"},
            "snapshotSetHash": "snapshot-hash",
            "catalogs": {},
        },
    )
    run_id = "stale-workspace-run"
    token = _token(run_id, workspace_id="workspace-1")
    raw = _write_purchase(tmp_path / "stale-workspace.xlsx")
    with raw.open("rb") as handle:
        analyze = client.post(
            "/api/v1/reconstructions/analyze",
            data={"context_token": token, "mode": "purchase"},
            files={
                "file": (
                    raw.name,
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert analyze.status_code == 200, analyze.text

    def stale_context(_token):
        raise ConversionContextError(
            "Danh mục hoặc alias MISA đã thay đổi",
            status_code=409,
        )

    monkeypatch.setattr(
        "app.reconstruction_workflow.fetch_master_data_context",
        stale_context,
    )
    validation = client.post(
        f"/api/v1/reconstructions/{run_id}/validate",
        headers={"x-reconstruction-context": token},
        json={},
    )
    assert validation.status_code == 409


def test_reconstruction_approve_rejects_stale_profile(tmp_path, monkeypatch):
    monkeypatch.setenv("VOUCHER_RECONSTRUCTION_ENABLED", "true")
    monkeypatch.setenv("CONVERSION_CONTEXT_SECRET", "reconstruction-test-secret")
    monkeypatch.setenv("RECONSTRUCTION_STORE_PROVIDER", "filesystem")
    monkeypatch.setenv("RECONSTRUCTION_STORE_DIR", str(tmp_path / "store"))
    monkeypatch.setenv("RECONSTRUCTION_NOTIFY_NODE", "false")
    monkeypatch.setattr(
        "app.reconstruction_workflow.fetch_master_data_context",
        lambda _token: {
            "workspace": {"id": "workspace-1"},
            "snapshotSetHash": "snapshot-hash",
            "catalogs": {},
        },
    )
    monkeypatch.setattr(
        "app.reconstruction_workflow.find_reconstruction_profile",
        lambda *_args, **_kwargs: {
            "id": "profile-1",
            "version": 1,
            "directionScope": "purchase",
            "fieldRoles": {},
            "groupingKeys": [],
            "fillDownFields": [],
            "templateRouting": {},
        },
    )
    run_id = "stale-profile-run"
    token = _token(run_id, workspace_id="workspace-1")
    raw = _write_purchase(tmp_path / "stale-profile.xlsx")
    with raw.open("rb") as handle:
        analyze = client.post(
            "/api/v1/reconstructions/analyze",
            data={"context_token": token, "mode": "purchase"},
            files={
                "file": (
                    raw.name,
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert analyze.status_code == 200, analyze.text
    monkeypatch.setattr(
        "app.reconstruction_workflow.assert_reconstruction_profile_current",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            ReconstructionProfileClientError("Profile đã bị thay thế")
        ),
    )

    approved = client.post(
        f"/api/v1/reconstructions/{run_id}/approve",
        headers={"x-reconstruction-context": token},
        json={"acknowledge_warnings": True},
    )
    assert approved.status_code == 409
