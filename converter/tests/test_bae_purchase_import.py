from pathlib import Path

import openpyxl
import xlrd

from app.converter import convert_file, validate_file
from app.excel_io import read_input_table, read_template
from app.conversion_types import get_conversion_type


SMART_HEADERS = [
    "NIBOT_GHICHU",
    "LCTG",
    "SR_HD",
    "SOCT",
    "NGAY_KY",
    "NGAYCT",
    "SO_HD",
    "NGAY_HD",
    "DIENGIAI",
    "HTTT",
    "TKNO",
    "MADTPNNO",
    "TKCO",
    "MADTPNCO",
    "MADMNO",
    "MADMCO",
    "TENDM",
    "MATHANG",
    "Column1",
    "Phân loại",
    "DONVI",
    "LUONG",
    "DGUSD",
    "TTUSD",
    "TYGIA",
    "DGVND",
    "TTVND",
    "PT_CK",
    "CHIETKHAU",
    "HDVAT",
    "TKTHUE",
    "TS_GTGT",
    "THUEUSD",
    "THUEVND",
    "TTUSD_TT",
    "TTVND_TT",
    "MAKH",
    "TENKH",
    "KHACHHANG",
    "DIACHI_NGD",
    "MS_DN",
    "DIACHI",
    "TK_XUATKHO",
    "ID_NGHIEPVU",
    "GHICHU",
    "GUID",
]


def _write_bae_purchase_workbook(path: Path) -> Path:
    workbook = openpyxl.Workbook()
    summary = workbook.active
    summary.title = "HoaDon_TongQuat"
    summary.append(["HÓA ĐƠN MUA VÀO"])
    summary.append(["MST: 0317262773"])
    summary.append(["Tên DN: CÔNG TY TNHH BAE FOODS VIỆT NAM"])
    summary.append(["Loại HĐ", "MST người bán", "Người bán", "Ngày", "Ký hiệu", "Số"])
    summary.append(["V", "0312583082", "NGUYỄN THỊ TƯƠI", "2026-04-07", "2C26TYY", "27"])

    detail = workbook.create_sheet("Smart_KTSC_OK")
    detail.append(SMART_HEADERS)
    detail.append(
        [
            "HĐ cùng ngày, cùng ĐTPN và có tổng giá trị > 5,000,000 đồng",
            "",
            "2C26TYY",
            "27",
            "2026-04-07 11:22:33",
            "2026-04-07 00:00:00",
            27,
            "2026-04-07 00:00:00",
            "",
            "Tiền mặt/Chuyển khoản",
            "",
            "",
            "331",
            "0312583082",
            "",
            "",
            "Chà bông gà cay",
            "Chà bông gà cay",
            0,
            "Hàng hóa",
            "Kg",
            295,
            0,
            0,
            0,
            168000,
            49560000,
            0,
            0,
            "V",
            "1331",
            "",
            0,
            0,
            0,
            49560000,
            "0312583082",
            "NGUYỄN THỊ TƯƠI",
            "NGUYỄN THỊ TƯƠI",
            "C1 Phạm Hùng",
            "0312583082",
            "C1 Phạm Hùng",
            "",
            "TIENHANG",
            "",
            "goods-guid",
        ]
    )
    detail.append(
        [
            "HĐ có giá trị > 5,000,000 đồng",
            "",
            "1C26TQP",
            "148",
            "2026-04-03 10:58:44",
            "2026-04-03 00:00:00",
            148,
            "2026-04-03 00:00:00",
            "",
            "Chuyển khoản",
            "",
            "",
            "331",
            "0316292260",
            "",
            "",
            "Phí dịch vụ tư vấn tháng 04/2026",
            "Phí dịch vụ tư vấn tháng 04/2026",
            "",
            "Dịch vụ",
            "Tháng",
            1,
            0,
            0,
            0,
            8000000,
            8000000,
            0,
            0,
            "V",
            "1331",
            "8",
            0,
            640000,
            0,
            8640000,
            "0316292260",
            "CÔNG TY TNHH DỊCH VỤ TƯ VẤN NAAG",
            "CÔNG TY TNHH DỊCH VỤ TƯ VẤN NAAG",
            "1115/7C Huỳnh Tấn Phát",
            "0316292260",
            "1115/7C Huỳnh Tấn Phát",
            "",
            "TIENHANG",
            "",
            "service-guid",
        ]
    )
    workbook.save(path)
    return path


def _header_map(sheet: xlrd.sheet.Sheet) -> dict[str, int]:
    header_row = 7
    return {
        str(value).strip(): index
        for index, value in enumerate(sheet.row_values(header_row))
        if str(value).strip()
    }


def test_bae_smart_purchase_workbook_validates_and_converts_to_full_misa_template(tmp_path):
    input_path = _write_bae_purchase_workbook(tmp_path / "bae_purchase.xlsx")
    output_path = tmp_path / "misa_purchase.xls"

    table = read_input_table(input_path)
    assert table.sheet_name == "Smart_KTSC_OK"
    assert len(table.rows) == 2

    report = validate_file(input_path, "purchase_goods")
    assert report.ok is True
    assert report.detected_columns["purchase_receipt"] == "SOCT"
    assert report.detected_columns["supplier_code"] == "MAKH"
    assert report.detected_columns["item_code"] == "MATHANG"
    assert report.detected_columns["item_type"] == "Phân loại"

    template = read_template(get_conversion_type("purchase_goods").template_path)
    assert len(template.headers) == 58
    assert "Mẫu số HĐ" in template.headers
    assert "Ký hiệu HĐ" in template.headers
    assert "CP không hợp lý" in template.headers

    converted = convert_file(input_path, "purchase_goods", output_path)
    assert converted.ok is True

    sheet = xlrd.open_workbook(str(output_path)).sheet_by_index(0)
    headers = _header_map(sheet)
    goods_row = 8
    service_row = 9

    assert sheet.cell_value(goods_row, headers["Hình thức mua hàng"]) == "Mua hàng trong nước nhập kho"
    assert sheet.cell_value(goods_row, headers["Số phiếu nhập (*)"]) == "27"
    assert sheet.cell_value(goods_row, headers["Số hóa đơn"]) == "27"
    assert sheet.cell_value(goods_row, headers["Ký hiệu HĐ"]) == "2C26TYY"
    assert sheet.cell_value(goods_row, headers["Mã nhà cung cấp"]) == "0312583082"
    assert sheet.cell_value(goods_row, headers["Mã số thuế"]) == "0312583082"
    assert sheet.cell_value(goods_row, headers["Mã hàng (*)"]) == "Chà bông gà cay"
    assert sheet.cell_value(goods_row, headers["ĐVT"]) == "Kg"
    assert sheet.cell_value(goods_row, headers["Số lượng"]) == 295
    assert sheet.cell_value(goods_row, headers["Đơn giá"]) == 168000
    assert sheet.cell_value(goods_row, headers["Thành tiền"]) == 49560000

    assert sheet.cell_value(service_row, headers["Hình thức mua hàng"]) == "Mua hàng trong nước không qua kho"
    assert sheet.cell_value(service_row, headers["Số phiếu nhập (*)"]) == ""
    assert sheet.cell_value(service_row, headers["Số chứng từ ghi nợ/Số chứng từ thanh toán"]) == "148"
    assert sheet.cell_value(service_row, headers["Số hóa đơn"]) == "148"
    assert sheet.cell_value(service_row, headers["Ký hiệu HĐ"]) == "1C26TQP"
    assert sheet.cell_value(service_row, headers["% thuế GTGT"]) == "8"
    assert sheet.cell_value(service_row, headers["Tiền thuế GTGT"]) == 640000
    assert sheet.cell_value(service_row, headers["TK thuế GTGT"]) == "1331"


def test_common_purchase_total_column_does_not_override_misa_line_amount(tmp_path):
    input_path = tmp_path / "common_purchase.xlsx"
    output_path = tmp_path / "common_purchase.xls"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Số PN",
            "Ngày nhập",
            "Mã NCC",
            "Tên NCC",
            "Mã SKU mua",
            "SL nhập",
            "Giá mua",
            "Tổng dòng mua",
            "Giảm giá mua",
        ]
    )
    sheet.append(["PNX00001", "2025-12-01", "NCC001", "Nhà cung cấp 001", "SKU-M00001", 1, 8000, 7760, 240])
    workbook.save(input_path)

    report = convert_file(input_path, "purchase_goods", output_path, {"allow_calculation_warnings": True})
    assert report.ok is True

    output_sheet = xlrd.open_workbook(str(output_path)).sheet_by_index(0)
    headers = _header_map(output_sheet)
    assert output_sheet.cell_value(8, headers["Thành tiền"]) == 8000
    assert output_sheet.cell_value(8, headers["Tiền chiết khấu"]) == 240
