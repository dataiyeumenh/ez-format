from __future__ import annotations

import openpyxl
import xlwt

from app.document_structure import (
    enforce_workbook_limits,
    inspect_workbook_structure,
    validate_excel_magic,
)


def test_xlsx_structure_reports_hidden_formula_and_merged_cells(tmp_path):
    path = tmp_path / "structure.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "Dữ liệu"
    sheet.append(["Số hóa đơn", "Số lượng", "Thành tiền"])
    sheet.append(["HD01", 2, "=B3*1000"])
    sheet.row_dimensions[3].hidden = True
    sheet.column_dimensions["C"].hidden = True
    workbook.save(path)

    result = inspect_workbook_structure(path)

    assert result["sheet_count"] == 1
    assert result["formula_cell_count"] == 1
    assert result["hidden_row_count"] == 1
    assert result["hidden_column_count"] == 1
    assert result["merged_range_count"] == 1
    assert {item["code"] for item in result["warnings"]} == {
        "formula_cells_detected",
        "hidden_rows_or_columns_detected",
        "merged_cells_detected",
    }


def test_xls_structure_warns_when_formula_detection_is_unavailable_and_normalizes_hidden_cells(
    tmp_path,
):
    path = tmp_path / "structure.xls"
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Data")
    sheet.write(0, 0, "Mã hóa đơn")
    sheet.write(1, 0, "HD01")
    sheet.write(1, 1, xlwt.Formula("1+1"))
    sheet.row(1).hidden = True
    sheet.col(0).hidden = True
    workbook.save(str(path))

    result = inspect_workbook_structure(path)

    assert result["format"] == "xls"
    assert result["formula_cell_count"] == 0
    assert result["formula_detection"] == "unavailable"
    assert result["sheets"] == [
        {
            "name": "Data",
            "max_row": 2,
            "max_column": 2,
            "hidden_rows": [2],
            "hidden_columns": ["A"],
            "formula_cells": [],
            "merged_ranges": [],
        }
    ]
    assert {item["code"] for item in result["warnings"]} == {
        "formula_detection_unavailable",
        "hidden_rows_or_columns_detected",
    }


def test_workbook_limits_reject_excessive_rows(monkeypatch):
    monkeypatch.setenv("RECONSTRUCTION_MAX_ROWS", "10")
    try:
        enforce_workbook_limits(content_size=10, row_count=11, column_count=2)
    except ValueError as exc:
        assert "10 dòng" in str(exc)
    else:
        raise AssertionError("Expected row limit error")


def test_excel_magic_rejects_renamed_non_excel_file():
    try:
        validate_excel_magic("fake.xlsx", b"not-an-excel-file")
    except ValueError as exc:
        assert "OpenXML" in str(exc)
    else:
        raise AssertionError("Expected invalid xlsx magic error")
