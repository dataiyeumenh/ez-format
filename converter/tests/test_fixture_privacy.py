import hashlib
from io import BytesIO
import json
from datetime import datetime
from pathlib import Path

import openpyxl
import olefile
import pytest
import xlrd

from app import misa_biff, misa_templates
from app.excel_io import find_header_row
from scripts.generate_synthetic_sales_fixtures import generate_fixtures


ROOT = Path(__file__).resolve().parents[1]
REPOSITORY_ROOT = ROOT.parent
FIXTURES = ROOT / "fixtures"
SAMPLES = FIXTURES / "samples"
FIXTURE_MANIFEST = ROOT / "config" / "converter-fixture-manifest.json"
EXPECTED_WORKBOOKS = {
    "samples/golden_sales_import.xls",
    "samples/raw_sales_sample.xlsx",
    "templates/bsn_purchase.xls",
    "templates/bsn_sales.xls",
    "templates/mua_hang_trong_nuoc_full.xls",
    "templates/purchase_goods.xls",
    "templates/purchase_service.xls",
    "templates/sales_goods.xls",
    "templates/sales_service.xls",
}
SAFE_GOLDEN_VALUES = {
    "131",
    "1561",
    "33311",
    "5111",
    "632",
    "Bán hàng hóa trong nước",
    "Chưa thu tiền",
    "Có",
    "Không",
    "KHO_BSN",
    "Hộp",
    "không",
    "Đã lập",
}


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _accepted_trust_levels(path: Path) -> set[str]:
    prefix = "MISA_TEMPLATE_ACCEPTED_TRUST_LEVELS="
    line = next(
        line for line in path.read_text(encoding="utf-8").splitlines() if line.startswith(prefix)
    )
    return {value.strip() for value in line[len(prefix) :].split(",") if value.strip()}


@pytest.fixture
def _allow_uncertified_test_exports(monkeypatch):
    monkeypatch.setattr(misa_templates, "_enforce_export_capability", lambda *args: None)


def test_converter_fixture_manifest_pins_only_deterministic_synthetic_samples():
    manifest = json.loads(FIXTURE_MANIFEST.read_text(encoding="utf-8"))

    assert manifest["schema_version"] == 2
    assert manifest["fixture_version"] == "2026-08-01.1"
    assert set(manifest["fixtures"]) == {
        "golden_sales_import.xls",
        "raw_sales_sample.xlsx",
    }
    for filename, entry in manifest["fixtures"].items():
        path = SAMPLES / filename
        assert entry["source_kind"] == "deterministic_synthetic"
        assert entry["fixture_kind"] == "synthetic"
        assert entry["privacy_classification"] == "synthetic_no_customer_data"
        assert entry["contains_customer_data"] is False
        assert entry["generator"] == "scripts/generate_synthetic_sales_fixtures.py"
        assert entry["reviewer"] == "fixture-privacy-reviewer"
        assert entry["approval_status"] == "approved"
        assert entry["approved_at_utc"] == "2026-08-01T00:00:00+00:00"
        assert entry["synthetic_fixture_id"].startswith("synthetic-sales-")
        assert entry["path"] == f"converter/fixtures/samples/{filename}"
        assert entry["sha256"] == _sha256(path)


def test_fixture_generator_reproduces_committed_hashes(
    tmp_path,
    _allow_uncertified_test_exports,
):
    expected = json.loads(FIXTURE_MANIFEST.read_text(encoding="utf-8"))
    generated = generate_fixtures(
        tmp_path / "samples",
        tmp_path / "converter-fixture-manifest.json",
    )

    assert generated == expected


def test_all_fixture_workbooks_have_synthetic_cells_and_clean_metadata():
    workbooks = {
        path.relative_to(FIXTURES).as_posix()
        for path in FIXTURES.rglob("*")
        if path.suffix.lower() in {".xls", ".xlsx"}
    }
    assert workbooks == EXPECTED_WORKBOOKS

    raw_book = openpyxl.load_workbook(SAMPLES / "raw_sales_sample.xlsx")
    assert raw_book.sheetnames == ["SYN-SALES"]
    assert raw_book.properties.creator == "SYN-FIXTURE-GENERATOR"
    assert raw_book.properties.lastModifiedBy == "SYN-FIXTURE-GENERATOR"
    assert raw_book.properties.created == datetime(2000, 1, 1)
    assert raw_book.properties.modified == datetime(2000, 1, 1)
    for value in vars(raw_book.properties).values():
        if isinstance(value, str) and value.strip():
            assert value.startswith("SYN-")
    assert not list(raw_book.defined_names.values())
    if hasattr(raw_book, "custom_doc_props"):
        assert not list(raw_book.custom_doc_props)
    raw_sheet = raw_book.active
    assert (raw_sheet.max_row, raw_sheet.max_column) == (1931, 66)
    for row in raw_sheet.iter_rows(min_row=2):
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.strip():
                assert value.startswith("SYN-"), cell.coordinate
            assert cell.comment is None, cell.coordinate
            assert cell.hyperlink is None, cell.coordinate

    golden_bytes = (SAMPLES / "golden_sales_import.xls").read_bytes()
    metadata_scanner = getattr(misa_biff, "scan_ole_metadata", None)
    assert metadata_scanner is not None
    assert metadata_scanner(golden_bytes).clean
    golden_book = xlrd.open_workbook(file_contents=golden_bytes, formatting_info=True)
    golden_sheet = golden_book.sheet_by_index(0)
    header_row = find_header_row(golden_sheet)
    for row_index in range(header_row + 1, golden_sheet.nrows):
        for value in golden_sheet.row_values(row_index):
            if not isinstance(value, str) or not value.strip():
                continue
            assert "SYN-" in value or value in SAFE_GOLDEN_VALUES, (
                row_index,
                value,
            )

    for template in misa_templates.list_misa_templates():
        assert metadata_scanner(template.workbook.file_contents).clean, template.id


def test_root_and_converter_env_accept_manifest_trust_vocabulary():
    manifest = json.loads(misa_templates.DEFAULT_MANIFEST_PATH.read_text(encoding="utf-8"))
    manifest_levels = {
        entry["provenance"]["trust_level"] for entry in manifest["templates"].values()
    }

    assert manifest_levels == {"partner_sample_derived"}
    assert _accepted_trust_levels(REPOSITORY_ROOT / ".env.example") == manifest_levels
    assert _accepted_trust_levels(ROOT / ".env.example") == manifest_levels


def test_ole_scanner_fails_closed_on_future_unknown_stream(monkeypatch):
    original_listdir = olefile.OleFileIO.listdir

    def listdir_with_unknown(self, *args, **kwargs):
        return original_listdir(self, *args, **kwargs) + [["FutureMetadata"]]

    monkeypatch.setattr(olefile.OleFileIO, "listdir", listdir_with_unknown)
    contents = (FIXTURES / "templates" / "bsn_sales.xls").read_bytes()

    scan = misa_biff.scan_ole_metadata(contents)

    assert scan.unknown_ole_stream_count == 1
    assert not scan.clean


def test_ole_scanner_rejects_ambiguous_duplicate_workbook_stream(monkeypatch):
    original_listdir = olefile.OleFileIO.listdir

    def listdir_with_second_workbook(self, *args, **kwargs):
        return original_listdir(self, *args, **kwargs) + [["Book"]]

    monkeypatch.setattr(olefile.OleFileIO, "listdir", listdir_with_second_workbook)
    contents = (FIXTURES / "templates" / "bsn_sales.xls").read_bytes()

    scan = misa_biff.scan_ole_metadata(contents)

    assert scan.unsafe_ole_stream_count == 1
    assert not scan.clean


def test_ole_scanner_rejects_non_allowlisted_binary_stream_content(monkeypatch):
    original_openstream = olefile.OleFileIO.openstream

    def openstream_with_customer_text(self, filename):
        stream_name = "/".join(filename) if isinstance(filename, (list, tuple)) else filename
        if stream_name == "\x01CompObj":
            return BytesIO(b"Customer User Name")
        return original_openstream(self, filename)

    monkeypatch.setattr(olefile.OleFileIO, "openstream", openstream_with_customer_text)
    contents = (FIXTURES / "templates" / "purchase_goods.xls").read_bytes()

    scan = misa_biff.scan_ole_metadata(contents)

    assert scan.unsafe_ole_stream_count == 1
    assert not scan.clean


def test_ole_scanner_includes_custom_document_properties(monkeypatch):
    monkeypatch.setattr(
        olefile.OleFileIO,
        "get_userdefined_properties",
        lambda *args, **kwargs: [{"property_name": "Customer", "value": "User Name"}],
    )
    contents = (FIXTURES / "templates" / "sales_goods.xls").read_bytes()

    scan = misa_biff.scan_ole_metadata(contents)

    assert scan.ole_property_value_count >= 2
    assert not scan.clean


def test_ole_scanner_rejects_property_stream_padding_markers(monkeypatch):
    original_openstream = olefile.OleFileIO.openstream
    marker = b"CUSTOMER-PADDING-MARKER"

    def openstream_with_padding_marker(self, filename):
        stream_name = "/".join(filename) if isinstance(filename, (list, tuple)) else filename
        stream = original_openstream(self, filename)
        if stream_name != "\x05SummaryInformation":
            return stream
        contents = stream.read()
        return BytesIO(contents[: -len(marker)] + marker)

    monkeypatch.setattr(olefile.OleFileIO, "openstream", openstream_with_padding_marker)
    contents = (FIXTURES / "templates" / "sales_goods.xls").read_bytes()

    scan = misa_biff.scan_ole_metadata(contents)

    assert scan.property_stream_residual_count == 1
    assert not scan.clean


def test_ole_scanner_rejects_property_stream_structural_markers(monkeypatch):
    original_openstream = olefile.OleFileIO.openstream

    def openstream_with_structural_marker(self, filename):
        stream_name = "/".join(filename) if isinstance(filename, (list, tuple)) else filename
        stream = original_openstream(self, filename)
        if stream_name != "\x05SummaryInformation":
            return stream
        contents = bytearray(stream.read())
        contents[28:44] = b"CUSTOMER-MARKER!"
        return BytesIO(bytes(contents))

    monkeypatch.setattr(olefile.OleFileIO, "openstream", openstream_with_structural_marker)
    contents = (FIXTURES / "templates" / "sales_goods.xls").read_bytes()

    scan = misa_biff.scan_ole_metadata(contents)

    assert scan.property_stream_residual_count == 1
    assert not scan.clean
