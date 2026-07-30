from __future__ import annotations

from io import BytesIO
import re
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
import xlwt


def build_import_result_xlsx(
    *,
    headers: list[object],
    rows: list[list[object]],
    hidden_rows: tuple[int, ...] = (),
    formula_cells: dict[str, str] | None = None,
) -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Import result"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    for row_number in hidden_rows:
        sheet.row_dimensions[row_number].hidden = True
    for coordinate, formula in (formula_cells or {}).items():
        sheet[coordinate] = formula
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_import_result_xls(*, headers: list[object], rows: list[list[object]]) -> bytes:
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Import result")
    for column, value in enumerate(headers):
        sheet.write(0, column, value)
    for row_number, row in enumerate(rows, start=1):
        for column, value in enumerate(row):
            sheet.write(row_number, column, value)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def add_xlsx_zip_payload(content: bytes, *, name: str, payload: bytes) -> bytes:
    output = BytesIO()
    with ZipFile(BytesIO(content)) as source, ZipFile(output, "w", ZIP_DEFLATED) as target:
        for item in source.infolist():
            target.writestr(item, source.read(item.filename))
        target.writestr(name, payload)
    return output.getvalue()


def add_xlsx_external_link(content: bytes) -> bytes:
    output = BytesIO()
    with ZipFile(BytesIO(content)) as source, ZipFile(output, "w", ZIP_DEFLATED) as target:
        for item in source.infolist():
            payload = source.read(item.filename)
            if item.filename == "xl/_rels/workbook.xml.rels":
                payload = payload.replace(
                    b"</Relationships>",
                    b'<Relationship Id="rIdExternal" '
                    b'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLink" '
                    b'Target="externalLinks/externalLink1.xml"/></Relationships>',
                )
            target.writestr(item, payload)
        target.writestr(
            "xl/externalLinks/externalLink1.xml",
            b'<externalLink xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            b"<externalBook/></externalLink>",
        )
        target.writestr(
            "xl/externalLinks/_rels/externalLink1.xml.rels",
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            b'<Relationship Id="rId1" '
            b'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLinkPath" '
            b'Target="https://example.invalid/source.xlsx" TargetMode="External"/>'
            b"</Relationships>",
        )
    return output.getvalue()


def build_import_result_xlsx_sheets(
    sheets: list[tuple[str, list[list[object]]]],
) -> bytes:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for title, rows in sheets:
        sheet = workbook.create_sheet(title)
        for row in rows:
            sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def set_xlsx_sheet_dimension(content: bytes, reference: str) -> bytes:
    output = BytesIO()
    with ZipFile(BytesIO(content)) as source, ZipFile(output, "w", ZIP_DEFLATED) as target:
        for item in source.infolist():
            payload = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                payload = re.sub(
                    rb'<dimension ref="[^"]+"\s*/>',
                    f'<dimension ref="{reference}"/>'.encode("ascii"),
                    payload,
                    count=1,
                )
            target.writestr(item, payload)
    return output.getvalue()
