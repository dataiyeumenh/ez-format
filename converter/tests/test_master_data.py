from pathlib import Path

from openpyxl import Workbook
import xlwt

from app.master_data import parse_master_data_file
from app.master_data_resolver import resolve_master_data


def _write_xlsx(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def test_parse_supplier_catalog_preserves_codes_and_tax_codes(tmp_path):
    path = tmp_path / "suppliers.xlsx"
    _write_xlsx(
        path,
        ["Mã nhà cung cấp", "Tên nhà cung cấp", "Mã số thuế"],
        [
            ["001-NCC", "Công ty TNHH Bảo An", "0317262773-001"],
            ["NCC002", "Công ty B", "0101234567"],
        ],
    )

    result = parse_master_data_file(path, "supplier")

    assert result["sheet_name"] == "Sheet"
    assert result["entries"][0]["code"] == "001-NCC"
    assert result["entries"][0]["normalizedCode"] == "001-NCC"
    assert result["entries"][0]["normalizedName"] == "cong ty tnhh bao an"
    assert result["entries"][0]["normalizedTaxCode"] == "0317262773-001"


def test_parse_catalog_rejects_unsupported_type(tmp_path):
    path = tmp_path / "data.xlsx"
    _write_xlsx(path, ["Mã"], [["A"]])

    try:
        parse_master_data_file(path, "unknown")
    except ValueError as exc:
        assert "không được hỗ trợ" in str(exc)
    else:
        raise AssertionError("Expected unsupported catalog type to fail")


def test_parse_catalog_reports_duplicate_codes(tmp_path):
    path = tmp_path / "items.xlsx"
    _write_xlsx(
        path,
        ["Mã hàng", "Tên hàng", "ĐVT"],
        [["HH001", "Hàng A", "Cái"], ["hh001", "Hàng A duplicate", "Cái"]],
    )

    result = parse_master_data_file(path, "item")

    assert len(result["entries"]) == 1
    assert result["warnings"] == ["Mã HH001 xuất hiện nhiều lần; chỉ giữ dòng đầu tiên."]


def test_parse_xlsx_preserves_leading_zero_from_number_format(tmp_path):
    path = tmp_path / "suppliers.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Mã nhà cung cấp", "Tên nhà cung cấp", "Mã số thuế"])
    sheet.append([12, "Nhà cung cấp 12", 317262773])
    sheet["A2"].number_format = "0000"
    sheet["C2"].number_format = "0000000000"
    workbook.save(path)

    result = parse_master_data_file(path, "supplier")

    assert result["entries"][0]["code"] == "0012"
    assert result["entries"][0]["taxCode"] == "0317262773"


def test_parse_xls_preserves_leading_zero_from_number_format(tmp_path):
    path = tmp_path / "suppliers.xls"
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Suppliers")
    for index, header in enumerate(["Mã nhà cung cấp", "Tên nhà cung cấp", "Mã số thuế"]):
        sheet.write(0, index, header)
    sheet.write(1, 0, 12, xlwt.easyxf(num_format_str="0000"))
    sheet.write(1, 1, "Nhà cung cấp 12")
    sheet.write(1, 2, 317262773, xlwt.easyxf(num_format_str="0000000000"))
    workbook.save(str(path))

    result = parse_master_data_file(path, "supplier")

    assert result["entries"][0]["code"] == "0012"
    assert result["entries"][0]["taxCode"] == "0317262773"


def _context() -> dict:
    return {
        "workspace": {"id": "workspace-1", "name": "BAE"},
        "snapshotSetHash": "hash-1",
        "catalogs": {
            "supplier": {
                "entries": [
                    {
                        "code": "NCC001",
                        "normalizedCode": "NCC001",
                        "name": "Công ty BAE Foods",
                        "normalizedName": "cong ty bae foods",
                        "taxCode": "0317262773",
                        "normalizedTaxCode": "0317262773",
                        "active": True,
                    }
                ],
                "aliases": [
                    {
                        "normalizedRawValue": "bae food",
                        "targetCode": "NCC001",
                        "normalizedTargetCode": "NCC001",
                    }
                ],
            },
            "account": {
                "entries": [
                    {
                        "code": "331",
                        "normalizedCode": "331",
                        "name": "Phải trả người bán",
                        "normalizedName": "phai tra nguoi ban",
                        "active": True,
                    }
                ],
                "aliases": [],
            },
        },
    }


def test_resolver_applies_confirmed_alias_and_groups_repeated_values():
    rows = [
        {"Mã nhà cung cấp": "BAE FOOD", "TK công nợ/TK tiền (*)": "331"},
        {"Mã nhà cung cấp": "BAE FOOD", "TK công nợ/TK tiền (*)": "331"},
    ]

    result = resolve_master_data(rows, _context())

    assert result.rows[0]["Mã nhà cung cấp"] == "NCC001"
    assert result.rows[1]["Mã nhà cung cấp"] == "NCC001"
    supplier = next(item for item in result.resolutions if item.catalog_type == "supplier")
    assert supplier.status == "verified"
    assert supplier.match_method == "confirmed_alias"
    assert supplier.affected_rows == 2


def test_resolver_prefers_alias_for_current_source_system():
    context = _context()
    context["catalogs"]["supplier"]["entries"].append(
        {
            "code": "NCC002",
            "normalizedCode": "NCC002",
            "name": "BAE source 2",
            "normalizedName": "bae source 2",
            "active": True,
        }
    )
    context["catalogs"]["supplier"]["aliases"] = [
        {
            "sourceSystem": "source-a",
            "normalizedRawValue": "bae food",
            "targetCode": "NCC001",
            "normalizedTargetCode": "NCC001",
        },
        {
            "sourceSystem": "source-b",
            "normalizedRawValue": "bae food",
            "targetCode": "NCC002",
            "normalizedTargetCode": "NCC002",
        },
    ]

    result = resolve_master_data(
        [{"Mã nhà cung cấp": "BAE FOOD"}],
        context,
        source_system="source-b",
    )

    assert result.rows[0]["Mã nhà cung cấp"] == "NCC002"


def test_resolver_never_fuzzy_matches_account_codes():
    rows = [{"TK công nợ/TK tiền (*)": "33l"}]

    result = resolve_master_data(rows, _context())

    resolution = result.resolutions[0]
    assert resolution.catalog_type == "account"
    assert resolution.status == "missing"
    assert resolution.candidates == []


def test_resolver_suggests_name_match_without_applying_it():
    rows = [{"Mã nhà cung cấp": "Công ty BAE Foods"}]

    result = resolve_master_data(rows, _context())

    resolution = result.resolutions[0]
    assert resolution.status == "suggested"
    assert resolution.candidates[0]["code"] == "NCC001"
    assert result.rows[0]["Mã nhà cung cấp"] == "Công ty BAE Foods"


def test_resolver_marks_missing_catalog_as_not_checked():
    rows = [{"Mã hàng (*)": "HH001"}]

    result = resolve_master_data(rows, _context())

    assert result.resolutions[0].status == "not_checked"
