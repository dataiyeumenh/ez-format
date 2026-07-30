from pathlib import Path
from io import BytesIO
from zipfile import ZipFile

import pytest
import app.student_anonymization as anonymization_module
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as WorksheetImage
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table
from openpyxl.styles import NamedStyle
from openpyxl.packaging.relationship import Relationship
from openpyxl.packaging.custom import StringProperty
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.workbook.external_link.external import (
    ExternalBook,
    ExternalLink,
    ExternalSheetNames,
)
import xlrd
import xlwt
from PIL import Image as PillowImage

from app.student_anonymization import (
    AnonymizationSession,
    AnonymizationExportError,
    AnonymizationUnsupportedLayerError,
    anonymize_workbook_bytes,
    scan_confidential_values,
)
from app.student_workflow import (
    _student_anonymization_secret,
    assert_student_anonymization_config,
)


ROOT = Path(__file__).resolve().parents[2]


def test_disabled_student_assistant_does_not_require_anonymization_secret(
    monkeypatch,
):
    monkeypatch.setenv("NODE_ENV", "production")
    monkeypatch.setenv("STUDENT_ASSISTANT_ENABLED", "false")
    monkeypatch.delenv("STUDENT_ANONYMIZATION_SECRET", raising=False)

    assert_student_anonymization_config()


@pytest.mark.parametrize(
    "secret,conversion_secret,service_token",
    [
        ("", "conversion-secret-value-1234567890", "service-token-value-123456789012"),
        ("short", "conversion-secret-value-1234567890", "service-token-value-123456789012"),
        (
            "same-secret-value-12345678901234567890",
            "same-secret-value-12345678901234567890",
            "service-token-value-123456789012",
        ),
        (
            "same-secret-value-12345678901234567890",
            "conversion-secret-value-1234567890",
            "same-secret-value-12345678901234567890",
        ),
    ],
)
def test_enabled_production_requires_a_distinct_32_character_anonymization_secret(
    monkeypatch,
    secret,
    conversion_secret,
    service_token,
):
    monkeypatch.setenv("NODE_ENV", "production")
    monkeypatch.setenv("STUDENT_ASSISTANT_ENABLED", "true")
    monkeypatch.setenv("STUDENT_ANONYMIZATION_SECRET", secret)
    monkeypatch.setenv("CONVERSION_CONTEXT_SECRET", conversion_secret)
    monkeypatch.setenv("CONVERTER_SERVICE_TOKEN", service_token)

    with pytest.raises(ValueError, match="STUDENT_ANONYMIZATION_SECRET"):
        assert_student_anonymization_config()


def test_enabled_production_accepts_a_distinct_anonymization_secret(monkeypatch):
    monkeypatch.setenv("NODE_ENV", "production")
    monkeypatch.setenv("STUDENT_ASSISTANT_ENABLED", "true")
    monkeypatch.setenv(
        "STUDENT_ANONYMIZATION_SECRET",
        "student-anonymization-secret-value-1234",
    )
    monkeypatch.setenv(
        "CONVERSION_CONTEXT_SECRET",
        "conversion-context-secret-value-12345",
    )
    monkeypatch.setenv(
        "CONVERTER_SERVICE_TOKEN",
        "converter-service-token-value-123456",
    )

    assert_student_anonymization_config()


def test_student_anonymization_never_falls_back_to_auth_secrets_in_production(
    monkeypatch,
):
    monkeypatch.setenv("NODE_ENV", "production")
    monkeypatch.delenv("STUDENT_ANONYMIZATION_SECRET", raising=False)
    monkeypatch.setenv("CONVERSION_CONTEXT_SECRET", "conversion-secret")
    monkeypatch.setenv("JWT_SECRET", "auth-secret")
    monkeypatch.setenv("STUDENT_ANONYMIZATION_ALLOW_SHARED_SECRET_FALLBACK", "true")

    with pytest.raises(ValueError, match="Student anonymization secret"):
        _student_anonymization_secret()


def test_student_anonymization_shared_fallback_is_explicitly_development_only(
    monkeypatch,
):
    monkeypatch.setenv("NODE_ENV", "development")
    monkeypatch.delenv("STUDENT_ANONYMIZATION_SECRET", raising=False)
    monkeypatch.setenv("CONVERSION_CONTEXT_SECRET", "conversion-secret")
    monkeypatch.setenv("STUDENT_ANONYMIZATION_ALLOW_SHARED_SECRET_FALLBACK", "true")

    assert _student_anonymization_secret() == "conversion-secret"
FEATURE_FLAGS = {
    "STUDENT_ASSISTANT_ENABLED",
    "STUDENT_FILE_EXPLAIN_ENABLED",
    "STUDENT_FILE_QA_ENABLED",
    "STUDENT_ACCOUNTING_MAP_ENABLED",
    "STUDENT_RECONCILIATION_ENABLED",
    "STUDENT_INTERNSHIP_ENABLED",
}


def test_anonymization_is_stable_within_the_same_session_and_category():
    first = AnonymizationSession("session-1", "secret")
    second = AnonymizationSession("session-1", "secret")

    replacement = first.replace("company", "Công ty TNHH Sao Mai")

    assert replacement == first.replace("company", "Công ty TNHH Sao Mai")
    assert replacement == second.replace("company", "Công ty TNHH Sao Mai")
    assert replacement != AnonymizationSession("session-2", "secret").replace(
        "company", "Công ty TNHH Sao Mai"
    )


def test_anonymization_categories_do_not_collide_for_the_same_source_value():
    session = AnonymizationSession("session-1", "secret")

    replacements = {
        category: session.replace(category, "0012345678")
        for category in (
            "company",
            "counterparty",
            "tax_code",
            "address",
            "email",
            "phone",
            "bank_account",
            "document_number",
        )
    }

    assert len(set(replacements.values())) == len(replacements)


def test_anonymization_preserves_blanks_and_numeric_identifiers_as_text():
    session = AnonymizationSession("session-1", "secret")

    assert session.replace("company", None) is None
    assert session.replace("company", "") == ""
    assert session.replace("company", "   ") == "   "
    tax_code = session.replace("tax_code", "0012345678")

    assert isinstance(tax_code, str)
    assert tax_code.startswith("TAX-00")


def test_anonymization_rejects_unknown_categories():
    session = AnonymizationSession("session-1", "secret")

    with pytest.raises(ValueError, match="category"):
        session.replace("password", "secret")


def test_confidential_scanner_reports_categories_without_returning_raw_values():
    matches = scan_confidential_values(
        {
            "summary": "Công ty TNHH Sao Mai",
            "rows": [{"tax_code": "0012345678"}, {"note": "safe"}],
        },
        {
            "company": ["Công ty TNHH Sao Mai"],
            "tax_code": ["0012345678"],
            "phone": ["0900000000"],
        },
    )

    assert matches == ("company", "tax_code")
    assert "Công ty TNHH Sao Mai" not in matches
    assert "0012345678" not in matches


def test_anonymize_xlsx_returns_a_scanned_copy_without_mutating_original_bytes():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Company"])
    worksheet.append(["Công ty TNHH Sao Mai"])
    original_stream = BytesIO()
    workbook.save(original_stream)
    original = original_stream.getvalue()

    exported = anonymize_workbook_bytes(
        filename="student.xlsx",
        content=original,
        session=AnonymizationSession("session-1", "secret"),
        confidential_values={"company": ["Công ty TNHH Sao Mai"]},
    )

    anonymized = load_workbook(BytesIO(exported.content), data_only=False)
    assert original == original_stream.getvalue()
    assert anonymized.active["A2"].value != "Công ty TNHH Sao Mai"
    assert exported.replaced_categories == ("company",)


def test_anonymize_xlsx_refuses_export_when_a_confidential_formula_remains():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = '=CONCAT("Công ty TNHH Sao Mai")'
    stream = BytesIO()
    workbook.save(stream)

    with pytest.raises(AnonymizationExportError, match="company"):
        anonymize_workbook_bytes(
            filename="student.xlsx",
            content=stream.getvalue(),
            session=AnonymizationSession("session-1", "secret"),
            confidential_values={"company": ["Công ty TNHH Sao Mai"]},
        )


def test_anonymize_xlsx_replaces_confidential_tokens_case_insensitively():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "CÔNG TY TNHH SAO MAI"
    stream = BytesIO()
    workbook.save(stream)

    exported = anonymize_workbook_bytes(
        filename="student.xlsx",
        content=stream.getvalue(),
        session=AnonymizationSession("session-1", "secret"),
        confidential_values={"company": ["Công ty TNHH Sao Mai"]},
    )

    assert "công ty tnhh sao mai" not in str(
        load_workbook(BytesIO(exported.content)).active["A1"].value
    ).casefold()


def test_anonymize_xlsx_accepts_a_single_confidential_string_value():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "Công ty TNHH Sao Mai"
    stream = BytesIO()
    workbook.save(stream)

    exported = anonymize_workbook_bytes(
        filename="student.xlsx",
        content=stream.getvalue(),
        session=AnonymizationSession("session-1", "secret"),
        confidential_values={"company": "Công ty TNHH Sao Mai"},
    )

    assert exported.replaced_categories == ("company",)


def test_anonymize_xlsx_sanitizes_metadata_hyperlinks_comments_and_hidden_sheets():
    secret = "Công ty TNHH Sao Mai"
    workbook = Workbook()
    workbook.code_name = secret
    worksheet = workbook.active
    worksheet["A1"] = "Nội dung an toàn"
    worksheet["A1"].comment = Comment(f"Ghi chú của {secret}", secret)
    worksheet["A2"] = "Liên kết"
    worksheet["A2"].hyperlink = f"https://example.test/?company={secret}"
    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    hidden["A1"] = secret
    workbook.properties.creator = secret
    workbook.properties.description = f"Dữ liệu của {secret}"
    workbook.defined_names.add(
        DefinedName("SensitiveCompany", attr_text=f'"{secret}"')
    )
    stream = BytesIO()
    workbook.save(stream)

    exported = anonymize_workbook_bytes(
        filename="student.xlsx",
        content=stream.getvalue(),
        session=AnonymizationSession("session-metadata", "secret"),
        confidential_values={"company": [secret]},
    )

    sanitized = load_workbook(BytesIO(exported.content), data_only=False)
    assert secret.casefold() not in str(sanitized.code_name).casefold()
    assert secret.casefold() not in str(sanitized.properties.creator).casefold()
    assert secret.casefold() not in str(sanitized.properties.description).casefold()
    assert sanitized.sheetnames == ["Sheet1"]
    assert sanitized.active["A1"].comment is None
    assert sanitized.active["A2"].hyperlink is None
    assert "SensitiveCompany" not in sanitized.defined_names
    assert {
        "hidden_sheet_cells",
        "hidden_sheets_removed",
        "comments",
        "hyperlinks",
        "workbook_properties",
        "defined_names",
    } <= set(exported.replaced_layers)


def test_anonymize_xlsx_removes_comments_instead_of_persisting_their_content():
    company = "Công ty TNHH Sao Mai"
    email = "alice@example.com"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"].comment = Comment(
        f"Liên hệ {company} qua {email}",
        f"{company} / {email}",
    )
    stream = BytesIO()
    workbook.save(stream)

    exported = anonymize_workbook_bytes(
        filename="student.xlsx",
        content=stream.getvalue(),
        session=AnonymizationSession("session-comment", "secret"),
        confidential_values={"company": [company], "email": [email]},
    )

    comment = load_workbook(BytesIO(exported.content)).active["A1"].comment
    assert comment is None
    assert "comments" in exported.replaced_layers


def test_anonymize_xlsx_removes_custom_document_properties():
    secret = "Công ty TNHH Sao Mai"
    workbook = Workbook()
    workbook.custom_doc_props.append(
        StringProperty(name="Customer", value=secret)
    )
    stream = BytesIO()
    workbook.save(stream)

    exported = anonymize_workbook_bytes(
        filename="student.xlsx",
        content=stream.getvalue(),
        session=AnonymizationSession("session-custom-props", "secret"),
        confidential_values={"company": [secret]},
    )

    sanitized = load_workbook(BytesIO(exported.content), data_only=False)
    assert len(sanitized.custom_doc_props) == 0
    assert "custom_document_properties" in exported.replaced_layers


def test_anonymize_xlsx_sanitizes_all_ooxml_text_parts():
    secret = "Công ty TNHH Sao Mai"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    workbook.create_sheet(secret)
    worksheet.sheet_properties.codeName = secret
    worksheet.append(["Company", "Amount"])
    worksheet.append([secret, 10])
    worksheet.oddHeader.center.text = secret
    validation = DataValidation(
        type="list",
        formula1='"safe"',
        prompt=secret,
        error=secret,
    )
    worksheet.add_data_validation(validation)
    validation.add(worksheet["A2"])
    worksheet.auto_filter.ref = "A1:B2"
    worksheet.auto_filter.add_filter_column(0, [secret])
    table = Table(displayName="SensitiveTable", ref="A1:B2")
    table.comment = secret
    table.totalsRowLabel = secret
    worksheet.add_table(table)
    chart = BarChart()
    chart.title = secret
    chart.add_data(Reference(worksheet, min_col=2, min_row=1, max_row=2))
    worksheet.add_chart(chart, "D1")
    workbook.properties.revision = secret
    stream = BytesIO()
    workbook.save(stream)

    exported = anonymize_workbook_bytes(
        filename="student.xlsx",
        content=stream.getvalue(),
        session=AnonymizationSession("session-ooxml", "secret"),
        confidential_values={"company": [secret]},
    )

    with ZipFile(BytesIO(exported.content)) as archive:
        text_parts = b"\n".join(
            archive.read(name)
            for name in archive.namelist()
            if name.lower().endswith((".xml", ".rels", ".txt"))
        ).decode("utf-8", errors="ignore")
    assert secret.casefold() not in text_parts.casefold()
    assert "workbook_properties" in exported.replaced_layers


def test_anonymize_xlsx_rejects_oversized_archive_before_decompression(monkeypatch):
    workbook = Workbook()
    workbook.active["A1"] = "safe"
    stream = BytesIO()
    workbook.save(stream)
    monkeypatch.setattr(anonymization_module, "MAX_XLSX_UNCOMPRESSED_BYTES", 64)

    with pytest.raises(
        AnonymizationUnsupportedLayerError,
        match="archive_uncompressed_size",
    ):
        anonymize_workbook_bytes(
            filename="student.xlsx",
            content=stream.getvalue(),
            session=AnonymizationSession("session-archive-limit", "secret"),
            confidential_values={"company": ["Công ty TNHH Sao Mai"]},
        )


def test_formula_protection_never_reuses_existing_workbook_text():
    secret = "Công ty TNHH Sao Mai"
    sentinel_like_text = "__EZFORMAT_PROTECTED_FORMULA_0__"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "=1+1"
    worksheet["A2"] = sentinel_like_text
    validation = DataValidation(type="whole", prompt=secret)
    worksheet.add_data_validation(validation)
    validation.add(worksheet["A2"])
    stream = BytesIO()
    workbook.save(stream)

    exported = anonymize_workbook_bytes(
        filename="student.xlsx",
        content=stream.getvalue(),
        session=AnonymizationSession("session-formula-token", "secret"),
        confidential_values={"company": [secret]},
    )

    sanitized = load_workbook(BytesIO(exported.content), data_only=False)
    assert sanitized.active["A1"].value == "=1+1"
    assert sanitized.active["A2"].value == sentinel_like_text


def test_short_confidential_value_never_rewrites_ooxml_markup():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "a"
    validation = DataValidation(type="whole", prompt="a")
    worksheet.add_data_validation(validation)
    validation.add(worksheet["A1"])
    stream = BytesIO()
    workbook.save(stream)

    exported = anonymize_workbook_bytes(
        filename="student.xlsx",
        content=stream.getvalue(),
        session=AnonymizationSession("session-short-secret", "secret"),
        confidential_values={"company": ["a"]},
    )

    sanitized = load_workbook(BytesIO(exported.content), data_only=False)
    assert sanitized.active["A1"].value != "a"


def test_ooxml_scanner_detects_confidential_text_split_across_rich_text_runs():
    xml = """
    <a:p xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
      <a:r><a:t>Secret</a:t></a:r>
      <a:r><a:t>Co</a:t></a:r>
    </a:p>
    """

    values = anonymization_module._ooxml_text_values(xml)

    assert scan_confidential_values(values, {"company": ["SecretCo"]}) == (
        "company",
    )


def test_anonymize_xlsx_rebuild_removes_unclassified_style_metadata():
    secret = "SecretCo"
    workbook = Workbook()
    workbook.add_named_style(NamedStyle(name=secret))
    stream = BytesIO()
    workbook.save(stream)

    exported = anonymize_workbook_bytes(
        filename="student.xlsx",
        content=stream.getvalue(),
        session=AnonymizationSession("session-style-name", "secret"),
        confidential_values={"company": [secret]},
    )

    sanitized = load_workbook(BytesIO(exported.content), data_only=False)
    assert secret not in sanitized.named_styles


def test_anonymize_xlsx_fails_closed_when_external_links_cannot_be_scanned():
    workbook = Workbook()
    external_link = ExternalLink(
        externalBook=ExternalBook(
            sheetNames=ExternalSheetNames(sheetName=["Sheet1"])
        )
    )
    external_link.file_link = Relationship(
        type="externalLinkPath",
        Target="file:///C:/Sensitive.xlsx",
        TargetMode="External",
    )
    workbook._external_links.append(external_link)
    stream = BytesIO()
    workbook.save(stream)

    with pytest.raises(AnonymizationUnsupportedLayerError, match="external_links"):
        anonymize_workbook_bytes(
            filename="student.xlsx",
            content=stream.getvalue(),
            session=AnonymizationSession("session-external", "secret"),
            confidential_values={"company": ["Công ty TNHH Sao Mai"]},
        )


def test_anonymize_xlsx_fails_closed_for_embedded_binary_images():
    workbook = Workbook()
    image_stream = BytesIO()
    PillowImage.new("RGB", (1, 1), color="white").save(image_stream, format="PNG")
    image_stream.seek(0)
    workbook.active.add_image(WorksheetImage(image_stream), "A1")
    stream = BytesIO()
    workbook.save(stream)

    with pytest.raises(
        AnonymizationUnsupportedLayerError,
        match="embedded_binary_objects",
    ):
        anonymize_workbook_bytes(
            filename="student.xlsx",
            content=stream.getvalue(),
            session=AnonymizationSession("session-image", "secret"),
            confidential_values={"company": ["Công ty TNHH Sao Mai"]},
        )


def test_anonymize_xls_fails_closed_when_metadata_layers_cannot_be_verified():
    source = xlwt.Workbook()
    worksheet = source.add_sheet("Students")
    values = {
        "company": "Công ty TNHH Sao Mai",
        "counterparty": "Nhà cung cấp Sao Mai",
        "tax_code": "0012345678",
        "address": "1 Đường Mới",
        "email": "mai@example.com",
        "phone": "0900000000",
        "bank_account": "0011223344",
        "document_number": "HD-001",
    }
    style = xlwt.easyxf("font: bold on")
    for column_index, value in enumerate(values.values()):
        worksheet.write(0, column_index, value, style)
    original_stream = BytesIO()
    source.save(original_stream)
    with pytest.raises(
        AnonymizationUnsupportedLayerError,
        match="xls_metadata_layers",
    ):
        anonymize_workbook_bytes(
            filename="student.xls",
            content=original_stream.getvalue(),
            session=AnonymizationSession("session-1", "secret"),
            confidential_values={category: [value] for category, value in values.items()},
            full_document_numbers=True,
        )


def test_student_feature_flags_and_retention_values_are_documented():
    root_env = (ROOT / ".env.example").read_text(encoding="utf-8")
    converter_env = (ROOT / "converter" / ".env.example").read_text(encoding="utf-8")
    frontend_env = (ROOT / "frontend" / ".env.example").read_text(encoding="utf-8")

    for flag in FEATURE_FLAGS:
        assert f"{flag}=false" in root_env
        assert f"{flag}=false" in converter_env
        assert f"VITE_{flag}=false" in frontend_env
    assert "STUDENT_CHECK_WORK_ENABLED" not in root_env
    assert "STUDENT_CHECK_WORK_ENABLED" not in converter_env
    assert "VITE_STUDENT_CHECK_WORK_ENABLED" not in frontend_env
    assert "CONVERSION_CONTEXT_SECRET=" in root_env
    assert "CONVERSION_CONTEXT_SECRET=" in converter_env
    for documented_env in (root_env, converter_env):
        assert "STUDENT_ANONYMIZATION_SECRET=" in documented_env
        assert "at least 32 characters" in documented_env
        assert "distinct from CONVERSION_CONTEXT_SECRET and CONVERTER_SERVICE_TOKEN" in documented_env
    assert "STUDENT_UPLOAD_RETENTION_SECONDS=86400" in root_env
    assert "STUDENT_UPLOAD_RETENTION_SECONDS=86400" in converter_env
    assert "STUDENT_UPLOAD_CLEANUP_INTERVAL_SECONDS=300" in converter_env
    assert "LOCAL_MAPPING_OWNER_SCOPE=local:default" in converter_env
