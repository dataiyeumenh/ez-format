from __future__ import annotations

import base64
import hashlib
import hmac
import json
import time
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
from pathlib import Path
from threading import Event, Lock

import openpyxl
import pytest
import xlwt
from fastapi.testclient import TestClient
from openpyxl.comments import Comment
from openpyxl.packaging.custom import StringProperty
from openpyxl.workbook.defined_name import DefinedName

from app import main as main_module
from app import student_store, student_workflow
from app.main import app, clear_student_rate_limits
from app.misa_workflow import _read_metadata, _write_metadata, save_upload
from app.student_anonymization import _workbook_values, scan_confidential_values
from app.student_context import verify_student_context
from app.student_session_client import (
    StudentSessionClientError,
    assert_student_session_active,
    record_analysis_completed,
    record_question_event,
)
from app.student_store import find_student_upload_id, student_upload_retention_seconds


TEST_SERVICE_TOKEN = "converter-service-token-for-student-tests"


def _test_client():
    return TestClient(
        app,
        headers={"X-Converter-Service-Token": TEST_SERVICE_TOKEN},
    )


def _encode_part(payload):
    raw = json.dumps(payload, separators=(",", ":")).encode("utf-8")
    return base64.urlsafe_b64encode(raw).rstrip(b"=").decode("ascii")


def _student_token(secret="student-secret", **overrides):
    payload = {
        "purpose": "student_file_session",
        "session_id": "507f1f77bcf86cd799439011",
        "user_id": "507f1f77bcf86cd799439012",
        "owner_scope": "user:507f1f77bcf86cd799439012",
        "workspace_id": None,
        "snapshot_set_hash": None,
        "allowed_scopes": [
            "analyze",
            "explain",
            "ask",
            "accounting_map",
            "reconcile",
            "export",
        ],
        "iat": int(time.time()),
        "exp": int(time.time()) + 600,
        "retention_expires_at": int(time.time()) + 24 * 60 * 60,
    }
    payload.update(overrides)
    header_part = _encode_part({"alg": "HS256", "typ": "JWT"})
    payload_part = _encode_part(payload)
    signed = f"{header_part}.{payload_part}".encode("ascii")
    signature = hmac.new(secret.encode("utf-8"), signed, hashlib.sha256).digest()
    signature_part = base64.urlsafe_b64encode(signature).rstrip(b"=").decode("ascii")
    return f"{header_part}.{payload_part}.{signature_part}"


def _student_operation_token(secret="student-secret", **overrides):
    session_id = "507f1f77bcf86cd799439011"
    user_id = "507f1f77bcf86cd799439012"
    payload = {
        "purpose": "misa_conversion",
        "user_id": user_id,
        "owner_scope": f"user:{user_id}",
        "workspace_id": None,
        "snapshot_set_hash": None,
        "conversion_run_id": f"student:{session_id}",
        "operation_session_id": session_id,
        "upload_id": "",
        "target_template_id": "bsn_sales",
        "scopes": ["analyze"],
        "exp": int(time.time()) + 600,
    }
    payload.update(overrides)
    header_part = _encode_part({"alg": "HS256", "typ": "JWT"})
    payload_part = _encode_part(payload)
    signed = f"{header_part}.{payload_part}".encode("ascii")
    signature = hmac.new(secret.encode("utf-8"), signed, hashlib.sha256).digest()
    signature_part = base64.urlsafe_b64encode(signature).rstrip(b"=").decode("ascii")
    return f"{header_part}.{payload_part}.{signature_part}"


def _workbook_bytes():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(
        [
            "Mã hóa đơn",
            "Thời gian",
            "Tên khách hàng",
            "Mã hàng",
            "Số lượng",
            "Đơn giá",
        ]
    )
    sheet.append(["HD001", "01/01/2026", "Khách A", "SP001", 2, 100000])
    sheet.append(["HD001", "01/01/2026", "Khách A", "SP002", 1, 50000])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _accounting_workbook_bytes():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(
        [
            "Mã hóa đơn",
            "Thời gian",
            "Tên khách hàng",
            "Mã hàng",
            "Tên hàng",
            "Loại HHDV",
            "ĐVT",
            "Số lượng",
            "Đơn giá",
            "Thành tiền",
            "TK Tiền/Chi phí/Nợ (*)",
            "TK Doanh thu/Có (*)",
        ]
    )
    sheet.append(
        [
            "HD001",
            "01/01/2026",
            "Khách A",
            "SP001",
            "Sản phẩm A",
            "Hàng hóa",
            "Cái",
            2,
            100000,
            200000,
            "131",
            "5111",
        ]
    )
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _structured_workbook_bytes():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(
        [
            "Mã hóa đơn",
            "Thời gian",
            "Tên khách hàng",
            "Mã hàng",
            "Số lượng",
            "Đơn giá",
            "Thành tiền",
        ]
    )
    sheet.append(["HD001", "01/01/2026", "Khách A", "SP001", 2, 100000, "=E2*F2"])
    sheet.row_dimensions[2].hidden = True
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _structured_xls_workbook_bytes():
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Data")
    headers = [
        "Mã hóa đơn",
        "Thời gian",
        "Tên khách hàng",
        "Mã hàng",
        "Số lượng",
        "Đơn giá",
        "Thành tiền",
    ]
    for column, header in enumerate(headers):
        sheet.write(0, column, header)
    for column, value in enumerate(["HD001", "01/01/2026", "Khách A", "SP001", 2, 100000]):
        sheet.write(1, column, value)
    sheet.write(1, 6, xlwt.Formula("E2*F2"))
    sheet.row(1).hidden = True
    sheet.col(0).hidden = True
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _privacy_layer_workbook_bytes(secret: str):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(
        [
            "Mã hóa đơn",
            "Thời gian",
            "Tên khách hàng",
            "Mã hàng",
            "Số lượng",
            "Đơn giá",
        ]
    )
    sheet.append(["HD001", "01/01/2026", "Khách an toàn", "SP001", 2, 100000])
    sheet["A1"].comment = Comment(f"comment:{secret}", secret)
    hidden = workbook.create_sheet(f"hidden-{secret}")
    hidden.sheet_state = "hidden"
    hidden["A1"] = secret
    workbook.properties.creator = secret
    workbook.properties.description = f"metadata:{secret}"
    workbook.custom_doc_props.append(StringProperty(name="Confidential", value=secret))
    workbook.defined_names.add(DefinedName("ConfidentialName", attr_text=f'"{secret}"'))
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


@pytest.fixture
def student_api(tmp_path, monkeypatch):
    clear_student_rate_limits()
    upload_root = tmp_path / "uploads"
    monkeypatch.setenv("STUDENT_ASSISTANT_ENABLED", "true")
    monkeypatch.setenv("STUDENT_FILE_EXPLAIN_ENABLED", "true")
    monkeypatch.setenv("STUDENT_FILE_QA_ENABLED", "true")
    monkeypatch.setenv("STUDENT_ACCOUNTING_MAP_ENABLED", "false")
    monkeypatch.setenv("STUDENT_RECONCILIATION_ENABLED", "false")
    monkeypatch.setenv("STUDENT_INTERNSHIP_ENABLED", "false")
    monkeypatch.setenv("CONVERSION_CONTEXT_SECRET", "student-secret")
    monkeypatch.setenv("CONVERTER_SERVICE_TOKEN", TEST_SERVICE_TOKEN)
    monkeypatch.setenv("STUDENT_ANONYMIZATION_SECRET", "student-anonymization-secret")
    monkeypatch.setenv("MAPPING_DB_PATH", str(tmp_path / "profiles.sqlite"))
    monkeypatch.setenv("AI_PROVIDER", "disabled")
    monkeypatch.setattr("app.misa_workflow.UPLOAD_ROOT", upload_root)
    monkeypatch.setattr("app.student_store.UPLOAD_ROOT", upload_root)
    monkeypatch.setattr("app.misa_workflow.find_mapping_profile", lambda *args, **kwargs: None)
    sync_events = []
    monkeypatch.setattr(
        "app.student_workflow.record_analysis_completed",
        lambda token, payload: sync_events.append(payload),
    )
    monkeypatch.setattr(
        "app.student_workflow.assert_student_session_active",
        lambda token, session_id, upload_id, **kwargs: None,
    )
    return _test_client(), sync_events


def _analyze(client, token=None):
    return _analyze_bytes(client, _workbook_bytes(), token=token)


def _analyze_bytes(client, content, token=None, filename="sales.xlsx"):
    data = {"target_template_id": "bsn_sales"}
    if token is not None:
        data["context_token"] = token
    headers = {"X-Student-Context": token} if token is not None else {}
    return client.post(
        "/api/v1/student/sessions/analyze",
        headers=headers,
        data=data,
        files={
            "file": (
                filename,
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


def _ask(client, session_id, token, question):
    return client.post(
        f"/api/v1/student/sessions/{session_id}/questions",
        headers={"X-Student-Context": token},
        json={"question": question},
    )


def _source_row(client, session_id, token, worksheet_row):
    return client.get(
        f"/api/v1/student/sessions/{session_id}/source-rows/{worksheet_row}",
        headers={"X-Student-Context": token},
    )


def test_converter_startup_validates_student_anonymization_config(monkeypatch):
    checks = []
    monkeypatch.setattr(
        main_module,
        "assert_secure_production_config",
        lambda: checks.append("gateway"),
    )
    monkeypatch.setattr(
        main_module,
        "assert_student_anonymization_config",
        lambda: checks.append("student"),
        raising=False,
    )

    main_module._assert_converter_security_config()

    assert checks == ["gateway", "student"]


def test_student_analyze_requires_valid_signed_context(student_api):
    client, _ = student_api

    assert _analyze(client).status_code == 422
    invalid = _analyze(client, "not-a-token")
    assert invalid.status_code == 401
    assert "context" in invalid.json()["detail"].lower()


def test_student_analyze_uses_node_binding_with_metadata_only_remote_state(
    student_api,
    monkeypatch,
    tmp_path,
):
    client, _ = student_api
    captured_puts = []
    session_id = "507f1f77bcf86cd799439011"
    upload_id = "00000000-0000-4000-8000-000000000011"

    class CapturingRemoteStore:
        def __init__(self):
            self.run_id = f"student:{session_id}"
            self.session_id = session_id

        def put_state(self, **payload):
            captured_puts.append(payload)
            return {"session": {"revision": payload["revision"]}}

        def put_artifact(self, **_payload):
            raise AssertionError("Student analysis must not persist raw upload artifacts remotely")

    remote = CapturingRemoteStore()
    monkeypatch.setenv("OPERATION_STORE_PROVIDER", "node")
    monkeypatch.setenv("OPERATION_SESSION_DIR", str(tmp_path / "operation-sessions"))
    monkeypatch.setattr(
        "app.operation_store.NodeOperationStoreClient",
        lambda _token: remote,
    )
    student_token = _student_token()

    response = client.post(
        "/api/v1/student/sessions/analyze",
        headers={
            "X-Student-Context": student_token,
            "X-Conversion-Context": _student_operation_token(upload_id=upload_id),
        },
        data={
            "context_token": student_token,
            "target_template_id": "bsn_sales",
        },
        files={
            "file": (
                "sales.xlsx",
                _workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    assert len(captured_puts) == 1
    state = captured_puts[0]["state"]
    assert state["contract"] == "student_metadata_v1"
    assert "audit_events" not in state["session"]
    retention_expires_at = verify_student_context(student_token, "analyze").retention_expires_at
    assert time.time() + 23 * 60 * 60 < captured_puts[0]["expires_at"].timestamp()
    assert captured_puts[0]["expires_at"].timestamp() <= retention_expires_at
    serialized = json.dumps(state, ensure_ascii=False)
    for forbidden in (
        "headers",
        "rows",
        "values",
        "Mã hóa đơn",
        "Tên khách hàng",
        "HD001",
        "Khách A",
        "SP001",
    ):
        assert forbidden not in serialized


@pytest.mark.parametrize(
    "failure_mode",
    [
        "sync_error",
        "missing_upload_id",
        "missing_upload_id_purge_partial",
        "post_save_value_error",
    ],
)
def test_production_student_analysis_fails_with_truthful_cleanup_status(
    student_api,
    monkeypatch,
    tmp_path,
    failure_mode,
):
    client, _ = student_api
    session_id = "507f1f77bcf86cd799439011"
    upload_id = "00000000-0000-4000-8000-000000000021"
    deleted_remote = []
    remote_state = {}

    class RemoteStore:
        run_id = "student:507f1f77bcf86cd799439011"
        session_id = "507f1f77bcf86cd799439011"

        @staticmethod
        def put_state(**payload):
            remote_state.update(payload)
            return {"session": {"revision": payload["revision"]}}

        @staticmethod
        def get_state(**_payload):
            return {
                "state": remote_state["state"],
                "session": {"revision": remote_state["revision"]},
            }

        @staticmethod
        def delete_session_artifacts(**payload):
            deleted_remote.append(payload)
            return {
                "success": True,
                "session_id": payload["session_id"],
                "run_id": payload["run_id"],
                "purge_scope": "all_artifacts",
                "remaining_metadata": (
                    1 if failure_mode == "missing_upload_id_purge_partial" else 0
                ),
                "remaining_bytes": 0,
                "remote_operation_session_deleted": (
                    failure_mode != "missing_upload_id_purge_partial"
                ),
            }

    operation_root = tmp_path / "operation-sessions"
    monkeypatch.setenv("OPERATION_STORE_PROVIDER", "node")
    monkeypatch.setenv("OPERATION_SESSION_DIR", str(operation_root))
    monkeypatch.setattr(
        "app.operation_store.NodeOperationStoreClient",
        lambda _token: RemoteStore(),
    )
    if failure_mode == "sync_error":
        monkeypatch.setattr(
            student_workflow,
            "record_analysis_completed",
            lambda *_args, **_kwargs: (_ for _ in ()).throw(
                student_workflow.StudentSessionClientError("backend unavailable")
            ),
        )
    elif failure_mode.startswith("missing_upload_id"):
        original_payload = student_workflow._analysis_completed_payload
        monkeypatch.setattr(
            student_workflow,
            "_analysis_completed_payload",
            lambda overview: {
                **original_payload(overview),
                "converterUploadId": "",
            },
        )
    else:
        monkeypatch.setattr(
            student_workflow,
            "_build_current_overview",
            lambda **_kwargs: (_ for _ in ()).throw(
                ValueError("post-save overview validation failed")
            ),
        )

    response = client.post(
        "/api/v1/student/sessions/analyze",
        headers={
            "X-Student-Context": _student_token(),
            "X-Conversion-Context": _student_operation_token(upload_id=upload_id),
        },
        data={
            "context_token": _student_token(),
            "target_template_id": "bsn_sales",
        },
        files={
            "file": (
                "sales.xlsx",
                _workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    expected_status = 400 if failure_mode == "post_save_value_error" else 503
    assert response.status_code == expected_status, response.text
    if failure_mode == "missing_upload_id_purge_partial":
        assert "purge không hoàn tất" in response.json()["detail"]
        assert "đã được purge" not in response.json()["detail"]
    assert not (student_store.UPLOAD_ROOT / upload_id).exists()
    assert not (operation_root / session_id).exists()
    assert deleted_remote == [
        {"session_id": session_id, "run_id": f"student:{session_id}"}
    ]


def test_student_purge_contract_removes_raw_upload_and_operation_session(
    student_api,
    monkeypatch,
    tmp_path,
):
    client, _ = student_api
    session_id = "507f1f77bcf86cd799439011"
    upload_id = "00000000-0000-4000-8000-000000000031"
    captured = {}
    remote_deletes = []

    class RemoteStore:
        def __init__(self):
            self.run_id = f"student:{session_id}"
            self.session_id = session_id

        @staticmethod
        def put_state(**payload):
            captured.update(payload)
            return {"session": {"revision": payload["revision"]}}

        @staticmethod
        def get_state(**_payload):
            return {
                "state": captured["state"],
                "session": {"revision": captured["revision"]},
            }

        @staticmethod
        def delete_session_artifacts(**payload):
            remote_deletes.append(payload)
            return {
                "success": True,
                "session_id": payload["session_id"],
                "run_id": payload["run_id"],
                "purge_scope": "all_artifacts",
                "remaining_metadata": 0,
                "remaining_bytes": 0,
                "remote_operation_session_deleted": True,
            }

    remote = RemoteStore()
    operation_root = tmp_path / "operation-sessions"
    monkeypatch.setenv("OPERATION_STORE_PROVIDER", "node")
    monkeypatch.setenv("OPERATION_SESSION_DIR", str(operation_root))
    monkeypatch.setattr(
        "app.operation_store.NodeOperationStoreClient",
        lambda _token: remote,
    )
    student_token = _student_token()
    analyzed = client.post(
        "/api/v1/student/sessions/analyze",
        headers={
            "X-Student-Context": student_token,
            "X-Conversion-Context": _student_operation_token(upload_id=upload_id),
        },
        data={"context_token": student_token, "target_template_id": "bsn_sales"},
        files={
            "file": (
                "sales.xlsx",
                _workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert analyzed.status_code == 200, analyzed.text
    assert analyzed.json()["upload_id"] == upload_id
    assert (student_store.UPLOAD_ROOT / upload_id / "input.xlsx").is_file()
    assert (operation_root / session_id / "table.json").is_file()

    purged = client.delete(
        f"/api/v1/student/sessions/{session_id}/purge",
        headers={
            "X-Student-Context": student_token,
            "X-Conversion-Context": _student_operation_token(upload_id=upload_id),
        },
    )

    assert purged.status_code == 200, purged.text
    assert purged.json() == {
        "success": True,
        "session_id": session_id,
        "upload_id": upload_id,
        "raw_upload_deleted": True,
        "operation_session_deleted": True,
        "local_operation_session_deleted": True,
        "remote_operation_session_deleted": True,
    }
    assert not (student_store.UPLOAD_ROOT / upload_id).exists()
    assert not (operation_root / session_id).exists()
    assert remote_deletes == [
        {"session_id": session_id, "run_id": f"student:{session_id}"}
    ]


def test_student_analyze_rejects_upload_over_configured_maximum(
    student_api,
    monkeypatch,
):
    client, _ = student_api
    monkeypatch.setenv("STUDENT_MAX_FILE_BYTES", "8")

    response = _analyze_bytes(
        client,
        b"PK\x03\x04" + b"x" * 5,
        token=_student_token(),
    )

    assert response.status_code == 413
    assert "8 bytes" in response.json()["detail"]


@pytest.mark.parametrize(
    ("filename", "content"),
    [
        ("sales.csv", b"PK\x03\x04not-an-excel-file"),
        ("sales.xlsx", b"not-an-excel-file"),
        ("sales.xlsx", b"PK-not-an-openxml-archive"),
    ],
)
def test_student_analyze_rejects_unsupported_extension_or_bad_excel_magic(
    student_api,
    filename,
    content,
):
    client, _ = student_api

    response = _analyze_bytes(
        client,
        content,
        token=_student_token(),
        filename=filename,
    )

    assert response.status_code == 415


def test_student_analyze_is_hidden_when_phase_flag_is_disabled(student_api, monkeypatch):
    client, _ = student_api
    monkeypatch.setenv("STUDENT_FILE_EXPLAIN_ENABLED", "false")

    response = _analyze(client, _student_token())

    assert response.status_code == 404
    assert "chưa được bật" in response.json()["detail"]


def test_student_analyze_surfaces_formula_and_hidden_row_warnings(student_api):
    client, _ = student_api
    response = _analyze_bytes(client, _structured_workbook_bytes(), token=_student_token())

    assert response.status_code == 200
    payload = response.json()
    assert payload["workbook_structure"]["formula_cell_count"] == 1
    assert payload["workbook_structure"]["hidden_row_count"] == 1
    issues = {
        issue["code"]: issue
        for issue in payload["readiness"]["issues"]
        if issue["code"] in {"formula_cells_detected", "hidden_rows_or_columns_detected"}
    }
    assert set(issues) == {
        "formula_cells_detected",
        "hidden_rows_or_columns_detected",
    }
    assert all(issue["severity"] == "warning" for issue in issues.values())
    explanation_codes = {item.get("issue_code") for item in payload["explanations"]}
    assert set(issues).issubset(explanation_codes)


def test_student_analyze_surfaces_xls_formula_capability_warning_and_normalized_hidden_evidence(
    student_api,
):
    client, _ = student_api
    response = _analyze_bytes(
        client,
        _structured_xls_workbook_bytes(),
        token=_student_token(),
        filename="sales.xls",
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["workbook_structure"]["format"] == "xls"
    assert payload["workbook_structure"]["formula_cell_count"] == 0
    assert payload["workbook_structure"]["formula_detection"] == "unavailable"
    issues = {
        issue["code"]: issue
        for issue in payload["readiness"]["issues"]
        if issue["code"]
        in {"formula_detection_unavailable", "hidden_rows_or_columns_detected"}
    }
    assert set(issues) == {
        "formula_detection_unavailable",
        "hidden_rows_or_columns_detected",
    }
    assert all(issue["severity"] == "warning" for issue in issues.values())
    assert issues["hidden_rows_or_columns_detected"]["evidence"] == [
        {
            "kind": "source_cell",
            "sheet": "Data",
            "row": 2,
            "source_ref": "sheet:Data:row:2:hidden",
        },
        {
            "kind": "source_column",
            "sheet": "Data",
            "column": "A",
            "source_ref": "sheet:Data:column:A:hidden",
        },
    ]
    explanation_codes = {item.get("issue_code") for item in payload["explanations"]}
    assert set(issues).issubset(explanation_codes)


def test_student_analyze_returns_summary_preview_and_evidence_backed_explanations(
    student_api,
):
    client, sync_events = student_api
    response = _analyze(client, _student_token())

    assert response.status_code == 200
    payload = response.json()
    assert payload["target_template_id"] == "bsn_sales"
    assert payload["student_summary"]["session_id"] == "507f1f77bcf86cd799439011"
    assert payload["student_summary"]["data_row_count"] == 2
    assert payload["student_summary"]["document_count"] == 1
    assert payload["student_summary"]["recognized_columns"] == 6
    assert payload["student_summary"]["mapping_counts"]["mapped"] >= 6
    assert payload["student_summary"]["mapping_counts"]["default"] >= 1
    assert payload["student_summary"]["mapping_counts"]["formula"] >= 1
    assert len(payload["student_preview"]["rows"]) == 2

    explanations = payload["explanations"]
    assert explanations
    assert all(item["evidence"] for item in explanations)
    assert all(item["state_hash"] == payload["student_state_hash"] for item in explanations)
    assert any(item["kind"] == "mapping" for item in explanations)
    assert any(item["kind"] == "field" for item in explanations)
    assert any(item["kind"] == "calculation" for item in explanations)
    assert any(item["kind"] == "issue" for item in explanations)
    assert any(
        item["kind"] == "field" and item["target_field"] == "Ngày hạch toán (*)"
        for item in explanations
    )
    assert any(
        item["kind"] == "calculation" and item["target_field"] == "Thành tiền"
        for item in explanations
    )
    date_normalizations = [
        item
        for item in explanations
        if item["kind"] == "normalization"
        and item["target_field"] == "Ngày hạch toán (*)"
    ]
    assert {item["preview_row"] for item in date_normalizations} == {1, 2}
    for item in [entry for entry in explanations if entry["kind"] == "mapping"]:
        evidence_sources = sorted(
            {
                evidence["column"]
                for evidence in item["evidence"]
                if evidence["kind"] in {"source_column", "source_cell"}
                and evidence.get("column")
            }
        )
        assert sorted(item["claim_sources"]) == evidence_sources

    assert len(sync_events) == 1
    sync_payload = sync_events[0]
    assert sync_payload["event"] == "analysis_completed"
    assert sync_payload["converterUploadId"] == payload["upload_id"]
    assert sync_payload["targetTemplateId"] == "bsn_sales"
    assert sync_payload["status"] == "analyzed"
    assert "rows" not in json.dumps(sync_payload).lower()


def test_student_analyze_rejects_retry_when_session_already_has_upload(student_api):
    client, _ = student_api
    token = _student_token()

    first = _analyze(client, token)
    retry = _analyze(client, token)

    assert first.status_code == 200
    assert retry.status_code == 409
    assert "đã có upload" in retry.json()["detail"].lower()


def test_concurrent_student_analyze_allows_exactly_one_active_upload(
    student_api, monkeypatch
):
    _, _ = student_api
    token = _student_token()
    first_entered = Event()
    release_first = Event()
    call_lock = Lock()
    call_count = 0

    def controlled_analyze_upload(**kwargs):
        nonlocal call_count
        with call_lock:
            call_count += 1
            current_call = call_count
        if current_call == 1:
            first_entered.set()
            assert release_first.wait(timeout=5)
        claims = verify_student_context(kwargs["student_context_token"], "analyze")
        upload_id, _ = save_upload(
            kwargs["filename"],
            kwargs["content"],
            student_claims=claims,
            student_ttl_seconds=student_upload_retention_seconds(),
        )
        return {"upload_id": upload_id}

    def fast_overview(*, upload_id, **_kwargs):
        return {
            "upload_id": upload_id,
            "target_template_id": "bsn_sales",
            "detected": {"source_signature_hash": "test-signature"},
            "student_summary": {
                "session_id": "507f1f77bcf86cd799439011",
                "data_row_count": 0,
                "document_count": 0,
                "recognized_columns": 0,
                "unresolved_columns": 0,
                "mapping_counts": {},
                "issue_counts": {},
                "master_data_status": "not_configured",
                "explanation_count": 0,
            },
            "student_state_hash": "test-state",
            "readiness": {"status": "ready"},
        }

    monkeypatch.setattr(
        student_workflow,
        "analyze_upload",
        controlled_analyze_upload,
    )
    monkeypatch.setattr(student_workflow, "_build_current_overview", fast_overview)

    with _test_client() as first_client, _test_client() as second_client:
        with ThreadPoolExecutor(max_workers=2) as executor:
            first_future = executor.submit(_analyze, first_client, token)
            assert first_entered.wait(timeout=5)
            second_future = executor.submit(_analyze, second_client, token)
            try:
                second_response = second_future.result(timeout=5)
            finally:
                release_first.set()
            first_response = first_future.result(timeout=5)

    assert sorted(
        [first_response.status_code, second_response.status_code]
    ) == [200, 409]
    assert call_count == 1
    claims = verify_student_context(token, "analyze")
    active_upload_id = find_student_upload_id(claims)
    upload_root = student_store.UPLOAD_ROOT
    active_uploads = [
        path.name
        for path in upload_root.iterdir()
        if path.is_dir() and (path / "student.json").is_file()
    ]
    assert active_uploads == [active_upload_id]


def test_student_overview_enforces_session_owner_binding(student_api):
    client, _ = student_api
    token = _student_token()
    analyzed = _analyze(client, token).json()

    missing = client.get(
        "/api/v1/student/sessions/507f1f77bcf86cd799439011/overview"
    )
    assert missing.status_code == 401

    other_token = _student_token(
        session_id="507f1f77bcf86cd799439099",
        user_id="507f1f77bcf86cd799439098",
        owner_scope="user:507f1f77bcf86cd799439098",
    )
    denied = client.get(
        "/api/v1/student/sessions/507f1f77bcf86cd799439011/overview",
        headers={"X-Student-Context": other_token},
    )
    assert denied.status_code == 403

    allowed = client.get(
        "/api/v1/student/sessions/507f1f77bcf86cd799439011/overview",
        headers={"X-Student-Context": token},
    )
    assert allowed.status_code == 200
    assert allowed.json()["upload_id"] == analyzed["upload_id"]


def test_student_overview_rebuilds_when_mapping_state_changes(student_api):
    client, _ = student_api
    token = _student_token()
    analyzed = _analyze(client, token).json()
    first_state_hash = analyzed["student_state_hash"]

    metadata = _read_metadata(analyzed["upload_id"])
    suggestion = metadata["suggestion"]
    changed_defaults = dict(suggestion["defaults"])
    changed_defaults["Loại tiền"] = "USD"
    metadata["confirmed"] = {
        "mapping": suggestion["mapping"],
        "defaults": changed_defaults,
        "formulas": suggestion["formulas"],
    }
    _write_metadata(analyzed["upload_id"], metadata)

    refreshed = client.get(
        "/api/v1/student/sessions/507f1f77bcf86cd799439011/overview",
        headers={"X-Student-Context": token},
    )
    assert refreshed.status_code == 200
    payload = refreshed.json()
    assert payload["student_state_hash"] != first_state_hash
    assert payload["mapping_suggestion"]["source"] == "confirmed"
    assert all(item["state_hash"] == payload["student_state_hash"] for item in payload["explanations"])
    assert all(item["stale"] is False for item in payload["explanations"])
    assert any(
        item["kind"] == "field"
        and item["target_field"] == "Loại tiền"
        and item["reason_vi"].startswith("Giá trị mặc định")
        for item in payload["explanations"]
    )


def test_unchanged_confirmed_mapping_invalidates_heuristic_overview_cache(student_api):
    client, _ = student_api
    token = _student_token()
    analyzed = _analyze(client, token).json()
    first_state_hash = analyzed["student_state_hash"]

    metadata = _read_metadata(analyzed["upload_id"])
    suggestion = metadata["suggestion"]
    metadata["profile_id"] = "507f1f77bcf86cd799439055"
    metadata["confirmed"] = {
        "mapping": suggestion["mapping"],
        "defaults": suggestion["defaults"],
        "formulas": suggestion["formulas"],
    }
    _write_metadata(analyzed["upload_id"], metadata)

    refreshed = client.get(
        "/api/v1/student/sessions/507f1f77bcf86cd799439011/overview",
        headers={"X-Student-Context": token},
    ).json()

    assert refreshed["mapping_suggestion"]["source"] == "confirmed"
    assert refreshed["student_state_hash"] != first_state_hash


def test_analysis_completed_client_sends_service_and_student_context_headers(
    monkeypatch,
):
    captured = {}

    class Response:
        status_code = 200

    def fake_post(url, **kwargs):
        captured["url"] = url
        captured.update(kwargs)
        return Response()

    monkeypatch.setenv("CONVERTER_SERVICE_TOKEN", "service-secret")
    monkeypatch.setenv("NODE_INTERNAL_API_URL", "http://node.test/api/internal")
    monkeypatch.setattr("app.student_session_client.httpx.post", fake_post)
    token = _student_token()
    payload = {
        "event": "analysis_completed",
        "sessionId": "507f1f77bcf86cd799439011",
        "converterUploadId": "upload-1",
    }

    record_analysis_completed(token, payload)

    assert captured["url"].endswith(
        "/student/sessions/507f1f77bcf86cd799439011/events"
    )
    assert captured["headers"] == {
        "x-converter-service-token": "service-secret",
        "x-student-context": token,
    }
    assert captured["json"] == payload


def test_student_question_requires_phase_flag_ask_scope_and_matching_session(
    student_api,
    monkeypatch,
):
    client, _ = student_api
    session_id = "507f1f77bcf86cd799439011"
    token = _student_token()
    assert _analyze(client, token).status_code == 200

    monkeypatch.setenv("STUDENT_FILE_QA_ENABLED", "false")
    disabled = _ask(client, session_id, token, "File này có bao nhiêu dòng?")
    assert disabled.status_code == 404

    monkeypatch.setenv("STUDENT_FILE_QA_ENABLED", "true")
    missing_scope = _ask(
        client,
        session_id,
        _student_token(allowed_scopes=["analyze", "explain"]),
        "File này có bao nhiêu dòng?",
    )
    assert missing_scope.status_code == 401
    assert "ask" in missing_scope.json()["detail"]

    other_session = _ask(
        client,
        "507f1f77bcf86cd799439099",
        token,
        "File này có bao nhiêu dòng?",
    )
    assert other_session.status_code == 403


def test_student_question_returns_valid_evidence_and_sanitized_best_effort_event(
    student_api,
    monkeypatch,
):
    client, _ = student_api
    token = _student_token()
    analyzed = _analyze(client, token).json()
    captured = []
    active_checks = []
    monkeypatch.setattr(
        "app.student_workflow.assert_student_session_active",
        lambda context_token, session_id, upload_id: active_checks.append(
            (context_token, session_id, upload_id)
        ),
    )
    monkeypatch.setattr(
        "app.student_workflow.record_question_event",
        lambda context_token, payload: captured.append((context_token, payload)),
    )

    response = _ask(
        client,
        "507f1f77bcf86cd799439011",
        token,
        "Những dòng nào có hóa đơn HD001?",
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["intent"] == "locate_rows"
    assert payload["outcome"] == "supported"
    assert payload["evidence"]
    active_headers = set(analyzed["detected"]["headers"])
    first_data_row = analyzed["detected"]["header_row"] + 1
    last_data_row = first_data_row + analyzed["detected"]["row_count"] - 1
    assert all(item["field"] in active_headers for item in payload["evidence"])
    assert all(first_data_row <= item["row"] <= last_data_row for item in payload["evidence"])
    assert payload["event_sync"]["status"] == "synced"
    assert active_checks == [
        (token, "507f1f77bcf86cd799439011", analyzed["upload_id"])
    ]

    assert len(captured) == 1
    context_token, event = captured[0]
    assert context_token == token
    assert event == {
        "event": "question_answered",
        "sessionId": "507f1f77bcf86cd799439011",
        "questionHash": hashlib.sha256(
            "Những dòng nào có hóa đơn HD001?".encode("utf-8")
        ).hexdigest(),
        "questionLength": len("Những dòng nào có hóa đơn HD001?"),
        "category": "locate_rows",
        "operation": "ask",
        "answerType": "deterministic_file_query",
        "evidenceIds": [item["id"] for item in payload["evidence"]],
        "evidenceCount": payload["evidence_count"],
        "outcome": "supported",
    }
    serialized = json.dumps(event, ensure_ascii=False).lower()
    assert "những dòng nào" not in serialized
    assert '"rows":' not in serialized
    assert "actual" not in serialized
    assert "expected" not in serialized


def test_student_question_is_rate_limited_per_session_context(
    student_api,
    monkeypatch,
):
    client, _ = student_api
    monkeypatch.setenv("STUDENT_QUESTION_LIMIT_PER_15_MINUTES", "1")
    session_id = "507f1f77bcf86cd799439011"
    token = _student_token()
    assert _analyze(client, token).status_code == 200

    first = _ask(client, session_id, token, "File này có bao nhiêu dòng?")
    second = _ask(client, session_id, token, "File này có bao nhiêu dòng?")

    assert first.status_code == 200
    assert second.status_code == 429
    assert "quá nhiều" in second.json()["detail"].lower()


def test_student_rate_limit_is_stable_across_context_refresh(student_api, monkeypatch):
    client, _ = student_api
    monkeypatch.setenv("STUDENT_QUESTION_LIMIT_PER_15_MINUTES", "1")
    session_id = "507f1f77bcf86cd799439011"
    old_token = _student_token(iat=int(time.time()) - 10, exp=int(time.time()) + 300)
    fresh_token = _student_token(iat=int(time.time()), exp=int(time.time()) + 600)
    assert _analyze(client, old_token).status_code == 200

    first = _ask(client, session_id, old_token, "File này có bao nhiêu dòng?")
    second = _ask(client, session_id, fresh_token, "File này có bao nhiêu dòng?")

    assert first.status_code == 200
    assert second.status_code == 429


def test_invalid_student_context_does_not_allocate_rate_limit_bucket(student_api):
    client, _ = student_api
    clear_student_rate_limits()

    response = _ask(
        client,
        "507f1f77bcf86cd799439011",
        "arbitrary.invalid.token",
        "File này có bao nhiêu dòng?",
    )

    assert response.status_code == 401
    assert main_module._STUDENT_RATE_BUCKETS == {}


def test_student_question_event_failure_does_not_hide_deterministic_answer(
    student_api,
    monkeypatch,
):
    client, _ = student_api
    token = _student_token()
    assert _analyze(client, token).status_code == 200

    def unavailable(*args, **kwargs):
        raise student_workflow.StudentSessionClientError("node offline")

    monkeypatch.setattr("app.student_workflow.record_question_event", unavailable)
    response = _ask(
        client,
        "507f1f77bcf86cd799439011",
        token,
        "Tổng đơn giá là bao nhiêu?",
    )

    assert response.status_code == 200
    assert response.json()["outcome"] == "supported"
    assert response.json()["event_sync"]["status"] == "unavailable"


def test_question_event_client_sends_only_sanitized_metadata(monkeypatch):
    captured = {}

    class Response:
        status_code = 202

    def fake_post(url, **kwargs):
        captured["url"] = url
        captured.update(kwargs)
        return Response()

    monkeypatch.setenv("CONVERTER_SERVICE_TOKEN", "service-secret")
    monkeypatch.setenv("NODE_INTERNAL_API_URL", "http://node.test/api/internal")
    monkeypatch.setattr("app.student_session_client.httpx.post", fake_post)
    token = _student_token()
    payload = {
        "event": "question_answered",
        "sessionId": "507f1f77bcf86cd799439011",
        "questionHash": hashlib.sha256("Có bao nhiêu hóa đơn?".encode("utf-8")).hexdigest(),
        "questionLength": len("Có bao nhiêu hóa đơn?"),
        "category": "count_documents",
        "operation": "ask",
        "answerType": "deterministic_file_query",
        "evidenceIds": ["question-evidence-1"],
        "evidenceCount": 1,
        "outcome": "supported",
    }

    record_question_event(token, payload)

    assert captured["url"].endswith(
        "/student/sessions/507f1f77bcf86cd799439011/questions"
    )
    assert captured["headers"] == {
        "x-converter-service-token": "service-secret",
        "x-student-context": token,
    }
    assert captured["json"] == payload


@pytest.mark.parametrize(
    ("node_status", "expected_status"),
    [(503, 503), (410, 410), (403, 403), (409, 409)],
)
def test_student_question_fails_closed_when_node_session_is_unavailable_or_inactive(
    student_api,
    monkeypatch,
    node_status,
    expected_status,
):
    client, _ = student_api
    token = _student_token()
    assert _analyze(client, token).status_code == 200
    query_called = False

    def fail_active_check(context_token, session_id, upload_id):
        raise StudentSessionClientError("node session inactive", status_code=node_status)

    def should_not_query(*args, **kwargs):
        nonlocal query_called
        query_called = True
        raise AssertionError("query must not run before active Node session check")

    monkeypatch.setattr(
        "app.student_workflow.assert_student_session_active",
        fail_active_check,
    )
    monkeypatch.setattr("app.student_workflow.answer_question", should_not_query)

    response = _ask(
        client,
        "507f1f77bcf86cd799439011",
        token,
        "Có bao nhiêu hóa đơn?",
    )

    assert response.status_code == expected_status
    assert query_called is False


def test_active_session_client_authenticates_with_service_and_signed_context(monkeypatch):
    captured = {}

    class Response:
        status_code = 200

        @staticmethod
        def json():
            return {"success": True, "active": True}

    def fake_get(url, **kwargs):
        captured["url"] = url
        captured.update(kwargs)
        return Response()

    monkeypatch.setenv("CONVERTER_SERVICE_TOKEN", "service-secret")
    monkeypatch.setenv("NODE_INTERNAL_API_URL", "http://node.test/api/internal")
    monkeypatch.setattr("app.student_session_client.httpx.get", fake_get)
    token = _student_token()

    assert_student_session_active(token, "507f1f77bcf86cd799439011", "upload-1")

    assert captured["url"].endswith(
        "/student/sessions/507f1f77bcf86cd799439011/active"
    )
    assert captured["headers"] == {
        "x-converter-service-token": "service-secret",
        "x-student-context": token,
    }
    assert captured["params"] == {"uploadId": "upload-1"}


def test_source_row_endpoint_returns_exact_owner_bound_worksheet_row(student_api):
    client, _ = student_api
    session_id = "507f1f77bcf86cd799439011"
    token = _student_token()
    analyzed = _analyze(client, token).json()

    response = _source_row(client, session_id, token, 3)

    assert response.status_code == 200
    payload = response.json()
    assert payload["session_id"] == session_id
    assert payload["upload_id"] == analyzed["upload_id"]
    assert payload["sheet"] == "Data"
    assert payload["worksheet_row"] == 3
    assert payload["header_row"] == 1
    assert payload["fields"] == [
        {"field": "Mã hóa đơn", "value": "HD001"},
        {"field": "Thời gian", "value": "01/01/2026"},
        {"field": "Tên khách hàng", "value": "Khách A"},
        {"field": "Mã hàng", "value": "SP002"},
        {"field": "Số lượng", "value": 1},
        {"field": "Đơn giá", "value": 50000},
    ]


@pytest.mark.parametrize("worksheet_row", [1, 99])
def test_source_row_endpoint_rejects_header_and_out_of_range_rows(
    student_api,
    worksheet_row,
):
    client, _ = student_api
    token = _student_token()
    assert _analyze(client, token).status_code == 200

    response = _source_row(
        client,
        "507f1f77bcf86cd799439011",
        token,
        worksheet_row,
    )

    assert response.status_code == 404


def test_source_row_endpoint_is_ask_scope_and_session_bounded(student_api):
    client, _ = student_api
    token = _student_token()
    assert _analyze(client, token).status_code == 200

    missing_scope = _source_row(
        client,
        "507f1f77bcf86cd799439011",
        _student_token(allowed_scopes=["analyze", "explain"]),
        2,
    )
    wrong_session = _source_row(
        client,
        "507f1f77bcf86cd799439099",
        token,
        2,
    )

    assert missing_scope.status_code == 401
    assert wrong_session.status_code == 403


def test_accounting_map_endpoint_is_flag_scope_and_owner_bounded(
    student_api,
    monkeypatch,
):
    client, _ = student_api
    session_id = "507f1f77bcf86cd799439011"
    token = _student_token()
    analyzed = _analyze_bytes(client, _accounting_workbook_bytes(), token=token)
    assert analyzed.status_code == 200

    disabled = client.get(
        f"/api/v1/student/sessions/{session_id}/accounting-map",
        headers={"X-Student-Context": token},
    )
    assert disabled.status_code == 404

    monkeypatch.setenv("STUDENT_ACCOUNTING_MAP_ENABLED", "true")
    missing_scope = client.get(
        f"/api/v1/student/sessions/{session_id}/accounting-map",
        headers={
            "X-Student-Context": _student_token(
                allowed_scopes=["analyze", "explain"]
            )
        },
    )
    assert missing_scope.status_code == 401
    wrong_owner = client.get(
        f"/api/v1/student/sessions/{session_id}/accounting-map",
        headers={
            "X-Student-Context": _student_token(
                session_id="507f1f77bcf86cd799439099"
            )
        },
    )
    assert wrong_owner.status_code == 403

    activities = []
    monkeypatch.setattr(
        student_workflow,
        "record_activity_event",
        lambda context_token, payload: activities.append(payload),
        raising=False,
    )
    response = client.get(
        f"/api/v1/student/sessions/{session_id}/accounting-map",
        headers={"X-Student-Context": token},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["student_state_hash"] == analyzed.json()["student_state_hash"]
    assert payload["maps"][0]["business_event"] == "sales_goods"
    assert payload["maps"][0]["balanced"] is True
    assert [entry["account"] for entry in payload["maps"][0]["entries"]] == [
        "131",
        "5111",
    ]
    assert all(entry["evidence"] for entry in payload["maps"][0]["entries"])
    assert activities == [
        {
            "sessionId": session_id,
            "eventType": "accounting_map_reviewed",
            "evidenceCount": 2,
        }
    ]


def test_reconciliation_endpoint_keeps_insufficient_data_visible(
    student_api,
    monkeypatch,
):
    client, _ = student_api
    session_id = "507f1f77bcf86cd799439011"
    token = _student_token()
    analyzed = _analyze(client, token)
    assert analyzed.status_code == 200
    monkeypatch.setenv("STUDENT_RECONCILIATION_ENABLED", "true")
    activities = []
    monkeypatch.setattr(
        student_workflow,
        "record_activity_event",
        lambda context_token, payload: activities.append(payload),
        raising=False,
    )

    response = client.get(
        f"/api/v1/student/sessions/{session_id}/reconciliation",
        headers={"X-Student-Context": token},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "insufficient_data"
    assert payload["ok"] is False
    insufficient = [
        item for item in payload["items"] if item["status"] == "insufficient_data"
    ]
    assert insufficient
    assert all(item["deterministic"] is False for item in insufficient)
    assert activities[0]["eventType"] == "reconciliation_completed"


def test_anonymization_preview_and_export_are_copy_only_and_scanner_gated(
    student_api,
    monkeypatch,
):
    client, _ = student_api
    session_id = "507f1f77bcf86cd799439011"
    token = _student_token()
    analyzed = _analyze(client, token)
    assert analyzed.status_code == 200
    upload_id = analyzed.json()["upload_id"]
    input_path = _read_metadata(upload_id)["input_path"]
    original = open(input_path, "rb").read()
    monkeypatch.setenv("STUDENT_INTERNSHIP_ENABLED", "true")
    activities = []
    monkeypatch.setattr(
        student_workflow,
        "record_activity_event",
        lambda context_token, payload: activities.append(payload),
        raising=False,
    )

    preview = client.post(
        f"/api/v1/student/sessions/{session_id}/anonymization/preview",
        headers={"X-Student-Context": token},
        json={"full_document_numbers": False},
    )
    assert preview.status_code == 200
    preview_payload = preview.json()
    assert preview_payload["scanner_status"] == "passed"
    assert preview_payload["replaced_categories"] == ["counterparty"]
    assert "Khách A" not in json.dumps(preview_payload, ensure_ascii=False)
    assert open(input_path, "rb").read() == original

    exported = client.post(
        f"/api/v1/student/sessions/{session_id}/anonymization/export",
        headers={"X-Student-Context": token},
        json={"full_document_numbers": False},
    )
    assert exported.status_code == 200
    assert exported.content != original
    assert open(input_path, "rb").read() == original
    workbook = openpyxl.load_workbook(BytesIO(exported.content), data_only=False)
    assert workbook.active["C2"].value != "Khách A"
    assert activities == [
        {
            "sessionId": session_id,
            "eventType": "anonymized_export_created",
            "evidenceCount": 1,
        }
    ]


def test_anonymized_export_removes_confidential_hidden_comment_and_metadata_layers(
    student_api,
    monkeypatch,
):
    client, _ = student_api
    session_id = "507f1f77bcf86cd799439011"
    token = _student_token()
    secret = "HIDDEN-META-SECRET-9472"
    analyzed = _analyze_bytes(
        client,
        _privacy_layer_workbook_bytes(secret),
        token=token,
        filename="privacy.xlsx",
    )
    assert analyzed.status_code == 200
    monkeypatch.setenv("STUDENT_INTERNSHIP_ENABLED", "true")

    preview = client.post(
        f"/api/v1/student/sessions/{session_id}/anonymization/preview",
        headers={"X-Student-Context": token},
        json={"full_document_numbers": False},
    )
    exported = client.post(
        f"/api/v1/student/sessions/{session_id}/anonymization/export",
        headers={"X-Student-Context": token},
        json={"full_document_numbers": False},
    )

    assert preview.status_code == 200
    assert preview.json()["scanner_status"] == "passed"
    assert exported.status_code == 200
    assert secret.encode("utf-8") not in exported.content
    sanitized = openpyxl.load_workbook(BytesIO(exported.content), data_only=False)
    assert sanitized.sheetnames == ["Data"]
    assert all(
        cell.comment is None
        for worksheet in sanitized.worksheets
        for row in worksheet.iter_rows()
        for cell in row
    )
    assert not sanitized.defined_names
    assert len(sanitized.custom_doc_props) == 0
    assert sanitized.properties.creator in {None, ""}
    assert sanitized.properties.description in {None, ""}


def test_anonymized_export_uses_analyzed_sheet_not_active_confidential_cover(
    student_api,
    monkeypatch,
):
    client, _ = student_api
    session_id = "507f1f77bcf86cd799439011"
    token = _student_token()
    cover_secret = "Khách A"
    source = openpyxl.load_workbook(BytesIO(_workbook_bytes()), data_only=False)
    cover_sheet = source.create_sheet("Confidential Cover", 0)
    cover_sheet["A1"] = cover_secret
    source.active = 0
    source_bytes = BytesIO()
    source.save(source_bytes)

    analyzed = _analyze_bytes(
        client,
        source_bytes.getvalue(),
        token=token,
        filename="two-visible-sheets.xlsx",
    )
    assert analyzed.status_code == 200
    monkeypatch.setenv("STUDENT_INTERNSHIP_ENABLED", "true")

    preview = client.post(
        f"/api/v1/student/sessions/{session_id}/anonymization/preview",
        headers={"X-Student-Context": token},
        json={"full_document_numbers": False},
    )
    exported = client.post(
        f"/api/v1/student/sessions/{session_id}/anonymization/export",
        headers={"X-Student-Context": token},
        json={"full_document_numbers": False},
    )

    assert preview.status_code == 200
    assert preview.json()["scanner_status"] == "passed"
    assert exported.status_code == 200
    sanitized = openpyxl.load_workbook(BytesIO(exported.content), data_only=False)
    assert sanitized.sheetnames == ["Data"]
    assert sanitized["Data"]["C2"].value != cover_secret
    assert "Confidential Cover" not in sanitized.sheetnames
    assert (
        scan_confidential_values(
            _workbook_values(exported.content, ".xlsx"),
            {"counterparty": [cover_secret]},
        )
        == ()
    )
    assert exported.headers["X-Anonymization-Scanner"] == "passed"


def test_student_analysis_persists_no_preview_or_mapped_row_values(student_api):
    client, _ = student_api
    response = _analyze(client, _student_token())
    assert response.status_code == 200
    assert response.json()["student_preview"]["rows"]

    upload_id = response.json()["upload_id"]
    input_path = _read_metadata(upload_id)["input_path"]
    json_paths = sorted(Path(input_path).parent.glob("*.json"))
    assert json_paths
    for path in json_paths:
        serialized = path.read_text(encoding="utf-8")
        assert "Khách A" not in serialized
        payload = json.loads(serialized)
        assert "rows" not in json.dumps(payload, ensure_ascii=False).lower()
        assert "student_preview" not in payload

    persisted_overview = json.loads(
        (Path(input_path).parent / "student-overview.json").read_text(
            encoding="utf-8"
        )
    )
    assert set(persisted_overview) <= {
        "cache_version",
        "upload_id",
        "student_state_hash",
        "target_template_id",
        "schema",
        "counts",
        "anonymized_metadata",
    }


def test_internship_report_requires_verified_activity_ids_and_safe_approved_notes(
    student_api,
    monkeypatch,
):
    client, _ = student_api
    session_id = "507f1f77bcf86cd799439011"
    token = _student_token()
    analyzed = _analyze(client, token)
    assert analyzed.status_code == 200
    monkeypatch.setenv("STUDENT_INTERNSHIP_ENABLED", "true")
    monkeypatch.setattr(
        student_workflow,
        "get_verified_activities",
        lambda context_token, requested_session_id: {
            "activities": [
                {
                    "id": "activity-1",
                    "event_type": "reconciliation_completed",
                    "skill": "VAT reconciliation",
                    "summary": "Completed a deterministic reconciliation review.",
                    "evidence_count": 8,
                    "resolved_issues": [],
                }
            ]
        },
        raising=False,
    )

    invented = client.post(
        f"/api/v1/student/sessions/{session_id}/internship-report",
        headers={"X-Student-Context": token},
        json={"activity_ids": ["activity-invented"], "approved_notes": []},
    )
    assert invented.status_code == 400

    unsafe = client.post(
        f"/api/v1/student/sessions/{session_id}/internship-report",
        headers={"X-Student-Context": token},
        json={"activity_ids": ["activity-1"], "approved_notes": ["<script>"]},
    )
    assert unsafe.status_code == 400

    response = client.post(
        f"/api/v1/student/sessions/{session_id}/internship-report",
        headers={"X-Student-Context": token},
        json={
            "activity_ids": ["activity-1"],
            "approved_notes": ["Reviewed with the supervisor."],
        },
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/markdown")
    assert "Completed a deterministic reconciliation review." not in response.text
    assert "Reviewed with the supervisor." not in response.text
    assert "PSEUDO-" in response.text
    assert "Khách A" not in response.text


def test_student_grading_endpoints_are_not_available(student_api):
    client, _ = student_api
    token = _student_token()

    attempt = client.post(
        "/api/v1/student/sessions/507f1f77bcf86cd799439011/attempts",
        headers={"X-Student-Context": token},
        json={},
    )
    hint = client.post(
        "/api/v1/student/sessions/507f1f77bcf86cd799439011/attempts/attempt-1/hints/0",
        headers={"X-Student-Context": token},
        json={},
    )

    assert attempt.status_code == 404
    assert hint.status_code == 404
