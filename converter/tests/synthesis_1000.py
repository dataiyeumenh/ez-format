"""Synthetic data generators for 1000× specialized backend scenarios."""

from __future__ import annotations

import random
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
import xlwt

from app.cell_ref import excel_cell
from app.document_import import classify_pdf_bytes

SALES_CALC_HEADERS = [
    "Mã hóa đơn",
    "Thời gian",
    "Tên khách hàng",
    "Mã hàng",
    "Số lượng",
    "Đơn giá",
    "Thành tiền",
    "% VAT",
    "Tiền thuế",
]


def vat_header_name() -> str:
    return "Tiền thuế"


def build_vat_mismatch_workbook(path: Path, *, row_count: int = 1000, seed: int = 0) -> dict[int, str]:
    """Each data row has intentional VAT mismatch; returns excel_row -> expected cell."""
    rng = random.Random(seed)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Stress VAT pinpoint"])
    ws.append(SALES_CALC_HEADERS)
    expected_cells: dict[int, str] = {}
    vat_col = SALES_CALC_HEADERS.index(vat_header_name())

    for i in range(row_count):
        excel_row = i + 3
        qty = (i % 9) + 1
        price = 10000 + (i % 50) * 100
        line = qty * price
        rate = 0.1
        wrong_vat = int(line * rate) + 2 + (i % 5)
        ws.append(
            [
                f"HD{i:05d}",
                f"{(i % 28) + 1:02d}/12/2025",
                f"Khách {i}",
                f"SP{i:04d}",
                qty,
                price,
                line,
                10,
                wrong_vat,
            ]
        )
        expected_cells[excel_row] = excel_cell(excel_row, vat_col)

    wb.save(path)
    return expected_cells


def build_multisheet_workbook(path: Path, scenario: int) -> tuple[str, int]:
    """Returns (expected_sheet_name, expected_data_rows)."""
    rng = random.Random(scenario)
    wb = openpyxl.Workbook()
    data_name = f"DuLieu_{scenario % 97}"
    decoy_name = f"TomTat_{scenario % 89}"

    decoy = wb.active
    decoy.title = decoy_name
    decoy.append(["Báo cáo tổng hợp — không phải dữ liệu nhập"])
    decoy.append(["Chỉ tiêu", "Giá trị"])
    decoy.append(["Doanh thu", scenario * 1000])

    data = wb.create_sheet(data_name)
    data.append(SALES_CALC_HEADERS)
    rows = (scenario % 5) + 3
    for i in range(rows):
        data.append(
            [
                f"HD{scenario}-{i}",
                "01/01/2026",
                "Khách",
                f"M{i}",
                1,
                1000,
                1000,
                0,
                0,
            ]
        )
    wb.save(path)
    return data_name, rows


def build_valid_xls(path: Path, rows: int = 3) -> None:
    wb = xlwt.Workbook()
    sh = wb.add_sheet("Sheet1")
    sh.write(0, 0, "Mã hóa đơn")
    sh.write(0, 1, "Thời gian")
    sh.write(0, 2, "Tên khách hàng")
    sh.write(0, 3, "Mã hàng")
    sh.write(0, 4, "Số lượng")
    sh.write(0, 5, "Đơn giá")
    for r in range(rows):
        sh.write(r + 1, 0, f"HD{r}")
        sh.write(r + 1, 1, "01/01/2026")
        sh.write(r + 1, 2, "K")
        sh.write(r + 1, 3, f"S{r}")
        sh.write(r + 1, 4, 1)
        sh.write(r + 1, 5, 100)
    wb.save(str(path))


def corrupt_xls_bytes(scenario: int, *, valid_base: bytes | None = None) -> bytes:
    rng = random.Random(scenario)
    mode = scenario % 6
    if mode == 0:
        return b"NOT_XLS" + bytes([scenario % 256] * 40)
    if mode == 1 and valid_base:
        return valid_base[: max(20, len(valid_base) // (2 + scenario % 5))]
    if mode == 2:
        return b"\xd0\xcf\x11\xe0" + bytes(rng.getrandbits(8) for _ in range(64))
    if mode == 3 and valid_base:
        mutated = bytearray(valid_base)
        for _ in range(3 + scenario % 10):
            if len(mutated) > 10:
                mutated[rng.randint(0, len(mutated) - 1)] = rng.randint(0, 255)
        return bytes(mutated)
    if mode == 4:
        return b""
    if valid_base:
        return valid_base[:8]
    return b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 32


PDF_ASCII_HEADERS = [
    "invoice",
    "date",
    "customer",
    "sku",
    "qty",
    "price",
    "line",
    "vat_pct",
    "vat_amt",
]


def build_pdf_table_bytes(scenario: int) -> bytes:
    try:
        from fpdf import FPDF
    except ImportError:
        return b"%PDF-1.4\n% invalid without fpdf2\n"

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", size=8)
    sep = "  "
    pdf.cell(0, 5, sep.join(PDF_ASCII_HEADERS), new_x="LMARGIN", new_y="NEXT")
    for i in range(3):
        row = sep.join(
            [
                f"HD{scenario}-{i}",
                "01/01/2026",
                "Cust",
                f"M{i}",
                "1",
                "1000",
                "1000",
                "10",
                str(100 + (i % 3)),
            ]
        )
        pdf.cell(0, 5, row, new_x="LMARGIN", new_y="NEXT")
    out = pdf.output()
    if isinstance(out, bytes):
        return out
    if isinstance(out, bytearray):
        return bytes(out)
    return str(out).encode("latin-1")


def corrupt_pdf_bytes(scenario: int, *, valid: bytes | None = None) -> bytes:
    rng = random.Random(scenario)
    mode = scenario % 5
    if mode == 0:
        return b"PDF%" + bytes([scenario % 200] * 30)
    if mode == 1:
        return b"%PDF-1.4\n" + b"x" * (20 + scenario % 80)
    if mode == 2 and valid:
        return valid[: max(50, len(valid) // 3)]
    if mode == 3 and valid:
        return valid + b"\n% corrupted trailer"
    return b"%PDF-1.4\n%%EOF\n"


def build_ocr_sidecar_text(scenario: int) -> str:
    lines = ["\t".join(SALES_CALC_HEADERS)]
    for i in range(2):
        lines.append(
            "\t".join(
                [
                    f"OCR{scenario}-{i}",
                    "02/01/2026",
                    "Khách OCR",
                    f"SKU{i}",
                    "2",
                    "5000",
                    "10000",
                    "10",
                    "1000",
                ]
            )
        )
    return "\n".join(lines) + "\n"


def pdf_classify_scenario(scenario: int) -> str:
    valid = build_pdf_table_bytes(scenario) if scenario % 4 == 0 else None
    payload = corrupt_pdf_bytes(scenario, valid=valid) if scenario % 4 != 0 else valid
    assert payload is not None
    return classify_pdf_bytes(payload)
