from __future__ import annotations

from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app


ROOT = Path(__file__).resolve().parents[1]
SAMPLES = ROOT / "fixtures" / "samples"
client = TestClient(app)


def _safe_sales_defaults(defaults: dict) -> dict:
    safe = dict(defaults)
    safe.update(
        {
            "Hình thức bán hàng": "Bán hàng hóa trong nước",
            "Phương thức thanh toán": "Chưa thu tiền",
            "TK Tiền/Chi phí/Nợ (*)": "131",
            "TK Doanh thu/Có (*)": "5111",
        }
    )
    return safe


def _analyze_confirm_sales(
    tmp_path, monkeypatch, *, blank_item_code: bool = False
) -> tuple[str, str, dict]:
    monkeypatch.setenv("MAPPING_DB_PATH", str(tmp_path / "profiles.sqlite"))
    monkeypatch.setenv("AI_PROVIDER", "disabled")

    content = (SAMPLES / "raw_sales_sample.xlsx").read_bytes()
    if blank_item_code:
        workbook = load_workbook(BytesIO(content))
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        item_code_column = headers.index("Mã hàng") + 1
        sheet.cell(row=2, column=item_code_column).value = ""
        output = BytesIO()
        workbook.save(output)
        content = output.getvalue()

    analyze = client.post(
        "/api/v1/uploads/analyze",
        data={"target_template_id": "bsn_sales"},
        files={
            "file": (
                "raw_sales_sample.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert analyze.status_code == 200
    payload = analyze.json()
    suggestion = payload["mapping_suggestion"]
    defaults = _safe_sales_defaults(suggestion["defaults"])

    confirm = client.post(
        "/api/v1/mappings/confirm",
        json={
            "upload_id": payload["upload_id"],
            "target_template_id": payload["target_template_id"],
            "mapping": suggestion["mapping"],
            "defaults": defaults,
            "formulas": suggestion["formulas"],
            "profile_name": "Readiness test",
        },
    )
    assert confirm.status_code == 200
    return payload["upload_id"], confirm.json()["profile_id"], {
        "mapping": suggestion["mapping"],
        "defaults": defaults,
        "formulas": suggestion["formulas"],
    }


def _valid_sales_row(**overrides):
    row = {
        "Hình thức bán hàng": "Bán hàng hóa trong nước",
        "Phương thức thanh toán": "Chưa thu tiền",
        "Ngày hạch toán (*)": "01/01/2026",
        "Ngày chứng từ (*)": "01/01/2026",
        "Số chứng từ (*)": "HD001",
        "Mã hàng (*)": "SP01",
        "Tên hàng": "Hàng test",
        "TK Tiền/Chi phí/Nợ (*)": "131",
        "TK Doanh thu/Có (*)": "5111",
        "Số lượng": 2,
        "Đơn giá": 1000,
        "Thành tiền": 2000,
    }
    row.update(overrides)
    return row


def test_readiness_api_returns_blocker_for_blank_required_value(tmp_path, monkeypatch):
    upload_id, _profile_id, confirmed = _analyze_confirm_sales(
        tmp_path, monkeypatch, blank_item_code=True
    )

    response = client.post(
        "/api/v1/mappings/readiness",
        json={
            "upload_id": upload_id,
            "target_template_id": "bsn_sales",
            **confirmed,
            "rows": [_valid_sales_row(**{"Mã hàng (*)": ""})],
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "blocked"
    assert any(issue["code"] == "required_value_blank" for issue in payload["issues"])


def test_export_blocks_when_readiness_has_blocker(tmp_path, monkeypatch):
    upload_id, profile_id, _confirmed = _analyze_confirm_sales(
        tmp_path, monkeypatch, blank_item_code=True
    )

    response = client.post(
        "/api/v1/conversions/export",
        json={
            "upload_id": upload_id,
            "profile_id": profile_id,
            "rows": [_valid_sales_row()],
            "acknowledge_warnings": True,
        },
    )

    assert response.status_code == 422
    payload = response.json()
    assert payload["status"] == "blocked"
    assert any(issue["code"] == "required_value_blank" for issue in payload["issues"])


def test_export_blocks_warning_without_acknowledgement(tmp_path, monkeypatch):
    upload_id, profile_id, _confirmed = _analyze_confirm_sales(tmp_path, monkeypatch)

    response = client.post(
        "/api/v1/conversions/export",
        json={
            "upload_id": upload_id,
            "profile_id": profile_id,
            "rows": [_valid_sales_row(**{"Mã hàng (*)": "Hàng test"})],
        },
    )

    assert response.status_code == 422
    payload = response.json()
    assert payload["status"] == "needs_review"
    assert any(issue["code"] == "master_data_not_checked" for issue in payload["issues"])


def test_export_allows_warning_when_acknowledged(tmp_path, monkeypatch):
    upload_id, profile_id, _confirmed = _analyze_confirm_sales(tmp_path, monkeypatch)

    response = client.post(
        "/api/v1/conversions/export",
        json={
            "upload_id": upload_id,
            "profile_id": profile_id,
            "rows": [_valid_sales_row(**{"Mã hàng (*)": "Hàng test"})],
            "acknowledge_warnings": True,
        },
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.ms-excel")
