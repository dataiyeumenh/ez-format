"""
Extreme E2E: full user journey per conversion type.
validate → preview → export roundtrip + direct convert download.
"""

from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest
import xlrd
from fastapi.testclient import TestClient

from app.conversion_types import CONVERSION_TYPES
from app.main import app


ROOT = Path(__file__).resolve().parents[1]
SAMPLES = ROOT / "fixtures" / "samples"
client = TestClient(app)


@pytest.fixture(autouse=True)
def enable_legacy_export_for_legacy_api_tests(monkeypatch):
    monkeypatch.setenv("ALLOW_LEGACY_ROW_EXPORT", "true")

SALES_HEADERS = [
    "Mã hóa đơn",
    "Thời gian",
    "Tên khách hàng",
    "Mã hàng",
    "Số lượng",
    "Đơn giá",
    "Thành tiền",
]

# Headers must normalize to FIELD_ALIASES (see app/field_detection.py).
PURCHASE_HEADERS = [
    "Số phiếu nhập",
    "Ngày chứng từ",
    "Mã NCC",
    "Tên NCC",
    "Mã hàng",
    "Số lượng",
    "Đơn giá",
]


def _write_sales_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(SALES_HEADERS)
    ws.append(["HD-E2E-001", "25/12/2025", "Khách E2E", "SP-E2E", 2, 50000, 100000])
    wb.save(path)


def _write_purchase_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(PURCHASE_HEADERS)
    ws.append(["PN-E2E-001", "02/01/2026", "NCC01", "Nhà cung cấp E2E", "MH-E2E", 3, 20000])
    wb.save(path)


def _kind_path(conversion_type: str, tmp_path: Path) -> Path:
    kind = CONVERSION_TYPES[conversion_type].kind
    if kind.startswith("sales"):
        p = tmp_path / f"{conversion_type}_sales.xlsx"
        _write_sales_xlsx(p)
        return p
    p = tmp_path / f"{conversion_type}_purchase.xlsx"
    _write_purchase_xlsx(p)
    return p


def test_health_and_conversion_types_catalog():
    health = client.get("/healthz").json()
    assert health == {
        "status": "ok",
        "capabilities": {"converter": True, "operations": True},
    }
    payload = client.get("/api/v1/conversion-types").json()
    ids = {item["id"] for item in payload["items"]}
    assert ids == set(CONVERSION_TYPES.keys())


def test_raw_sales_sample_full_journey_all_sales_types(tmp_path):
    path = SAMPLES / "raw_sales_sample.xlsx"
    assert path.exists(), "missing fixtures/samples/raw_sales_sample.xlsx"
    for ct in ("bsn_sales", "sales_goods", "sales_service"):
        _journey_for_file(path, ct, tmp_path)


def test_e2e_validate_preview_export_roundtrip_all_types(tmp_path):
    for conversion_type in CONVERSION_TYPES:
        input_path = _kind_path(conversion_type, tmp_path)
        _journey_for_file(input_path, conversion_type, tmp_path)


def test_e2e_direct_convert_download_all_types(tmp_path):
    for conversion_type in CONVERSION_TYPES:
        input_path = _kind_path(conversion_type, tmp_path)
        with input_path.open("rb") as handle:
            response = client.post(
                "/api/v1/conversions",
                data={"conversion_type": conversion_type},
                files={
                    "file": (
                        input_path.name,
                        handle,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        assert response.status_code == 200, (conversion_type, response.text[:500])
        assert response.headers["content-type"].startswith("application/vnd.ms-excel")
        assert len(response.content) > 1000


def test_purchase_common_headers_validate_preview_export(tmp_path):
    """Real-world headers: Số PN, Ngày nhập (not only formal MISA names)."""
    path = tmp_path / "purchase_pn.xlsx"
    _write_purchase_xlsx_common(path)
    _journey_for_file(path, "purchase_goods", tmp_path)


def _write_purchase_xlsx_common(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(
        [
            "Số PN",
            "Ngày nhập",
            "Mã NCC",
            "Tên NCC",
            "Mã hàng",
            "Số lượng",
            "Đơn giá",
        ]
    )
    ws.append(["PN-E2E-002", "03/01/2026", "NCC02", "NCC E2E", "MH2", 1, 15000])
    wb.save(path)


def test_e2e_rejects_invalid_file_type():
    response = client.post(
        "/api/v1/conversions",
        data={"conversion_type": "bsn_sales"},
        files={"file": ("bad.txt", b"not excel", "text/plain")},
    )
    assert response.status_code == 415


def test_e2e_preview_rejects_empty_workbook(tmp_path):
    path = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    wb.save(path)
    with path.open("rb") as handle:
        response = client.post(
            "/api/v1/conversions/preview",
            data={"conversion_type": "bsn_sales"},
            files={"file": (path.name, handle, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert response.status_code == 422


def _journey_for_file(input_path: Path, conversion_type: str, tmp_path: Path) -> None:
    # 1) validate
    with input_path.open("rb") as handle:
        validate_res = client.post(
            "/api/v1/conversions/validate",
            data={"conversion_type": conversion_type},
            files={
                "file": (
                    input_path.name,
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert validate_res.status_code == 200, validate_res.text
    report = validate_res.json()
    assert report["ok"] is True, report

    # 2) preview
    with input_path.open("rb") as handle:
        preview_res = client.post(
            "/api/v1/conversions/preview",
            data={"conversion_type": conversion_type},
            files={
                "file": (
                    input_path.name,
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert preview_res.status_code == 200, preview_res.text
    preview = preview_res.json()
    headers = preview["headers"]
    rows = preview["rows"]
    assert headers
    assert rows
    assert preview["report"]["ok"] is True

    # 3) export edited rows
    export_res = client.post(
        "/api/v1/conversions/export",
        json={"conversion_type": conversion_type, "rows": rows},
    )
    assert export_res.status_code == 200, export_res.text
    assert export_res.headers["content-type"].startswith("application/vnd.ms-excel")
    out = tmp_path / f"{conversion_type}_e2e_out.xls"
    out.write_bytes(export_res.content)
    book = xlrd.open_workbook(str(out))
    sheet = book.sheet_by_index(0)
    assert sheet.nrows > 8
