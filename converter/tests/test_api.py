import json
from pathlib import Path

import openpyxl
import xlrd
from fastapi.testclient import TestClient

from app.main import app


ROOT = Path(__file__).resolve().parents[1]
SAMPLES = ROOT / "fixtures" / "samples"


client = TestClient(app)

WARNING_HEADERS = [
    "Mã hóa đơn",
    "Thời gian",
    "Tên khách hàng",
    "Mã hàng",
    "Số lượng",
    "Đơn giá",
    "Thành tiền",
    "Tổng tiền hàng",
    "Khách cần trả",
]

MESSY_SALES_HEADERS = [
    "Số HĐ bán lẻ",
    "Ngày bán",
    "Người mua hàng",
    "Mã SKU",
    "Tên mặt hàng",
    "SL bán",
    "Giá bán",
    "Tổng dòng",
]

MESSY_SALES_MAPPING = {
    "invoice": "Số HĐ bán lẻ",
    "date": "Ngày bán",
    "customer_name": "Người mua hàng",
    "item_code": "Mã SKU",
    "item_name": "Tên mặt hàng",
    "quantity": "SL bán",
    "unit_price": "Giá bán",
    "line_amount": "Tổng dòng",
}

ACCOUNTING_HEADERS = [
    "Số HĐ",
    "Ngày CT",
    "Khách hàng",
    "Mã hàng",
    "Số lượng",
    "Đơn giá",
    "Thành tiền",
    "VAT %",
    "TK Nợ",
    "TK doanh thu",
    "TK thuế bán",
    "TK giá vốn",
    "TK kho",
]

ACCOUNTING_MAPPING = {
    "invoice": "Số HĐ",
    "date": "Ngày CT",
    "customer_name": "Khách hàng",
    "item_code": "Mã hàng",
    "quantity": "Số lượng",
    "unit_price": "Đơn giá",
    "line_amount": "Thành tiền",
    "vat_rate": "VAT %",
    "debit_account": "TK Nợ",
    "revenue_account": "TK doanh thu",
    "vat_account": "TK thuế bán",
    "cogs_account": "TK giá vốn",
    "inventory_account": "TK kho",
}


def _write_warning_workbook(path: Path) -> Path:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(WARNING_HEADERS)
    sheet.append(["HD001", "2025-12-25", "Khách A", "SP001", 2, 50, 90, 90, 90])
    workbook.save(path)
    return path


def _write_duplicate_warning_workbook(path: Path) -> Path:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Mã hóa đơn", "Thời gian", "Tên khách hàng", "Mã hàng", "Số lượng", "Đơn giá"])
    sheet.append(["HD-DUP-001", "2026-01-02", "Khách A", "SP-DUP", 1, 100000])
    sheet.append(["HD-DUP-001", "2026-01-02", "Khách A", "SP-DUP", 1, 100000])
    workbook.save(path)
    return path


def _write_messy_sales_workbook(path: Path) -> Path:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["EzFormat messy sales fixture"])
    sheet.append(MESSY_SALES_HEADERS)
    sheet.append(["HD-MESS-001", "25/12/2025", "Khách lẻ A", "SKU-001", "Hàng A", 2, "1.000", 2000])
    workbook.save(path)
    return path


def _write_wrong_accounting_workbook(path: Path) -> Path:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(ACCOUNTING_HEADERS)
    sheet.append(
        [
            "HD-ACC-001",
            "02/01/2026",
            "Khách sai TK",
            "SKU-ACC-001",
            1,
            100000,
            100000,
            "10%",
            "331",
            "331",
            "1331",
            "1561",
            "632",
        ]
    )
    workbook.save(path)
    return path


def _write_purchase_mapping_workbook(path: Path) -> Path:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Số PN",
            "Ngày CT",
            "Mã NCC",
            "Tên NCC",
            "Mã hàng",
            "Số lượng",
            "Giảm giá mua",
            "Giá mua",
        ]
    )
    sheet.append(["PN-MAP-001", "02/01/2026", "NCC001", "NCC A", "SKU001", 1, 0, 100000])
    workbook.save(path)
    return path


def test_healthz():
    response = client.get("/healthz")

    assert response.status_code == 200
    assert response.json() == {"status": "ok"}


def test_conversion_types_endpoint():
    response = client.get("/api/v1/conversion-types")

    assert response.status_code == 200
    payload = response.json()
    assert {item["id"] for item in payload["items"]} == {
        "bsn_sales",
        "bsn_purchase",
        "sales_goods",
        "sales_service",
        "purchase_goods",
        "purchase_service",
    }


def test_validate_endpoint_returns_validation_report():
    with (SAMPLES / "raw_sales_sample.xlsx").open("rb") as handle:
        response = client.post(
            "/api/v1/conversions/validate",
            data={"conversion_type": "bsn_sales"},
            files={
                "file": (
                    "raw_sales_sample.xlsx",
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["ok"] is True
    assert payload["summary"]["input_rows"] == 1930
    assert payload["detected_columns"]["invoice"] == "Mã hóa đơn"


def test_convert_endpoint_returns_xls_file_and_cleans_temp_storage():
    with (SAMPLES / "raw_sales_sample.xlsx").open("rb") as handle:
        response = client.post(
            "/api/v1/conversions",
            data={"conversion_type": "bsn_sales"},
            files={
                "file": (
                    "raw_sales_sample.xlsx",
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.ms-excel")
    assert response.headers["content-disposition"].endswith('filename="bsn_sales_import.xls"')

    workbook_path = ROOT / "tests" / "_tmp_api_response.xls"
    workbook_path.write_bytes(response.content)
    try:
        sheet = xlrd.open_workbook(str(workbook_path)).sheet_by_index(0)
        assert sheet.cell_value(8, 7) == "HD046178"
        assert sheet.cell_value(8, 25) == "SP094030"
    finally:
        workbook_path.unlink(missing_ok=True)

    temp_root = ROOT / ".tmp"
    assert not temp_root.exists() or not any(temp_root.iterdir())


def test_validate_endpoint_returns_calculation_warnings(tmp_path):
    input_path = _write_warning_workbook(tmp_path / "warning.xlsx")

    with input_path.open("rb") as handle:
        response = client.post(
            "/api/v1/conversions/validate",
            data={"conversion_type": "bsn_sales"},
            files={
                "file": (
                    "warning.xlsx",
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    payload = response.json()
    warning = next(
        warning
        for warning in payload["warnings"]
        if warning["code"] == "calculation_line_amount_mismatch"
    )
    assert warning["row"] == 2
    assert warning["invoice"] == "HD001"
    assert warning["expected"] == 100
    assert warning["actual"] == 90
    assert warning["tolerance"] == 1


def test_convert_endpoint_blocks_calculation_warnings_until_override(tmp_path):
    input_path = _write_warning_workbook(tmp_path / "warning.xlsx")

    with input_path.open("rb") as handle:
        blocked = client.post(
            "/api/v1/conversions",
            data={"conversion_type": "bsn_sales"},
            files={
                "file": (
                    "warning.xlsx",
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert blocked.status_code == 422
    blocked_payload = blocked.json()
    assert blocked_payload["ok"] is True
    assert any(
        warning["code"] == "calculation_line_amount_mismatch"
        for warning in blocked_payload["warnings"]
    )

    with input_path.open("rb") as handle:
        allowed = client.post(
            "/api/v1/conversions",
            data={
                "conversion_type": "bsn_sales",
                "options": json.dumps({"allow_calculation_warnings": True}),
            },
            files={
                "file": (
                    "warning.xlsx",
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert allowed.status_code == 200
    assert allowed.headers["content-type"].startswith("application/vnd.ms-excel")


def test_validate_endpoint_accepts_safe_column_mapping_override(tmp_path):
    input_path = _write_messy_sales_workbook(tmp_path / "messy.xlsx")

    with input_path.open("rb") as handle:
        response = client.post(
            "/api/v1/conversions/validate",
            data={
                "conversion_type": "bsn_sales",
                "options": json.dumps({"column_mapping": MESSY_SALES_MAPPING}),
            },
            files={
                "file": (
                    "messy.xlsx",
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["ok"] is True
    assert payload["detected_columns"]["invoice"] == "Số HĐ bán lẻ"
    assert payload["detected_columns"]["unit_price"] == "Giá bán"


def test_validate_endpoint_rejects_invalid_column_mapping(tmp_path):
    input_path = _write_messy_sales_workbook(tmp_path / "messy.xlsx")

    with input_path.open("rb") as handle:
        response = client.post(
            "/api/v1/conversions/validate",
            data={
                "conversion_type": "bsn_sales",
                "options": json.dumps(
                    {
                        "column_mapping": {
                            "not_a_semantic_field": "Số HĐ bán lẻ",
                            "invoice": "Không tồn tại",
                        }
                    }
                ),
            },
            files={
                "file": (
                    "messy.xlsx",
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["ok"] is False
    assert {
        "invalid_column_mapping_field",
        "missing_column_mapping_header",
    }.issubset({error["code"] for error in payload["errors"]})


def test_ai_mapping_suggestions_mock_provider(tmp_path, monkeypatch):
    monkeypatch.setenv("AI_MODE", "mock")
    input_path = _write_messy_sales_workbook(tmp_path / "messy.xlsx")

    with input_path.open("rb") as handle:
        response = client.post(
            "/api/v1/ai/mapping-suggestions",
            data={"conversion_type": "bsn_sales"},
            files={
                "file": (
                    "messy.xlsx",
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["ok"] is True
    assert payload["provider"] == "mock"
    assert payload["suggested_mapping"]["invoice"] == "Số HĐ bán lẻ"
    assert payload["suggested_mapping"]["quantity"] == "SL bán"
    assert payload["missing_fields"] == []


def test_ai_mapping_suggestions_disabled_provider(tmp_path, monkeypatch):
    monkeypatch.setenv("AI_MODE", "disabled")
    input_path = _write_messy_sales_workbook(tmp_path / "messy.xlsx")

    with input_path.open("rb") as handle:
        response = client.post(
            "/api/v1/ai/mapping-suggestions",
            data={"conversion_type": "bsn_sales"},
            files={
                "file": (
                    "messy.xlsx",
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["ok"] is False
    assert payload["provider"] == "disabled"
    assert payload["errors"]


def test_ai_mapping_suggestions_do_not_map_purchase_discount_as_unit_price(tmp_path, monkeypatch):
    monkeypatch.setenv("AI_MODE", "mock")
    input_path = _write_purchase_mapping_workbook(tmp_path / "purchase_mapping.xlsx")

    with input_path.open("rb") as handle:
        response = client.post(
            "/api/v1/ai/mapping-suggestions",
            data={"conversion_type": "purchase_goods"},
            files={
                "file": (
                    "purchase_mapping.xlsx",
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["suggested_mapping"]["unit_price"] == "Giá mua"
    assert payload["suggested_mapping"]["discount_amount"] == "Giảm giá mua"


def test_ai_explain_validation_mock_provider_does_not_change_report(monkeypatch):
    monkeypatch.setenv("AI_MODE", "mock")
    report = {
        "ok": True,
        "summary": {"input_rows": 1, "output_rows": 1, "error_count": 0, "warning_count": 1},
        "errors": [],
        "warnings": [
            {
                "row": 2,
                "field": "line_amount",
                "code": "calculation_line_amount_mismatch",
                "message": "Line amount mismatch.",
                "invoice": "HD001",
                "expected": 100,
                "actual": 90,
                "delta": -10,
                "tolerance": 1,
            }
        ],
        "detected_columns": {"invoice": "Số HĐ"},
    }

    response = client.post("/api/v1/ai/explain-validation", json=report)

    assert response.status_code == 200
    payload = response.json()
    assert payload["ok"] is True
    assert payload["provider"] == "mock"
    assert payload["explanations"][0]["code"] == "calculation_line_amount_mismatch"
    assert report["warnings"][0]["actual"] == 90


def test_ai_error_check_returns_accounting_issues(tmp_path, monkeypatch):
    monkeypatch.setenv("AI_MODE", "mock")
    input_path = _write_wrong_accounting_workbook(tmp_path / "wrong_accounts.xlsx")

    with input_path.open("rb") as handle:
        response = client.post(
            "/api/v1/ai/error-check",
            data={
                "conversion_type": "sales_goods",
                "options": json.dumps(
                    {
                        "column_mapping": ACCOUNTING_MAPPING,
                        "accounting_profile": "tt99_2026",
                    }
                ),
            },
            files={
                "file": (
                    "wrong_accounts.xlsx",
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["tool_name"] == "Kiểm tra lỗi bằng AI"
    assert payload["ok"] is False
    assert payload["accounting_profile"] == "tt99_2026"
    assert payload["summary"]["accounting_issue_count"] >= 5
    codes = {issue["code"] for issue in payload["issues"]}
    assert {
        "accounting_wrong_sales_debit_account",
        "accounting_wrong_sales_revenue_account",
        "accounting_wrong_output_vat_account",
        "accounting_wrong_cogs_account",
        "accounting_wrong_sales_inventory_account",
    }.issubset(codes)
    assert payload["ai_explanation"]


def test_ai_error_check_works_when_ai_disabled(tmp_path, monkeypatch):
    monkeypatch.setenv("AI_MODE", "disabled")
    input_path = _write_wrong_accounting_workbook(tmp_path / "wrong_accounts.xlsx")

    with input_path.open("rb") as handle:
        response = client.post(
            "/api/v1/ai/error-check",
            data={
                "conversion_type": "sales_goods",
                "options": json.dumps({"column_mapping": ACCOUNTING_MAPPING}),
            },
            files={
                "file": (
                    "wrong_accounts.xlsx",
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["ok"] is False
    assert payload["summary"]["accounting_issue_count"] >= 5
    assert payload["ai_explanation"].startswith("AI disabled")


def test_ai_error_check_strict_blocks_unreviewed_warnings(tmp_path, monkeypatch):
    monkeypatch.setenv("AI_MODE", "mock")
    input_path = _write_warning_workbook(tmp_path / "warning.xlsx")

    with input_path.open("rb") as handle:
        response = client.post(
            "/api/v1/ai/error-check",
            data={
                "conversion_type": "bsn_sales",
                "options": json.dumps({"strict": True}),
            },
            files={
                "file": (
                    "warning.xlsx",
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["strict_blocked"] is True
    assert "calculation_line_amount_mismatch" in {
        issue["code"] for issue in payload["blocking_issues"]
    }
    assert payload["remediation"]


def test_convert_strict_blocks_accounting_errors(tmp_path, monkeypatch):
    monkeypatch.setenv("AI_MODE", "mock")
    input_path = _write_wrong_accounting_workbook(tmp_path / "wrong_accounts.xlsx")

    with input_path.open("rb") as handle:
        response = client.post(
            "/api/v1/conversions",
            data={
                "conversion_type": "sales_goods",
                "options": json.dumps(
                    {
                        "strict": True,
                        "column_mapping": ACCOUNTING_MAPPING,
                        "accounting_profile": "tt99_2026",
                    }
                ),
            },
            files={
                "file": (
                    "wrong_accounts.xlsx",
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 422
    payload = response.json()
    assert payload["strict_blocked"] is True
    assert {
        "accounting_wrong_sales_debit_account",
        "accounting_wrong_sales_revenue_account",
    }.issubset({issue["code"] for issue in payload["blocking_issues"]})


def test_convert_strict_allows_reviewed_non_error_warning(tmp_path, monkeypatch):
    monkeypatch.setenv("AI_MODE", "mock")
    input_path = _write_duplicate_warning_workbook(tmp_path / "duplicate.xlsx")

    with input_path.open("rb") as handle:
        response = client.post(
            "/api/v1/conversions",
            data={
                "conversion_type": "bsn_sales",
                "options": json.dumps(
                    {
                        "strict": True,
                        "reviewed_issue_codes": ["duplicate_invoice_item"],
                    }
                ),
            },
            files={
                "file": (
                    "duplicate.xlsx",
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.ms-excel")


def test_convert_endpoint_rejects_invalid_file_type():
    response = client.post(
        "/api/v1/conversions",
        data={"conversion_type": "bsn_sales"},
        files={"file": ("input.txt", b"not excel", "text/plain")},
    )

    assert response.status_code == 415
    assert response.json()["detail"] == "Only .xls and .xlsx files are supported."


def test_preview_endpoint_returns_json_rows(tmp_path):
    input_path = _write_messy_sales_workbook(tmp_path / "messy.xlsx")
    options = json.dumps({"column_mapping": MESSY_SALES_MAPPING})

    with input_path.open("rb") as handle:
        response = client.post(
            "/api/v1/conversions/preview",
            data={"conversion_type": "bsn_sales", "options": options},
            files={
                "file": (
                    "messy.xlsx",
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["report"]["ok"] is True
    assert payload["headers"]
    assert len(payload["rows"]) == 1
    assert payload["rows"][0]["Số chứng từ (*)"] == "HD-MESS-001"


def test_export_rows_endpoint_returns_xls(tmp_path):
    input_path = _write_messy_sales_workbook(tmp_path / "messy.xlsx")
    options = json.dumps({"column_mapping": MESSY_SALES_MAPPING})

    with input_path.open("rb") as handle:
        preview = client.post(
            "/api/v1/conversions/preview",
            data={"conversion_type": "bsn_sales", "options": options},
            files={
                "file": (
                    "messy.xlsx",
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    rows = preview.json()["rows"]
    response = client.post(
        "/api/v1/conversions/export",
        json={
            "conversion_type": "bsn_sales",
            "rows": rows,
        },
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.ms-excel")
    assert "bsn_sales_import.xls" in response.headers["content-disposition"]
