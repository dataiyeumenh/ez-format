from __future__ import annotations

from io import BytesIO
from types import SimpleNamespace

import openpyxl
import pytest

from app.excel_io import InputTable
from app.mapping_profile_v2 import (
    MappingProfileV2,
    MappingProfileV2Error,
    MappingProfileV2Match,
    build_profile_identity,
    confirm_mapping_profile_v2,
    mapping_profile_v2_enabled,
    match_mapping_profile_v2,
    record_confirmed_export_v2,
)
from app.misa_workflow import analyze_upload, confirm_mapping, export_confirmed_profile


class FakeResponse:
    status_code = 200

    def json(self):
        return {
            "success": True,
            "match": {
                "tier": "exact",
                "driftFields": [],
                "approvalState": "approved",
                "approvalAppliesToMatch": True,
                "riskFlags": ["vat"],
                "approvedRiskFlags": ["vat"],
                "unapprovedRiskFlags": [],
                "canSuggest": True,
                "requiresPreview": True,
                "profile": {
                    "id": "profile-v2",
                    "version": 3,
                    "status": "active",
                    "ownerScope": "workspace:workspace-1",
                    "mapping": {"Ma NCC": "Mã nhà cung cấp"},
                    "defaults": {},
                    "formulas": {},
                    "riskFlags": ["vat"],
                    "stateHash": "profile-state-hash",
                    "approvedBy": "approver-1",
                },
            },
        }


def _table(extra_header: str | None = None) -> InputTable:
    headers = ["Ma NCC", "Ngay hoa don", "Tong tien"]
    if extra_header:
        headers.append(extra_header)
    return InputTable(
        headers=headers,
        rows=[
            {"Ma NCC": "NCC01", "Ngay hoa don": "01/07/2026", "Tong tien": 1000},
            {"Ma NCC": "NCC02", "Ngay hoa don": "02/07/2026", "Tong tien": 2000},
        ],
        sheet_name="Mua vao",
        header_row_index=0,
    )


def test_profile_identity_changes_when_header_or_data_shape_drifts():
    original = build_profile_identity(
        _table(),
        target_template_id="misa_purchase_domestic",
        target_template_version="template-v1",
    )
    header_drift = build_profile_identity(
        _table("Ghi chu"),
        target_template_id="misa_purchase_domestic",
        target_template_version="template-v1",
    )
    shape_drift_table = _table()
    shape_drift_table.rows[0]["Tong tien"] = "khong phai so"
    shape_drift_table.rows[1]["Tong tien"] = "khong phai so"
    shape_drift = build_profile_identity(
        shape_drift_table,
        target_template_id="misa_purchase_domestic",
        target_template_version="template-v1",
    )

    assert original.normalized_header_fingerprint != header_drift.normalized_header_fingerprint
    assert original.data_shape_fingerprint != shape_drift.data_shape_fingerprint
    assert original.document_type == "purchase"


def test_mapping_profile_v2_flag_defaults_false(monkeypatch):
    monkeypatch.delenv("FEATURE_MAPPING_PROFILE_V2", raising=False)

    assert mapping_profile_v2_enabled() is False


def test_match_sends_complete_identity_and_parses_exact_profile(monkeypatch):
    captured = {}

    def request(method, url, **kwargs):
        captured.update({"method": method, "url": url, **kwargs})
        return FakeResponse()

    monkeypatch.setenv("FEATURE_MAPPING_PROFILE_V2", "true")
    monkeypatch.setenv("NODE_INTERNAL_API_URL", "http://node/api/internal")
    monkeypatch.setattr("app.mapping_profile_v2.httpx.request", request)
    identity = build_profile_identity(
        _table(),
        target_template_id="misa_purchase_domestic",
        target_template_version="template-v1",
        source_family="invoice-puller",
    )

    result = match_mapping_profile_v2("context-token", identity)

    assert result is not None
    assert result.match_tier == "exact"
    assert result.approval_state == "approved"
    assert result.approval_applies_to_match is True
    assert result.approved_risk_flags == ("vat",)
    assert result.unapproved_risk_flags == ()
    assert result.can_suggest is True
    assert result.requires_preview is True
    assert result.profile.version == 3
    assert result.profile.state_hash == "profile-state-hash"
    assert captured["json"]["sourceFamily"] == "invoice-puller"
    assert captured["json"]["headerFingerprint"]
    assert "normalizedHeaderFingerprint" not in captured["json"]
    assert captured["json"]["dataShapeFingerprint"]


def test_confirmed_export_sends_stable_idempotency_contract(monkeypatch):
    captured = {}

    def request(method, url, **kwargs):
        captured.update({"method": method, "url": url, **kwargs})
        return FakeResponse()

    monkeypatch.setenv("FEATURE_MAPPING_PROFILE_V2", "true")
    monkeypatch.setenv("NODE_INTERNAL_API_URL", "http://node/api/internal")
    monkeypatch.setattr("app.mapping_profile_v2.httpx.request", request)

    record_confirmed_export_v2(
        "context-token",
        profile_id="profile-v2",
        version=3,
        upload_id="upload-1",
        state_hash="state-hash",
    )

    body = captured["json"]
    assert captured["url"].endswith("/mapping-profiles/v2/profile-v2/confirmed-export")
    assert body["version"] == 3
    assert body["stateHash"] == "state-hash"
    assert len(body["exportId"]) == 64

    first_export_id = body["exportId"]
    record_confirmed_export_v2(
        "context-token",
        profile_id="profile-v2",
        version=3,
        upload_id="upload-1",
        state_hash="state-hash",
    )
    assert captured["json"]["exportId"] == first_export_id


def test_confirm_mapping_profile_v2_sends_candidate_and_expected_version(monkeypatch):
    captured = {}

    class ConfirmResponse:
        status_code = 200

        def json(self):
            return {
                "success": True,
                "profile": {
                    "id": "profile-v2-next",
                    "version": 4,
                    "status": "active",
                    "ownerScope": "workspace:workspace-1",
                    "targetTemplateId": "bsn_sales",
                    "mapping": {"Mã hàng": "Mã hàng (*)"},
                    "defaults": {},
                    "formulas": {},
                    "riskFlags": [],
                    "stateHash": "next-state",
                },
            }

    def request(method, url, **kwargs):
        captured.update({"method": method, "url": url, **kwargs})
        return ConfirmResponse()

    monkeypatch.setenv("NODE_INTERNAL_API_URL", "http://node/api/internal")
    monkeypatch.setattr("app.mapping_profile_v2.httpx.request", request)
    result = confirm_mapping_profile_v2(
        "context-token",
        candidate_profile_id="profile-v2",
        source_signature_hash="source-hash",
        target_template_id="bsn_sales",
        mapping={"Mã hàng": "Mã hàng (*)"},
        defaults={},
        formulas={},
        expected_version=3,
    )

    assert result.id == "profile-v2-next"
    assert result.version == 4
    assert captured["url"].endswith("/mapping-profiles/v2/confirm")
    assert captured["json"]["candidate_profile_id"] == "profile-v2"
    assert captured["json"]["expected_version"] == 3


def test_analyze_prefers_exact_active_v2_and_does_not_read_v1(tmp_path, monkeypatch):
    import app.misa_workflow as workflow

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Mã hóa đơn", "Thời gian", "Mã hàng"])
    sheet.append(["HD001", "01/07/2026", "SP01"])
    buffer = BytesIO()
    workbook.save(buffer)
    monkeypatch.setenv("FEATURE_MAPPING_PROFILE_V2", "true")
    monkeypatch.setenv("OPERATION_SESSION_DIR", str(tmp_path / "sessions"))
    monkeypatch.setattr(workflow, "UPLOAD_ROOT", tmp_path / "uploads")
    monkeypatch.setattr(
        workflow,
        "_context_for_analyze",
        lambda _token: (
            {},
            "online",
            None,
            {
                "workspace_id": "workspace-1",
                "user_id": "user-1",
                "snapshot_set_hash": "s1",
            },
        ),
    )
    monkeypatch.setattr(
        workflow,
        "match_mapping_profile_v2",
        lambda *_args, **_kwargs: MappingProfileV2Match(
            match_tier="exact",
            approval_state="approved",
            approval_applies_to_match=True,
            approved_risk_flags=("vat",),
            can_suggest=True,
            requires_preview=True,
            profile=MappingProfileV2(
                id="v2-1",
                version=2,
                status="active",
                owner_scope="workspace:workspace-1",
                target_template_id="bsn_sales",
                mapping={
                    "Mã hóa đơn": "Số chứng từ (*)",
                    "Thời gian": ["Ngày hạch toán (*)", "Ngày chứng từ (*)"],
                    "Mã hàng": "Mã hàng (*)",
                },
                defaults={
                    "TK Tiền/Chi phí/Nợ (*)": "131",
                    "TK Doanh thu/Có (*)": "5111",
                },
                formulas={},
                risk_flags=["vat"],
                state_hash="profile-state-v2-1",
                approved_by="approver-1",
                confidence=1,
            ),
        ),
    )
    monkeypatch.setattr(
        workflow,
        "find_mapping_profile",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("V1 must not be read")),
    )

    result = analyze_upload(
        filename="raw.xlsx",
        content=buffer.getvalue(),
        requested_target_template_id="bsn_sales",
        conversion_context_token="context-token",
    )

    assert result["mapping_suggestion"]["source"] == "profile_v2"
    assert result["mapping_suggestion"]["profile_id"] == "v2-1"
    assert result["mapping_profile_v2"]["match_tier"] == "exact"
    assert result["mapping_profile_v2"]["mapping_source"] == "profile_v2"
    assert result["mapping_profile_v2"]["approval_state"] == "approved"
    assert result["mapping_profile_v2"]["approval_applies_to_match"] is True
    assert result["mapping_profile_v2"]["approved_risk_flags"] == ["vat"]
    assert result["mapping_profile_v2"]["unapproved_risk_flags"] == []
    assert result["mapping_profile_v2"]["requires_preview"] is True
    assert result["mapping_profile_v2"]["profile"]["version"] == 2


def test_analyze_exposes_compatible_v2_drift_without_applying_profile(
    tmp_path, monkeypatch
):
    import app.misa_workflow as workflow

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Mã hóa đơn", "Thời gian", "Mã hàng"])
    sheet.append(["HD001", "01/07/2026", "SP01"])
    buffer = BytesIO()
    workbook.save(buffer)
    monkeypatch.setenv("FEATURE_MAPPING_PROFILE_V2", "true")
    monkeypatch.setenv("OPERATION_SESSION_DIR", str(tmp_path / "sessions"))
    monkeypatch.setattr(workflow, "UPLOAD_ROOT", tmp_path / "uploads")
    monkeypatch.setattr(
        workflow,
        "_context_for_analyze",
        lambda _token: (
            {},
            "online",
            None,
            {
                "workspace_id": "workspace-1",
                "user_id": "user-1",
                "snapshot_set_hash": "s1",
            },
        ),
    )
    monkeypatch.setattr(
        workflow,
        "match_mapping_profile_v2",
        lambda *_args, **_kwargs: MappingProfileV2Match(
            match_tier="compatible",
            approval_state="approved",
            approval_applies_to_match=False,
            unapproved_risk_flags=("vat_rate",),
            can_suggest=False,
            requires_preview=True,
            profile=MappingProfileV2(
                id="v2-compatible",
                version=4,
                status="active",
                owner_scope="workspace:workspace-1",
                target_template_id="bsn_sales",
                mapping={"Mã hóa đơn": "Số chứng từ (*)"},
                defaults={},
                formulas={},
                risk_flags=["vat_rate"],
                state_hash="profile-state-v2-compatible",
                approved_by="approver-1",
                confidence=0.88,
                name="Nguồn hóa đơn",
                source_family="invoice-puller",
                document_type="sales",
                target_template_version="template-v4",
            ),
            warnings=("headerFingerprint",),
        ),
    )
    monkeypatch.setattr(
        workflow,
        "find_mapping_profile",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("V1 must not be read after a V2 drift match")
        ),
    )

    result = analyze_upload(
        filename="raw.xlsx",
        content=buffer.getvalue(),
        requested_target_template_id="bsn_sales",
        conversion_context_token="context-token",
    )

    assert result["mapping_suggestion"]["source"] == "heuristic"
    match = result["mapping_profile_v2"]
    assert match["match_tier"] == "compatible"
    assert match["mapping_source"] == "heuristic"
    assert match["approval_state"] == "approved"
    assert match["approval_applies_to_match"] is False
    assert match["approved_risk_flags"] == []
    assert match["unapproved_risk_flags"] == ["vat_rate"]
    assert match["can_suggest"] is False
    assert match["profile_id"] == "v2-compatible"
    assert match["drift"][0]["id"] == "headerFingerprint"
    assert match["risk_flags"] == ["vat_rate"]
    assert match["profile"]["name"] == "Nguồn hóa đơn"


def test_analyze_exposes_exact_unapproved_risk_without_using_candidate(
    tmp_path, monkeypatch
):
    import app.misa_workflow as workflow

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Mã hóa đơn", "Thời gian", "Mã hàng"])
    sheet.append(["HD001", "01/07/2026", "SP01"])
    buffer = BytesIO()
    workbook.save(buffer)
    monkeypatch.setenv("FEATURE_MAPPING_PROFILE_V2", "true")
    monkeypatch.setenv("OPERATION_SESSION_DIR", str(tmp_path / "sessions"))
    monkeypatch.setattr(workflow, "UPLOAD_ROOT", tmp_path / "uploads")
    monkeypatch.setattr(
        workflow,
        "_context_for_analyze",
        lambda _token: (
            {},
            "online",
            None,
            {
                "workspace_id": "workspace-1",
                "user_id": "user-1",
                "snapshot_set_hash": "s1",
            },
        ),
    )
    monkeypatch.setattr(
        workflow,
        "match_mapping_profile_v2",
        lambda *_args, **_kwargs: MappingProfileV2Match(
            match_tier="exact",
            approval_state="unapproved",
            approval_applies_to_match=False,
            unapproved_risk_flags=("vat",),
            can_suggest=False,
            requires_preview=True,
            profile=MappingProfileV2(
                id="v2-unapproved",
                version=2,
                status="active",
                owner_scope="workspace:workspace-1",
                target_template_id="bsn_sales",
                mapping={"Mã hóa đơn": "Số chứng từ (*)"},
                defaults={},
                formulas={},
                risk_flags=["vat"],
                state_hash="profile-state-unapproved",
                confidence=1,
            ),
        ),
    )
    monkeypatch.setattr(
        workflow,
        "find_mapping_profile",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("V1 must not hide an unapproved V2 candidate")
        ),
    )

    result = analyze_upload(
        filename="raw.xlsx",
        content=buffer.getvalue(),
        requested_target_template_id="bsn_sales",
        conversion_context_token="context-token",
    )

    assert result["mapping_suggestion"]["source"] == "heuristic"
    match = result["mapping_profile_v2"]
    assert match["match_tier"] == "exact"
    assert match["mapping_source"] == "heuristic"
    assert match["approval_state"] == "unapproved"
    assert match["approval_applies_to_match"] is False
    assert match["approved_risk_flags"] == []
    assert match["unapproved_risk_flags"] == ["vat"]
    assert match["can_suggest"] is False


def _prepare_v2_export(tmp_path, monkeypatch):
    import app.misa_workflow as workflow

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Mã hóa đơn", "Thời gian", "Mã hàng"])
    sheet.append(["HD001", "01/07/2026", "SP01"])
    buffer = BytesIO()
    workbook.save(buffer)
    profile = MappingProfileV2(
        id="v2-export",
        version=5,
        status="active",
        owner_scope="workspace:workspace-1",
        target_template_id="bsn_sales",
        mapping={
            "Mã hóa đơn": "Số chứng từ (*)",
            "Thời gian": ["Ngày hạch toán (*)", "Ngày chứng từ (*)"],
            "Mã hàng": "Mã hàng (*)",
        },
        defaults={
            "TK Tiền/Chi phí/Nợ (*)": "131",
            "TK Doanh thu/Có (*)": "5111",
        },
        formulas={},
        risk_flags=["vat"],
        state_hash="immutable-profile-state-hash",
        approved_by="approver-1",
        confidence=1,
    )
    match = MappingProfileV2Match(
        match_tier="exact",
        profile=profile,
        approval_state="approved",
        approval_applies_to_match=True,
        approved_risk_flags=("vat",),
        can_suggest=True,
        requires_preview=True,
    )
    monkeypatch.setenv("FEATURE_MAPPING_PROFILE_V2", "true")
    monkeypatch.setenv("OPERATION_SESSION_DIR", str(tmp_path / "sessions"))
    monkeypatch.setattr(workflow, "UPLOAD_ROOT", tmp_path / "uploads")
    monkeypatch.setattr(
        workflow,
        "_context_for_analyze",
        lambda _token: (
            {},
            "online",
            None,
            {
                "workspace_id": "workspace-1",
                "user_id": "user-1",
                "snapshot_set_hash": "s1",
                "conversion_run_id": "run-1",
            },
        ),
    )
    monkeypatch.setattr(workflow, "match_mapping_profile_v2", lambda *_args: match)
    monkeypatch.setattr(workflow, "find_mapping_profile", lambda *_args, **_kwargs: None)
    analyzed = analyze_upload(
        filename="raw.xlsx",
        content=buffer.getvalue(),
        requested_target_template_id="bsn_sales",
        conversion_context_token="context-token",
    )
    monkeypatch.setattr(workflow, "_context_for_upload", lambda *_args: ({}, "online", None))
    monkeypatch.setattr(
        workflow,
        "verify_conversion_context_token",
        lambda _token: {
            "owner_scope": "workspace:workspace-1",
            "workspace_id": "workspace-1",
            "user_id": "user-1",
            "snapshot_set_hash": "s1",
            "conversion_run_id": "run-1",
            "operation_session_id": analyzed["session"]["session_id"],
            "upload_id": analyzed["upload_id"],
            "target_template_id": "bsn_sales",
            "scopes": ["export"],
        },
    )
    monkeypatch.setattr(workflow, "get_mapping_profile_v2", lambda *_args: profile)
    monkeypatch.setattr(
        workflow,
        "resolve_master_data",
        lambda rows, *_args, **_kwargs: SimpleNamespace(rows=rows, resolutions=[]),
    )
    readiness = SimpleNamespace(summary=SimpleNamespace(blocker=0, warning=0))
    monkeypatch.setattr(workflow, "build_readiness_report", lambda *_args, **_kwargs: readiness)
    monkeypatch.setattr(
        workflow, "add_master_data_resolutions", lambda report, *_args, **_kwargs: report
    )
    monkeypatch.setattr(
        workflow,
        "write_xls_from_template",
        lambda _template, _rows, path: path.write_bytes(b"xls-output"),
    )
    return workflow, profile, analyzed


def test_export_confirmation_uses_profile_state_hash_and_records_success(
    tmp_path, monkeypatch
):
    workflow, profile, analyzed = _prepare_v2_export(tmp_path, monkeypatch)
    recorded = {"count": 0}

    def record(*_args, **kwargs):
        assert kwargs["state_hash"] == profile.state_hash
        recorded["count"] += 1

    monkeypatch.setattr(workflow, "record_confirmed_export_v2", record)
    session = analyzed["session"]

    content, _filename = export_confirmed_profile(
        analyzed["upload_id"],
        profile.id,
        acknowledge_warnings=True,
        conversion_context_token="context-token",
        session_id=session["session_id"],
        revision=session["active_revision"],
        state_hash=session["state_hash"],
    )

    trace = workflow._read_metadata(analyzed["upload_id"])[
        "mapping_profile_v2_confirmation"
    ]
    assert content == b"xls-output"
    assert recorded["count"] == 1
    assert trace["status"] == "recorded"
    assert trace["profile_state_hash"] == profile.state_hash
    assert trace["error"] is None


def test_confirm_mapping_preserves_active_v2_candidate_and_state_hash(tmp_path, monkeypatch):
    import app.misa_workflow as workflow

    monkeypatch.setenv("FEATURE_MAPPING_PROFILE_V2", "true")
    monkeypatch.setenv("OPERATION_SESSION_DIR", str(tmp_path / "sessions"))
    monkeypatch.setattr(workflow, "UPLOAD_ROOT", tmp_path / "uploads")
    upload_id, _ = workflow.save_upload("raw.xlsx", b"not-used")
    metadata = {
        "upload_id": upload_id,
        "filename": "raw.xlsx",
        "input_path": str(workflow._upload_dir(upload_id) / "input.xlsx"),
        "target_template_id": "bsn_sales",
        "signature": {
            "sheet_name": "Sheet",
            "header_row": 1,
            "row_count": 1,
            "headers": ["Mã hóa đơn"],
            "hash": "source-signature-hash",
        },
        "mapping_profile_v2_candidate": {
            "profile_id": "candidate-v2",
            "version": 1,
            "state_hash": "candidate-state-hash",
        },
        "conversion_context": None,
        "owner_scope": "user:user-1",
        "suggestion": {},
    }
    workflow._write_metadata(upload_id, metadata)
    monkeypatch.setattr(workflow, "_assert_operation_state", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(workflow, "_context_for_upload", lambda *_args: (None, "not_configured", None))
    monkeypatch.setattr(
        workflow,
        "get_misa_template",
        lambda _template_id: SimpleNamespace(headers=["Số chứng từ (*)"]),
    )
    monkeypatch.setattr(
        workflow,
        "confirm_mapping_profile_v2",
        lambda *_args, **_kwargs: {
            "profile_id": "confirmed-v2",
            "version": 2,
            "state_hash": "confirmed-state-hash",
            "status": "active",
        },
    )
    monkeypatch.setattr(
        workflow,
        "source_signature",
        lambda _table: SimpleNamespace(
            sheet_name="Sheet",
            header_row=1,
            row_count=1,
            headers=["Mã hóa đơn"],
            hash="source-signature-hash",
        ),
    )
    monkeypatch.setattr(workflow, "_read_upload_table", lambda _upload_id: SimpleNamespace())
    monkeypatch.setattr(workflow.ProfileStore, "record_run", lambda *_args, **_kwargs: None)

    result = confirm_mapping(
        upload_id=upload_id,
        target_template_id="bsn_sales",
        mapping={"Mã hóa đơn": "Số chứng từ (*)"},
        defaults={
            "TK Tiền/Chi phí/Nợ (*)": "131",
            "TK Doanh thu/Có (*)": "5111",
        },
        formulas={},
        conversion_context_token=None,
        session_id=None,
        revision=None,
        state_hash=None,
    )

    assert result["mapping_profile_kind"] == "v2"
    assert result["profile_id"] == "confirmed-v2"
    assert result["mapping_profile_version"] == 2
    assert result["profile_state_hash"] == "confirmed-state-hash"
    stored = workflow._read_metadata(upload_id)
    assert stored["mapping_profile_kind"] == "v2"
    assert stored["mapping_profile_version"] == 2
    assert stored["mapping_profile_state_hash"] == "confirmed-state-hash"


def test_export_rejects_stale_v2_profile_binding_before_writing_bytes(tmp_path, monkeypatch):
    workflow, profile, analyzed = _prepare_v2_export(tmp_path, monkeypatch)
    session = analyzed["session"]

    with pytest.raises(MappingProfileV2Error, match="version không khớp") as failure:
        workflow.export_confirmed_profile(
            analyzed["upload_id"],
            profile.id,
            acknowledge_warnings=True,
            conversion_context_token="context-token",
            session_id=session["session_id"],
            revision=session["active_revision"],
            state_hash=session["state_hash"],
            requested_profile_version=profile.version - 1,
            requested_profile_state_hash=profile.state_hash,
        )

    assert failure.value.status_code == 409
    assert not (workflow._upload_dir(analyzed["upload_id"]) / "misa_export.xls").exists()


def test_export_fails_closed_and_traces_stale_profile_confirmation_failure(
    tmp_path, monkeypatch
):
    workflow, profile, analyzed = _prepare_v2_export(tmp_path, monkeypatch)

    def reject_stale(*_args, **_kwargs):
        raise MappingProfileV2Error("stale profile state contract")

    monkeypatch.setattr(workflow, "record_confirmed_export_v2", reject_stale)
    session = analyzed["session"]

    with pytest.raises(MappingProfileV2Error, match="Không thể ghi nhận"):
        export_confirmed_profile(
            analyzed["upload_id"],
            profile.id,
            acknowledge_warnings=True,
            conversion_context_token="context-token",
            session_id=session["session_id"],
            revision=session["active_revision"],
            state_hash=session["state_hash"],
        )

    trace = workflow._read_metadata(analyzed["upload_id"])[
        "mapping_profile_v2_confirmation"
    ]
    assert trace["status"] == "failed"
    assert trace["profile_state_hash"] == profile.state_hash
    assert trace["error"] == "stale profile state contract"
    assert not (workflow._upload_dir(analyzed["upload_id"]) / "misa_export.xls").exists()
