from numbers import Number
from decimal import Decimal

import openpyxl
import pytest
import xlwt

from app.excel_io import InputReadError, _xls_cell_value, read_input_table


def _write_xlsx(path, headers, values, number_formats=None):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(headers)
    sheet.append(values)
    for col_idx, number_format in enumerate(number_formats or (), start=1):
        sheet.cell(2, col_idx).number_format = number_format
    workbook.save(path)
    return path


def _write_xls(path, headers, values, number_formats=None):
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Data")
    formats = number_formats or ["General"] * len(values)
    for col_idx, header in enumerate(headers):
        sheet.write(0, col_idx, header)
    for col_idx, value in enumerate(values):
        style = xlwt.easyxf(num_format_str=formats[col_idx])
        sheet.write(1, col_idx, value, style)
    workbook.save(str(path))
    return path


@pytest.mark.parametrize("suffix,writer", [(".xlsx", _write_xlsx), (".xls", _write_xls)])
def test_read_rejects_duplicate_normalized_headers_with_column_details(tmp_path, suffix, writer):
    path = writer(
        tmp_path / f"duplicate{suffix}",
        ["Mã hàng", " ma hang "],
        ["MH-001", "MH-002"],
    )

    with pytest.raises(InputReadError) as caught:
        read_input_table(path)

    assert caught.value.code == "duplicate_headers"
    assert "ma_hang" in caught.value.message
    assert "columns 1 and 2" in caught.value.message
    assert "Rename one column" in caught.value.message


@pytest.mark.parametrize("suffix,writer", [(".xlsx", _write_xlsx), (".xls", _write_xls)])
def test_read_preserves_zero_only_numeric_display_but_keeps_amount_numeric(
    tmp_path, suffix, writer
):
    path = writer(
        tmp_path / f"leading-zero{suffix}",
        ["Mã khách hàng", "Thành tiền"],
        [123, 2500],
        ["000000", "#,##0"],
    )

    table = read_input_table(path)

    assert table.rows == [{"Mã khách hàng": "000123", "Thành tiền": 2500}]
    assert isinstance(table.rows[0]["Thành tiền"], Number)
    assert not isinstance(table.rows[0]["Thành tiền"], str)


def test_xls_writer_blocks_unsafe_decimal_precision():
    value = Decimal("9007199254740993.0000000001")

    with pytest.raises(ValueError, match="safe precision"):
        _xls_cell_value(value, xlwt.Style.default_style)


@pytest.mark.parametrize("suffix,writer", [(".xlsx", _write_xlsx), (".xls", _write_xls)])
def test_read_blocks_numeric_cells_beyond_excel_safe_precision(tmp_path, suffix, writer):
    path = writer(
        tmp_path / f"unsafe-precision{suffix}",
        ["Thành tiền"],
        [9_007_199_254_740_993],
    )

    with pytest.raises(InputReadError) as caught:
        read_input_table(path)

    assert caught.value.code == "unsafe_numeric_precision"
