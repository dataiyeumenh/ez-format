from __future__ import annotations

import hashlib
import json
import os
import tempfile
import time
import xml.parsers.expat as expat
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
from struct import unpack_from
from typing import Any, Callable
from zipfile import BadZipFile, ZipFile

import openpyxl
import xlrd
from xlrd.compdoc import CompDoc
from openpyxl.utils.cell import range_boundaries
from pydantic import ValidationError

from app.document_structure import inspect_workbook_structure, validate_excel_magic
from app.excel_io import _numeric_display_value, _xls_input_cell_value
from app.import_result_models import (
    ImportResultColumnMapping,
    ImportResultInspection,
    ImportResultSchemaError,
    ImportResultSelectionCandidate,
    NormalizedImportIssue,
)
from app.normalization import is_blank, normalize_header


MAX_IMPORT_RESULT_STRING_LENGTH = 2000
MAX_IMPORT_RESULT_HEADER_LENGTH = 256
MAX_IMPORT_RESULT_SAMPLE_ROWS = 20
MAX_IMPORT_RESULT_CANDIDATES = 20
MAX_HEADER_SCAN_ROWS = 30
_COLUMN_ROLES = (
    "source_row_number",
    "document_number",
    "invoice_number",
    "document_date",
    "partner_code",
    "item_code",
    "amount",
)
_HEADER_KEYWORDS = (
    "ma",
    "ngay",
    "thoi_gian",
    "ten",
    "so",
    "so_luong",
    "don_gia",
    "thanh_tien",
)
_BIFF_BOF = 0x0809
_BIFF_BOUNDSHEET = 0x0085
_BIFF_DIMENSIONS = 0x0200
_BIFF_WORKSHEET_SUBSTREAM = 0x0010


@dataclass(frozen=True)
class _ImportResultTable:
    sheet_name: str
    header_row: int
    headers: list[str]
    rows: list[tuple[int, dict[str, Any]]]


@dataclass(frozen=True)
class _SelectionCandidate:
    sheet_name: str
    header_row: int
    headers: list[str]
    rank: tuple[int, int, int]


_RowConsumer = Callable[[int, dict[str, Any]], None]


def inspect_import_result(content: bytes, filename: str) -> ImportResultInspection:
    table, warnings, candidates = _read_safe_import_result(
        content,
        filename,
        retained_rows=MAX_IMPORT_RESULT_SAMPLE_ROWS,
    )
    sample_rows = [
        {header: _bounded_text(row.get(header)) for header in table.headers}
        for _, row in table.rows
    ]
    visible_candidates = [_candidate_model(candidate) for candidate in candidates]
    top_rank = max((candidate.rank[:2] for candidate in candidates), default=None)
    ambiguous = bool(
        top_rank
        and sum(candidate.rank[:2] == top_rank for candidate in candidates) > 1
    )
    return ImportResultInspection(
        sheet_name=_bounded_sheet_name(table.sheet_name),
        header_row=table.header_row,
        headers=[_validated_header(header) for header in table.headers],
        sample_rows=sample_rows,
        warnings=_bounded_warnings(warnings),
        candidates=visible_candidates,
        selection_ambiguous=ambiguous,
    )


def normalize_import_result(
    content: bytes,
    mapping: ImportResultColumnMapping | dict[str, Any],
    filename: str | None = None,
) -> list[NormalizedImportIssue]:
    resolved_mapping = _validate_mapping(mapping)
    issues: list[NormalizedImportIssue] = []
    decoded_text_bytes = 0
    max_decoded_text_bytes = _positive_limit(
        "IMPORT_RESULT_MAX_DECODED_TEXT_BYTES", 8 * 1024 * 1024
    )
    max_issues = _positive_limit("IMPORT_RESULT_MAX_ISSUES", 10000)
    mapped_headers = list(dict.fromkeys(_mapped_headers(resolved_mapping).values()))

    def consume_row(artifact_row_number: int, row: dict[str, Any]) -> None:
        nonlocal decoded_text_bytes
        for header in mapped_headers:
            value = row.get(header)
            if value is not None:
                decoded_text_bytes += len(str(value).encode("utf-8"))
                if decoded_text_bytes > max_decoded_text_bytes:
                    raise ImportResultSchemaError(
                        "Workbook vuot gioi han decoded text "
                        f"{max_decoded_text_bytes} bytes"
                    )
        technical_message = _bounded_text(
            row.get(resolved_mapping.columns.technical_message)
        ).strip()
        if not technical_message:
            return
        if len(issues) >= max_issues:
            raise ImportResultSchemaError(f"Workbook vuot gioi han {max_issues} issue")
        locator = _locator_for_row(row, resolved_mapping)
        issue_key = _issue_key(
            sheet_name=resolved_mapping.sheet_name,
            artifact_row_number=artifact_row_number,
            technical_message=technical_message,
            locator=locator,
        )
        issues.append(
            NormalizedImportIssue(
                issue_key=issue_key,
                artifact_row_number=artifact_row_number,
                technical_message=technical_message,
                locator=locator,
            )
        )

    table, _, _ = _read_safe_import_result(
        content,
        filename,
        selection=resolved_mapping,
        retained_rows=0,
        row_consumer=consume_row,
    )
    _validate_mapping_against_table(resolved_mapping, table)
    return issues


def _read_safe_import_result(
    content: bytes,
    filename: str | None,
    *,
    selection: ImportResultColumnMapping | None = None,
    retained_rows: int | None = None,
    row_consumer: _RowConsumer | None = None,
) -> tuple[_ImportResultTable, list[dict[str, str]], list[_SelectionCandidate]]:
    deadline = time.monotonic() + _positive_limit("IMPORT_RESULT_MAX_PARSE_SECONDS", 15)
    suffix = _workbook_suffix(content, filename)
    if suffix not in {".xls", ".xlsx"}:
        raise ImportResultSchemaError("Chi ho tro workbook .xls va .xlsx")
    if len(content) > _max_file_bytes():
        raise ImportResultSchemaError(f"File vuot gioi han {_max_file_bytes()} bytes")
    try:
        validate_excel_magic(f"workbook{suffix}", content)
    except ValueError as exc:
        raise ImportResultSchemaError(str(exc)) from exc

    xlsx_warnings = _preflight_xlsx(content, deadline=deadline) if suffix == ".xlsx" else []
    if suffix == ".xls":
        _preflight_xls(content, deadline=deadline)
    with tempfile.TemporaryDirectory(prefix="ezformat-import-result-") as directory:
        workbook_path = Path(directory) / f"input{suffix}"
        workbook_path.write_bytes(content)
        if suffix == ".xlsx":
            return _read_xlsx_import_result(
                workbook_path,
                selection=selection,
                retained_rows=retained_rows,
                warnings=xlsx_warnings,
                row_consumer=row_consumer,
                deadline=deadline,
            )
        return _read_xls_import_result(
            workbook_path,
            selection=selection,
            retained_rows=retained_rows,
            row_consumer=row_consumer,
            deadline=deadline,
        )


def _read_xlsx_import_result(
    path: Path,
    *,
    selection: ImportResultColumnMapping | None,
    retained_rows: int | None,
    warnings: list[dict[str, str]],
    row_consumer: _RowConsumer | None,
    deadline: float,
) -> tuple[_ImportResultTable, list[dict[str, str]], list[_SelectionCandidate]]:
    try:
        workbook = openpyxl.load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise ImportResultSchemaError(f"Cannot read Excel workbook: {exc}") from exc
    try:
        _enforce_deadline(deadline)
        candidates = _xlsx_candidates(workbook, deadline=deadline)
        selected = _select_candidate(candidates, selection)
        table = _xlsx_selected_table(
            workbook,
            selected,
            retained_rows,
            row_consumer=row_consumer,
            deadline=deadline,
        )
        return table, warnings, candidates
    finally:
        workbook.close()


def _read_xls_import_result(
    path: Path,
    *,
    selection: ImportResultColumnMapping | None,
    retained_rows: int | None,
    row_consumer: _RowConsumer | None,
    deadline: float,
) -> tuple[_ImportResultTable, list[dict[str, str]], list[_SelectionCandidate]]:
    try:
        workbook = xlrd.open_workbook(str(path), formatting_info=True)
    except Exception as exc:
        raise ImportResultSchemaError(f"Cannot read .xls file: {exc}") from exc
    _enforce_deadline(deadline)
    candidates = _xls_candidates(workbook, deadline=deadline)
    selected = _select_candidate(candidates, selection)
    table = _xls_selected_table(
        workbook,
        selected,
        retained_rows,
        row_consumer=row_consumer,
        deadline=deadline,
    )
    _enforce_deadline(deadline)
    structure = inspect_workbook_structure(path)
    _enforce_deadline(deadline)
    if structure.get("has_external_links"):
        raise ImportResultSchemaError("Workbook co external links khong duoc ho tro")
    warnings = list(structure.get("warnings") or [])
    warnings.append(
        {
            "code": "xls_external_link_detection_unavailable",
            "message": "Workbook .xls chi duoc xu ly thu cong; khong tu dong xac nhan external links.",
        }
    )
    return table, warnings, candidates


def _xlsx_candidates(
    workbook: openpyxl.Workbook,
    *,
    deadline: float,
) -> list[_SelectionCandidate]:
    candidates: list[_SelectionCandidate] = []
    for worksheet in workbook.worksheets:
        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_row=MAX_HEADER_SCAN_ROWS,
                values_only=True,
            ),
            start=1,
        ):
            _enforce_deadline(deadline)
            candidate = _candidate_from_values(worksheet.title, row_number, list(row))
            if candidate:
                candidates.append(candidate)
    return _limit_candidates(candidates)


def _xls_candidates(
    workbook: xlrd.book.Book,
    *,
    deadline: float,
) -> list[_SelectionCandidate]:
    candidates: list[_SelectionCandidate] = []
    for worksheet in workbook.sheets():
        for row_index in range(min(worksheet.nrows, MAX_HEADER_SCAN_ROWS)):
            _enforce_deadline(deadline)
            values = [worksheet.cell_value(row_index, column) for column in range(worksheet.ncols)]
            candidate = _candidate_from_values(worksheet.name, row_index + 1, values)
            if candidate:
                candidates.append(candidate)
    return _limit_candidates(candidates)


def _candidate_from_values(
    sheet_name: str,
    row_number: int,
    values: list[Any],
) -> _SelectionCandidate | None:
    headers = ["" if value is None else str(value).strip() for value in values]
    non_empty = sum(bool(header) for header in headers)
    if not non_empty:
        return None
    keyword_hits = sum(
        any(
            keyword in normalize_header(header).split("_")
            if "_" not in keyword
            else keyword in normalize_header(header)
            for keyword in _HEADER_KEYWORDS
        )
        for header in headers
        if header
    )
    return _SelectionCandidate(
        sheet_name=sheet_name,
        header_row=row_number,
        headers=headers,
        rank=(keyword_hits, non_empty, -row_number),
    )


def _limit_candidates(candidates: list[_SelectionCandidate]) -> list[_SelectionCandidate]:
    return sorted(candidates, key=lambda candidate: candidate.rank, reverse=True)[
        :MAX_IMPORT_RESULT_CANDIDATES
    ]


def _select_candidate(
    candidates: list[_SelectionCandidate],
    selection: ImportResultColumnMapping | None,
) -> _SelectionCandidate:
    if selection is not None:
        return _SelectionCandidate(
            sheet_name=selection.sheet_name,
            header_row=selection.header_row,
            headers=[],
            rank=(0, 0, 0),
        )
    if not candidates:
        raise ImportResultSchemaError("Workbook khong co sheet du lieu de map thu cong")
    return candidates[0]


def _xlsx_selected_table(
    workbook: openpyxl.Workbook,
    candidate: _SelectionCandidate,
    retained_rows: int | None,
    *,
    row_consumer: _RowConsumer | None,
    deadline: float,
) -> _ImportResultTable:
    if candidate.sheet_name not in workbook.sheetnames:
        raise ImportResultSchemaError("sheet_name khong ton tai trong workbook")
    worksheet = workbook[candidate.sheet_name]
    worksheet.reset_dimensions()
    header_cells = next(
        worksheet.iter_rows(
            min_row=candidate.header_row,
            max_row=candidate.header_row,
        ),
        (),
    )
    headers = _validated_headers([cell.value for cell in header_cells])
    if not any(headers):
        raise ImportResultSchemaError("header_row khong co header de map")
    rows: list[tuple[int, dict[str, Any]]] = []
    rows_seen = 1
    cells_seen = len(header_cells)
    for cells in worksheet.iter_rows(min_row=candidate.header_row + 1):
        rows_seen += 1
        cells_seen += len(cells)
        _enforce_stream_budget(
            rows_seen,
            cells_seen,
            columns=len(cells),
            deadline=deadline,
        )
        row = _record_from_values(
            headers,
            [_numeric_display_value(cell.value, cell.number_format) for cell in cells],
        )
        if row is None:
            continue
        artifact_row_number = (
            cells[0].row if cells else candidate.header_row + rows_seen - 1
        )
        if row_consumer is not None:
            row_consumer(artifact_row_number, row)
        if retained_rows is None or len(rows) < retained_rows:
            rows.append((artifact_row_number, row))
    return _ImportResultTable(candidate.sheet_name, candidate.header_row, headers, rows)


def _xls_selected_table(
    workbook: xlrd.book.Book,
    candidate: _SelectionCandidate,
    retained_rows: int | None,
    *,
    row_consumer: _RowConsumer | None,
    deadline: float,
) -> _ImportResultTable:
    try:
        worksheet = workbook.sheet_by_name(candidate.sheet_name)
    except xlrd.biffh.XLRDError as exc:
        raise ImportResultSchemaError("sheet_name khong ton tai trong workbook") from exc
    if candidate.header_row > worksheet.nrows:
        raise ImportResultSchemaError("header_row khong co header de map")
    headers = _validated_headers(worksheet.row_values(candidate.header_row - 1))
    if not any(headers):
        raise ImportResultSchemaError("header_row khong co header de map")
    rows: list[tuple[int, dict[str, Any]]] = []
    rows_seen = 1
    cells_seen = len(headers)
    for row_index in range(candidate.header_row, worksheet.nrows):
        rows_seen += 1
        cells_seen += worksheet.ncols
        _enforce_stream_budget(
            rows_seen,
            cells_seen,
            columns=worksheet.ncols,
            deadline=deadline,
        )
        row = _record_from_values(
            headers,
            [
                _xls_input_cell_value(workbook, worksheet.cell(row_index, column))
                for column in range(worksheet.ncols)
            ],
        )
        if row is None:
            continue
        if row_consumer is not None:
            row_consumer(row_index + 1, row)
        if retained_rows is None or len(rows) < retained_rows:
            rows.append((row_index + 1, row))
    return _ImportResultTable(candidate.sheet_name, candidate.header_row, headers, rows)


def _validated_headers(values: list[Any]) -> list[str]:
    headers = ["" if value is None else str(value).strip() for value in values]
    seen: dict[str, int] = {}
    for column_index, header in enumerate(headers, start=1):
        normalized = normalize_header(header)
        if not normalized:
            continue
        if normalized in seen:
            raise ImportResultSchemaError(
                f"Duplicate normalized header '{normalized}' at columns "
                f"{seen[normalized]} and {column_index}. Rename one column before importing."
            )
        if len(header) > MAX_IMPORT_RESULT_HEADER_LENGTH:
            raise ImportResultSchemaError("Header vuot gioi han do dai")
        seen[normalized] = column_index
    return headers


def _record_from_values(headers: list[str], values: list[Any]) -> dict[str, Any] | None:
    values = values[: len(headers)] + [None] * max(0, len(headers) - len(values))
    if all(is_blank(value) for value in values):
        return None
    return {header: value for header, value in zip(headers, values)}


def _enforce_deadline(deadline: float | None) -> None:
    if deadline is not None and time.monotonic() > deadline:
        raise ImportResultSchemaError("Workbook parse qua thoi gian cho phep")


def _enforce_stream_budget(
    rows_seen: int,
    cells_seen: int,
    *,
    columns: int | None = None,
    deadline: float | None = None,
) -> None:
    _enforce_deadline(deadline)
    max_rows = _positive_limit("IMPORT_RESULT_MAX_ROWS", 10000)
    max_cells = _positive_limit("IMPORT_RESULT_MAX_CELLS", 250000)
    max_columns = _positive_limit("IMPORT_RESULT_MAX_COLUMNS", 256)
    if rows_seen > max_rows:
        raise ImportResultSchemaError(f"Workbook vuot gioi han {max_rows} dong")
    if cells_seen > max_cells:
        raise ImportResultSchemaError(f"Workbook vuot gioi han {max_cells} o")
    if columns is not None and columns > max_columns:
        raise ImportResultSchemaError(f"Workbook vuot gioi han {max_columns} cot")


def _preflight_xlsx(content: bytes, *, deadline: float) -> list[dict[str, str]]:
    try:
        with ZipFile(BytesIO(content)) as archive:
            entries = archive.infolist()
            max_expanded_bytes = _positive_limit(
                "IMPORT_RESULT_MAX_EXPANDED_BYTES", 100 * 1024 * 1024
            )
            expanded_bytes = sum(entry.file_size for entry in entries)
            if expanded_bytes > max_expanded_bytes:
                raise ImportResultSchemaError(
                    f"Workbook vuot gioi han expanded ZIP {max_expanded_bytes} bytes"
                )
            max_entries = _positive_limit("IMPORT_RESULT_MAX_ZIP_ENTRIES", 10000)
            if len(entries) > max_entries:
                raise ImportResultSchemaError(
                    f"Workbook vuot gioi han {max_entries} ZIP entries"
                )
            _enforce_deadline(deadline)
            if _xlsx_has_external_links(archive, entries, deadline=deadline):
                raise ImportResultSchemaError("Workbook co external links khong duoc ho tro")
            decoded_text_bytes = _enforce_xlsx_shared_string_budget(
                archive,
                entries,
                deadline=deadline,
            )
            return _xlsx_preflight_structure(
                archive,
                entries,
                deadline=deadline,
                decoded_text_bytes=decoded_text_bytes,
            )
    except (BadZipFile, expat.ExpatError) as exc:
        raise ImportResultSchemaError("Workbook .xlsx khong doc duoc ZIP archive") from exc


def _preflight_xls(content: bytes, *, deadline: float) -> None:
    try:
        compound = CompDoc(content, logfile=StringIO())
        stream_memory, stream_offset, stream_length = compound.locate_named_stream("Workbook")
        if stream_memory is None:
            stream_memory, stream_offset, stream_length = compound.locate_named_stream("Book")
        if stream_memory is None or stream_length <= 0:
            raise ImportResultSchemaError("Workbook .xls khong co BIFF Workbook stream")
        max_stream_bytes = _positive_limit(
            "IMPORT_RESULT_MAX_XLS_BIFF_BYTES", _max_file_bytes()
        )
        if stream_length > max_stream_bytes:
            raise ImportResultSchemaError(
                f"Workbook .xls vuot gioi han BIFF {max_stream_bytes} bytes"
            )
        stream = memoryview(stream_memory)[stream_offset : stream_offset + stream_length]
        _enforce_biff_dimensions(stream, deadline=deadline)
    except ImportResultSchemaError:
        raise
    except Exception as exc:
        raise ImportResultSchemaError(
            "Khong the xac minh an toan BIFF workbook .xls truoc khi parse"
        ) from exc


def _enforce_biff_dimensions(stream: memoryview, *, deadline: float) -> None:
    offset = 0
    records = 0
    worksheet_count = 0
    dimensions_count = 0
    total_rows = 0
    total_cells = 0
    in_worksheet = False
    max_records = _positive_limit("IMPORT_RESULT_MAX_BIFF_RECORDS", 1000000)

    while offset < len(stream):
        _enforce_deadline(deadline)
        if offset + 4 > len(stream):
            if not any(stream[offset:]):
                break
            raise ImportResultSchemaError("BIFF workbook co record header khong hop le")
        record_id, data_length = unpack_from("<HH", stream, offset)
        if record_id == 0 and data_length == 0 and not any(stream[offset:]):
            break
        record_end = offset + 4 + data_length
        if record_end > len(stream):
            raise ImportResultSchemaError("BIFF workbook co record bi cat ngan")
        records += 1
        if records > max_records:
            raise ImportResultSchemaError(
                f"Workbook .xls vuot gioi han {max_records} BIFF records"
            )
        if record_id == _BIFF_BOUNDSHEET:
            if data_length < 6:
                raise ImportResultSchemaError("BIFF BOUNDSHEET record khong hop le")
            if stream[offset + 9] == 0:
                worksheet_count += 1
        elif record_id == _BIFF_BOF:
            if data_length < 4:
                raise ImportResultSchemaError("BIFF BOF record khong hop le")
            in_worksheet = unpack_from("<H", stream, offset + 6)[0] == _BIFF_WORKSHEET_SUBSTREAM
        elif record_id == _BIFF_DIMENSIONS and in_worksheet:
            if data_length < 14:
                raise ImportResultSchemaError("BIFF DIMENSIONS record khong the xac minh")
            first_row, last_row, first_column, last_column = unpack_from(
                "<IIHH", stream, offset + 4
            )
            if last_row < first_row or last_column < first_column:
                raise ImportResultSchemaError("BIFF DIMENSIONS record khong hop le")
            rows = last_row - first_row
            columns = last_column - first_column
            total_rows += rows
            total_cells += rows * max(1, columns)
            _enforce_stream_budget(
                total_rows,
                total_cells,
                columns=columns,
                deadline=deadline,
            )
            dimensions_count += 1
            in_worksheet = False
        offset = record_end

    if worksheet_count <= 0 or dimensions_count != worksheet_count:
        raise ImportResultSchemaError(
            "Khong the xac minh BIFF DIMENSIONS cho tat ca worksheet"
        )


def _xlsx_has_external_links(
    archive: ZipFile,
    entries: list[Any],
    *,
    deadline: float,
) -> bool:
    for entry in entries:
        _enforce_deadline(deadline)
        name = entry.filename.replace("\\", "/")
        if name.startswith("xl/externalLinks/"):
            return True
        if not name.endswith(".rels"):
            continue
        relationships = archive.read(entry)
        if b'TargetMode="External"' in relationships or b"TargetMode='External'" in relationships:
            return True
    return False


def _enforce_xlsx_shared_string_budget(
    archive: ZipFile,
    entries: list[Any],
    *,
    deadline: float,
) -> int:
    max_bytes = _positive_limit(
        "IMPORT_RESULT_MAX_DECODED_TEXT_BYTES", 8 * 1024 * 1024
    )
    decoded_bytes = 0
    for entry in entries:
        if entry.filename.replace("\\", "/") != "xl/sharedStrings.xml":
            continue
        text_depth = 0
        parser = expat.ParserCreate(namespace_separator="}")

        def start(name: str, _attributes: dict[str, str]) -> None:
            nonlocal text_depth
            if name.rsplit("}", 1)[-1] == "t":
                text_depth += 1

        def end(name: str) -> None:
            nonlocal text_depth
            if name.rsplit("}", 1)[-1] == "t":
                text_depth -= 1

        def character_data(value: str) -> None:
            nonlocal decoded_bytes
            if text_depth <= 0 or not value:
                return
            decoded_bytes += len(value.encode("utf-8"))
            if decoded_bytes > max_bytes:
                raise ImportResultSchemaError(
                    f"Workbook vuot gioi han decoded text {max_bytes} bytes"
                )

        parser.StartElementHandler = start
        parser.EndElementHandler = end
        parser.CharacterDataHandler = character_data
        parser.StartDoctypeDeclHandler = _reject_xlsx_doctype
        with archive.open(entry) as shared_strings:
            while chunk := shared_strings.read(64 * 1024):
                _enforce_deadline(deadline)
                parser.Parse(chunk, False)
            parser.Parse(b"", True)
    return decoded_bytes


def _reject_xlsx_doctype(*_args: object) -> None:
    raise ImportResultSchemaError("Workbook .xlsx khong cho phep DTD")


def _xlsx_preflight_structure(
    archive: ZipFile,
    entries: list[Any],
    *,
    deadline: float,
    decoded_text_bytes: int,
) -> list[dict[str, str]]:
    total_rows = 0
    total_cells = 0
    formulas = 0
    hidden_rows = 0
    hidden_columns = 0
    merged_ranges = 0
    for entry in entries:
        if not entry.filename.startswith("xl/worksheets/") or not entry.filename.endswith(".xml"):
            continue
        stats = _xlsx_sheet_stats(
            archive,
            entry,
            deadline=deadline,
            decoded_text_bytes=decoded_text_bytes,
        )
        decoded_text_bytes += stats["decoded_text_bytes"]
        total_rows += stats["rows"]
        total_cells += stats["cells"]
        _enforce_stream_budget(
            total_rows,
            total_cells,
            columns=stats["columns"],
            deadline=deadline,
        )
        formulas += stats["formulas"]
        hidden_rows += stats["hidden_rows"]
        hidden_columns += stats["hidden_columns"]
        merged_ranges += stats["merged_ranges"]
    warnings: list[dict[str, str]] = []
    if formulas:
        warnings.append(
            {"code": "formula_cells_detected", "message": "Workbook co o cong thuc; can kiem tra."}
        )
    if hidden_rows or hidden_columns:
        warnings.append(
            {"code": "hidden_rows_or_columns_detected", "message": "Workbook co dong hoac cot an."}
        )
    if merged_ranges:
        warnings.append(
            {"code": "merged_cells_detected", "message": "Workbook co merged cells."}
        )
    return warnings


def _xlsx_sheet_stats(
    archive: ZipFile,
    entry: Any,
    *,
    deadline: float,
    decoded_text_bytes: int,
) -> dict[str, int]:
    stats = {
        "rows": 0,
        "cells": 0,
        "formulas": 0,
        "hidden_rows": 0,
        "hidden_columns": 0,
        "merged_ranges": 0,
        "columns": 0,
        "decoded_text_bytes": 0,
    }
    dimension_rows = 0
    dimension_cells = 0
    text_depth = 0
    max_decoded_text_bytes = _positive_limit(
        "IMPORT_RESULT_MAX_DECODED_TEXT_BYTES", 8 * 1024 * 1024
    )
    parser = expat.ParserCreate(namespace_separator="}")

    def start(name: str, attributes: dict[str, str]) -> None:
        nonlocal dimension_rows, dimension_cells, text_depth
        _enforce_deadline(deadline)
        tag = name.rsplit("}", 1)[-1]
        if tag == "dimension":
            dimension_rows, dimension_cells, dimension_columns = _dimension_size(
                attributes.get("ref")
            )
            stats["columns"] = max(stats["columns"], dimension_columns)
        elif tag == "row":
            stats["rows"] += 1
            if attributes.get("hidden") in {"1", "true"}:
                stats["hidden_rows"] += 1
        elif tag == "c":
            stats["cells"] += 1
            try:
                column, _, _, _ = range_boundaries(str(attributes.get("r") or ""))
            except ValueError:
                column = 0
            stats["columns"] = max(stats["columns"], column)
        elif tag == "f":
            stats["formulas"] += 1
        elif tag == "col" and attributes.get("hidden") in {"1", "true"}:
            stats["hidden_columns"] += 1
        elif tag == "mergeCell":
            stats["merged_ranges"] += 1
        if tag in {"t", "v", "f"}:
            text_depth += 1

    def end(name: str) -> None:
        nonlocal text_depth
        if name.rsplit("}", 1)[-1] in {"t", "v", "f"}:
            text_depth -= 1

    def character_data(value: str) -> None:
        if text_depth <= 0 or not value:
            return
        stats["decoded_text_bytes"] += len(value.encode("utf-8"))
        if decoded_text_bytes + stats["decoded_text_bytes"] > max_decoded_text_bytes:
            raise ImportResultSchemaError(
                "Workbook vuot gioi han decoded text "
                f"{max_decoded_text_bytes} bytes"
            )

    parser.StartElementHandler = start
    parser.EndElementHandler = end
    parser.CharacterDataHandler = character_data
    parser.StartDoctypeDeclHandler = _reject_xlsx_doctype
    with archive.open(entry) as sheet_xml:
        while chunk := sheet_xml.read(64 * 1024):
            _enforce_deadline(deadline)
            parser.Parse(chunk, False)
        parser.Parse(b"", True)
    stats["rows"] = max(stats["rows"], dimension_rows)
    stats["cells"] = max(stats["cells"], dimension_cells)
    return stats


def _dimension_size(reference: object) -> tuple[int, int, int]:
    if not reference:
        return 0, 0, 0
    try:
        min_column, min_row, max_column, max_row = range_boundaries(str(reference))
    except ValueError as exc:
        raise ImportResultSchemaError("Workbook co worksheet dimension khong hop le") from exc
    rows = max_row - min_row + 1
    columns = max_column - min_column + 1
    return rows, rows * columns, columns


def _workbook_suffix(content: bytes, filename: str | None) -> str:
    suffix = Path(filename or "").suffix.lower()
    if suffix in {".xls", ".xlsx"}:
        return suffix
    if content.startswith(b"PK\x03\x04"):
        return ".xlsx"
    if content.startswith(bytes.fromhex("D0CF11E0A1B11AE1")):
        return ".xls"
    return suffix


def _validate_mapping(
    mapping: ImportResultColumnMapping | dict[str, Any],
) -> ImportResultColumnMapping:
    if isinstance(mapping, ImportResultColumnMapping):
        return mapping
    if not isinstance(mapping, dict):
        raise ImportResultSchemaError("mapping phai la JSON object")
    candidate = mapping
    if "columns" not in candidate:
        candidate = {
            "sheet_name": candidate.get("sheet_name") or "unknown",
            "header_row": candidate.get("header_row") or 1,
            "columns": candidate,
        }
    try:
        return ImportResultColumnMapping.model_validate(candidate)
    except ValidationError as exc:
        if any("technical_message" in str(error.get("loc", ())) for error in exc.errors()):
            raise ImportResultSchemaError("mapping requires technical_message") from exc
        raise ImportResultSchemaError("mapping import-result khong hop le") from exc


def _validate_mapping_against_table(
    mapping: ImportResultColumnMapping,
    table: _ImportResultTable,
) -> None:
    if mapping.sheet_name != table.sheet_name or mapping.header_row != table.header_row:
        raise ImportResultSchemaError("Lua chon sheet/header khong khop workbook")
    headers = set(table.headers)
    for role, header in _mapped_headers(mapping).items():
        if header not in headers:
            raise ImportResultSchemaError(f"Cot map cho {role} khong ton tai")


def _mapped_headers(mapping: ImportResultColumnMapping) -> dict[str, str]:
    headers = {"technical_message": mapping.columns.technical_message}
    for role in _COLUMN_ROLES:
        header = getattr(mapping.columns, role)
        if header:
            headers[role] = header
    return headers


def _locator_for_row(
    row: dict[str, Any],
    mapping: ImportResultColumnMapping,
) -> dict[str, str | int | None]:
    locator: dict[str, str | int | None] = {"sheet_name": mapping.sheet_name}
    for role in _COLUMN_ROLES:
        header = getattr(mapping.columns, role)
        if header:
            locator[role] = _bounded_optional_text(row.get(header))
    return locator


def _issue_key(
    *,
    sheet_name: str,
    artifact_row_number: int,
    technical_message: str,
    locator: dict[str, str | int | None],
) -> str:
    payload = json.dumps(
        {
            "sheet_name": sheet_name,
            "artifact_row_number": artifact_row_number,
            "technical_message": technical_message,
            "locator": locator,
        },
        ensure_ascii=True,
        sort_keys=True,
        separators=(",", ":"),
    )
    return f"manual_excel_v1:{hashlib.sha256(payload.encode('utf-8')).hexdigest()[:24]}"


def _candidate_model(candidate: _SelectionCandidate) -> ImportResultSelectionCandidate:
    return ImportResultSelectionCandidate(
        sheet_name=_bounded_sheet_name(candidate.sheet_name),
        header_row=candidate.header_row,
        headers=[_validated_header(header) for header in candidate.headers],
    )


def _max_file_bytes() -> int:
    return _positive_limit(
        "IMPORT_RESULT_MAX_FILE_BYTES",
        _positive_limit("MAX_UPLOAD_BYTES", 20 * 1024 * 1024),
    )


def _positive_limit(name: str, default: int) -> int:
    try:
        return max(1, int(os.getenv(name, str(default))))
    except ValueError:
        return default


def _validated_header(value: object) -> str:
    header = _bounded_text(value).strip()
    if not header:
        return ""
    if len(str(value or "")) > MAX_IMPORT_RESULT_HEADER_LENGTH:
        raise ImportResultSchemaError("Header vuot gioi han do dai")
    return header


def _bounded_sheet_name(value: object) -> str:
    return _bounded_text(value, limit=128)


def _bounded_warnings(warnings: list[object]) -> list[dict[str, str]]:
    bounded: list[dict[str, str]] = []
    for warning in warnings:
        if not isinstance(warning, dict):
            continue
        bounded.append(
            {
                "code": _bounded_text(warning.get("code"), limit=128),
                "message": _bounded_text(warning.get("message")),
            }
        )
    return bounded


def _bounded_optional_text(value: object) -> str | None:
    if value is None:
        return None
    return _bounded_text(value)


def _bounded_text(value: object, *, limit: int = MAX_IMPORT_RESULT_STRING_LENGTH) -> str:
    return str(value if value is not None else "")[:limit]
