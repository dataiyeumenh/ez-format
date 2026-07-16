from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Any

import openpyxl
import xlrd

from app.excel_io import InputTable, read_input_table


SUPPORTED_CATALOG_TYPES = frozenset(
    {
        "account",
        "supplier",
        "customer",
        "item",
        "warehouse",
        "unit",
        "employee",
        "bank_account",
    }
)


CATALOG_FIELDS: dict[str, dict[str, tuple[str, ...]]] = {
    "account": {
        "code": ("so tai khoan", "ma tai khoan", "tai khoan", "tk"),
        "name": ("ten tai khoan", "ten tk"),
    },
    "supplier": {
        "code": ("ma nha cung cap", "ma ncc", "ma doi tuong", "ma"),
        "name": ("ten nha cung cap", "ten ncc", "ten doi tuong", "ten"),
        "taxCode": ("ma so thue", "mst"),
    },
    "customer": {
        "code": ("ma khach hang", "ma kh", "ma doi tuong", "ma"),
        "name": ("ten khach hang", "ten kh", "ten doi tuong", "ten"),
        "taxCode": ("ma so thue", "mst"),
    },
    "item": {
        "code": ("ma hang", "ma vthh", "ma vat tu", "ma hang hoa", "ma"),
        "name": ("ten hang", "ten vthh", "ten vat tu", "ten hang hoa", "ten"),
        "unit": ("dvt", "don vi tinh"),
        "barcode": ("ma vach", "barcode"),
    },
    "warehouse": {
        "code": ("ma kho", "ma"),
        "name": ("ten kho", "ten"),
    },
    "unit": {
        "code": ("ma don vi tinh", "ma dvt", "dvt", "ma"),
        "name": ("ten don vi tinh", "ten dvt", "don vi tinh", "ten"),
    },
    "employee": {
        "code": ("ma nhan vien", "ma nv", "ma"),
        "name": ("ten nhan vien", "ho va ten", "ten"),
        "taxCode": ("ma so thue", "mst"),
    },
    "bank_account": {
        "code": ("so tai khoan", "tai khoan ngan hang", "stk", "ma"),
        "name": ("ten tai khoan", "ten ngan hang", "ten"),
    },
}


def normalize_code(value: Any) -> str:
    return _cell_text(value).upper()


def normalize_name(value: Any) -> str:
    text = unicodedata.normalize("NFD", _cell_text(value))
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text.lower())
    return re.sub(r"\s+", " ", text).strip()


def normalize_tax_code(value: Any) -> str:
    return re.sub(r"\s+", "", _cell_text(value).upper())


def parse_master_data_file(path: Path, catalog_type: str) -> dict[str, Any]:
    if catalog_type not in SUPPORTED_CATALOG_TYPES:
        raise ValueError(f"Loại danh mục không được hỗ trợ: {catalog_type}")

    table = read_input_table(path)
    if not table.headers:
        raise ValueError("Không xác định được dòng tiêu đề của danh mục MISA")

    aliases = CATALOG_FIELDS[catalog_type]
    detected = {
        field: _find_header(table.headers, candidates)
        for field, candidates in aliases.items()
    }
    if not detected.get("code") and not detected.get("name"):
        raise ValueError("Không tìm thấy cột mã hoặc tên danh mục")

    entries: list[dict[str, Any]] = []
    warnings: list[str] = []
    seen_codes: set[str] = set()
    for row in _catalog_rows(path, table, detected):
        code = row.get("code", "")
        name = row.get("name", "")
        tax_code = row.get("taxCode", "")
        if not code and not name and not tax_code:
            continue
        normalized_code = normalize_code(code)
        if normalized_code and normalized_code in seen_codes:
            warnings.append(
                f"Mã {normalized_code} xuất hiện nhiều lần; chỉ giữ dòng đầu tiên."
            )
            continue
        if normalized_code:
            seen_codes.add(normalized_code)

        attributes: dict[str, Any] = {}
        for field in ("unit", "barcode"):
            if detected.get(field):
                attributes[field] = row.get(field, "")

        entries.append(
            {
                "code": code,
                "normalizedCode": normalized_code,
                "name": name,
                "normalizedName": normalize_name(name),
                "taxCode": tax_code,
                "normalizedTaxCode": normalize_tax_code(tax_code),
                "active": True,
                "attributes": attributes,
            }
        )

    return {
        "catalog_type": catalog_type,
        "sheet_name": table.sheet_name,
        "header_row": table.header_row_index + 1,
        "detected_columns": detected,
        "entries": entries,
        "warnings": warnings,
    }


def _find_header(headers: list[str], aliases: tuple[str, ...]) -> str | None:
    normalized = {normalize_name(header): header for header in headers if header}
    for alias in aliases:
        if alias in normalized:
            return normalized[alias]
    for alias in aliases:
        for candidate, original in normalized.items():
            if alias and (candidate.startswith(alias) or alias in candidate):
                return original
    return None


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _catalog_rows(
    path: Path,
    table: InputTable,
    detected: dict[str, str | None],
) -> list[dict[str, str]]:
    try:
        if path.suffix.lower() == ".xlsx":
            return _xlsx_catalog_rows(
                path,
                table.sheet_name,
                table.header_row_index,
                detected,
            )
        return _xls_catalog_rows(
            path,
            table.sheet_name,
            table.header_row_index,
            detected,
        )
    except (OSError, ValueError, KeyError, IndexError, xlrd.XLRDError):
        return [
            {
                field: _cell_text(row.get(header))
                for field, header in detected.items()
                if header
            }
            for row in table.rows
        ]


def _xlsx_catalog_rows(
    path: Path,
    sheet_name: str | None,
    header_row_index: int,
    detected: dict[str, str | None],
) -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        header_cells = next(
            sheet.iter_rows(
                min_row=header_row_index + 1,
                max_row=header_row_index + 1,
            ),
            (),
        )
        columns = {
            _cell_text(cell.value): index
            for index, cell in enumerate(header_cells)
            if _cell_text(cell.value)
        }
        records: list[dict[str, str]] = []
        for cells in sheet.iter_rows(min_row=header_row_index + 2):
            record = {
                field: _formatted_cell_text(
                    cells[columns[header]].value,
                    cells[columns[header]].number_format,
                )
                for field, header in detected.items()
                if header in columns and columns[header] < len(cells)
            }
            if any(record.values()):
                records.append(record)
        return records
    finally:
        workbook.close()


def _xls_catalog_rows(
    path: Path,
    sheet_name: str | None,
    header_row_index: int,
    detected: dict[str, str | None],
) -> list[dict[str, str]]:
    workbook = xlrd.open_workbook(str(path), formatting_info=True)
    sheet = workbook.sheet_by_name(sheet_name) if sheet_name else workbook.sheet_by_index(0)
    columns = {
        _cell_text(sheet.cell_value(header_row_index, index)): index
        for index in range(sheet.ncols)
        if _cell_text(sheet.cell_value(header_row_index, index))
    }
    records: list[dict[str, str]] = []
    for row_index in range(header_row_index + 1, sheet.nrows):
        record: dict[str, str] = {}
        for field, header in detected.items():
            if header not in columns:
                continue
            cell = sheet.cell(row_index, columns[header])
            format_string = ""
            if cell.xf_index is not None and cell.xf_index < len(workbook.xf_list):
                format_key = workbook.xf_list[cell.xf_index].format_key
                format_info = workbook.format_map.get(format_key)
                format_string = format_info.format_str if format_info else ""
            record[field] = _formatted_cell_text(cell.value, format_string)
        if any(record.values()):
            records.append(record)
    return records


def _formatted_cell_text(value: Any, number_format: str | None) -> str:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        numeric = float(value)
        if numeric.is_integer():
            zero_pattern = str(number_format or "").split(";", 1)[0].strip()
            if re.fullmatch(r"0+", zero_pattern):
                return str(int(numeric)).zfill(len(zero_pattern))
    return _cell_text(value)
