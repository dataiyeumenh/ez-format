from __future__ import annotations

import hashlib
import hmac
import re
import xml.etree.ElementTree as ElementTree
from collections.abc import Iterable, Mapping
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from xlrd import open_workbook
from xlutils.copy import copy as copy_xls_workbook


ANONYMIZATION_CATEGORIES = (
    "company",
    "counterparty",
    "tax_code",
    "address",
    "email",
    "phone",
    "bank_account",
    "document_number",
)

MAX_XLSX_ARCHIVE_ENTRIES = 2048
MAX_XLSX_UNCOMPRESSED_BYTES = 96 * 1024 * 1024
MAX_XLSX_ENTRY_UNCOMPRESSED_BYTES = 48 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 200

_OOXML_REPLACEABLE_ATTRIBUTES = {
    "workbookPr": {"codeName"},
    "sheet": {"name"},
    "sheetPr": {"codeName"},
    "dataValidation": {"prompt", "promptTitle", "error", "errorTitle"},
    "filter": {"val"},
    "customFilter": {"val"},
    "table": {"comment"},
    "tableColumn": {"totalsRowLabel"},
    "cNvPr": {"name", "descr", "title"},
}
_OOXML_FORMULA_ELEMENTS = {"f", "formula", "formula1", "formula2"}
_OOXML_RICH_TEXT_CONTAINERS = {"p", "si", "is", "tx", "rich"}

_TEXT_PREFIXES = {
    "company": "COMPANY",
    "counterparty": "COUNTERPARTY",
    "address": "ADDRESS",
}
_NUMERIC_PREFIXES = {
    "tax_code": "TAX",
    "phone": "PHONE",
    "bank_account": "BANK",
    "document_number": "DOC",
}

_PII_PATTERNS = {
    "company": re.compile(
        r"((?:c[oô]ng\s*ty|cong\s*ty|cty|doanh\s*nghi[eệ]p|doanh\s*nghiep)"
        r"\s+[^;,|\n]{2,120})",
        re.IGNORECASE,
    ),
    "counterparty": re.compile(
        r"(?:kh[aá]ch\s*h[aà]ng|khach\s*hang|nh[aà]\s*cung\s*c[aấ]p|"
        r"nha\s*cung\s*cap|[dđ][oố]i\s*t[aá]c|doi\s*tac)"
        r"\s*[:#-]\s*([^;,|\n]{2,120})",
        re.IGNORECASE,
    ),
    "address": re.compile(
        r"(?:[dđ][iị]a\s*ch[iỉ]|dia\s*chi|address)"
        r"\s*[:#-]\s*([^;|\n]{4,180})",
        re.IGNORECASE,
    ),
    "email": re.compile(
        r"(?<![\w.+-])[A-Z0-9.!#$%&'*+/=?^_`{|}~-]+@"
        r"[A-Z0-9](?:[A-Z0-9.-]{0,251}[A-Z0-9])?\.[A-Z]{2,63}(?![\w.-])",
        re.IGNORECASE,
    ),
    "phone": re.compile(
        r"(?<!\d)(?:\+?84|0)(?:[\s().-]*\d){8,10}(?!\d)",
        re.IGNORECASE,
    ),
    "tax_code": re.compile(
        r"(?:m[aã]\s*s[oố]\s*thu[eế]|ma\s+so\s+thue|mst|tax\s*code)"
        r"\s*[:#-]?\s*(\d{10}(?:-\d{3})?)",
        re.IGNORECASE,
    ),
    "bank_account": re.compile(
        r"(?:s[oố]\s*t[aà]i\s*kho[aả]n|so\s+tai\s+khoan|stk|bank\s*account)"
        r"\s*[:#-]?\s*(\d(?:[\s.-]*\d){5,18})",
        re.IGNORECASE,
    ),
}

_VIETNAMESE_NAME_PATTERN = re.compile(
    r"(?<!\w)((?:nguy[eễ]n|tr[aầ]n|l[eê]|ph[aạ]m|ho[aà]ng|hu[yỳ]nh|phan|"
    r"v[uũ]|v[oõ]|[dđ][aặ]ng|b[uù]i|[dđ][oỗ]|h[oồ]|ng[oô]|d[uư][oơ]ng|l[yý])"
    r"(?:\s+[A-ZÀ-ỴĐ][A-Za-zÀ-ỹĐđ]{1,30}){1,4})(?!\w)",
    re.IGNORECASE,
)
_VIETNAMESE_ADDRESS_PATTERN = re.compile(
    r"(?<!\w)(\d{1,5}(?:[/-]\d{1,5})?\s+"
    r"(?:(?i:(?:đường|duong|phố|pho|ngõ|ngo|hẻm|hem|ấp|ap|thôn|thon)"
    r"\s+[^;|\n]{2,160})|"
    r"(?:[A-ZÀ-ỴĐ][a-zà-ỹđ]{1,30}|[A-ZÀ-ỴĐ]{2,30})"
    r"(?:\s+(?:[A-ZÀ-ỴĐ][a-zà-ỹđ]{1,30}|[A-ZÀ-ỴĐ]{2,30})){1,5}))"
    r"(?!\w)"
)
_CCCD_PATTERN = re.compile(r"(?<!\d)(\d{12})(?!\d)")

# Deliberately separate from primary discovery. Export validation must still
# detect PII when discovery or replacement inventory regresses.
_POST_SCAN_PII_PATTERNS = (
    (
        "company",
        re.compile(
            r"(?:c[oô]ng\s*ty|cong\s*ty|cty|doanh\s*nghi[eệ]p|doanh\s*nghiep)"
            r"\s+[^;,|\n]{2,120}",
            re.IGNORECASE,
        ),
    ),
    (
        "counterparty",
        re.compile(
            r"(?<!\w)(?:nguy[eễ]n|tr[aầ]n|l[eê]|ph[aạ]m|ho[aà]ng|hu[yỳ]nh|"
            r"phan|v[uũ]|v[oõ]|[dđ][aặ]ng|b[uù]i|[dđ][oỗ]|h[oồ]|ng[oô]|"
            r"d[uư][oơ]ng|l[yý])"
            r"(?:\s+[A-ZÀ-ỴĐ][A-Za-zÀ-ỹĐđ]{1,30}){1,4}"
            r"(?!\w)",
            re.IGNORECASE,
        ),
    ),
    (
        "tax_code",
        re.compile(
            r"(?:m[aã]\s*s[oố]\s*thu[eế]|ma\s+so\s+thue|mst|tax\s*code)"
            r"\s*[:#-]?\s*\d{10}(?:-\d{3})?",
            re.IGNORECASE,
        ),
    ),
    (
        "address",
        re.compile(
            r"(?:[dđ][iị]a\s*ch[iỉ]|dia\s*chi|address)\s*[:#-]\s*[^;|\n]{4,180}"
            r"|(?<!\w)\d{1,5}(?:[/-]\d{1,5})?\s+"
            r"(?:(?i:(?:đường|duong|phố|pho|ngõ|ngo|hẻm|hem|ấp|ap|thôn|thon)"
            r"\s+[^;|\n]{2,160})|"
            r"(?:[A-ZÀ-ỴĐ][a-zà-ỹđ]{1,30}|[A-ZÀ-ỴĐ]{2,30})"
            r"(?:\s+(?:[A-ZÀ-ỴĐ][a-zà-ỹđ]{1,30}|[A-ZÀ-ỴĐ]{2,30})){1,5})"
            r"(?!\w)",
        ),
    ),
    (
        "email",
        re.compile(
            r"(?<![\w.+-])[A-Z0-9.!#$%&'*+/=?^_`{|}~-]+@"
            r"[A-Z0-9](?:[A-Z0-9.-]{0,251}[A-Z0-9])?\.[A-Z]{2,63}(?![\w.-])",
            re.IGNORECASE,
        ),
    ),
    (
        "phone",
        re.compile(r"(?<!\d)(?:\+?84|0)(?:[\s().-]*\d){8,10}(?!\d)"),
    ),
    (
        "bank_account",
        re.compile(
            r"(?:s[oố]\s*t[aà]i\s*kho[aả]n|so\s+tai\s+khoan|stk|bank\s*account)"
            r"\s*[:#-]?\s*\d(?:[\s.-]*\d){5,18}",
            re.IGNORECASE,
        ),
    ),
    ("document_number", re.compile(r"(?<!\d)\d{12}(?!\d)")),
)


class AnonymizationExportError(ValueError):
    """Raised when a generated workbook still contains confidential content."""

    def __init__(self, matched_categories: Iterable[str]) -> None:
        self.matched_categories = tuple(matched_categories)
        super().__init__(
            "Cannot export workbook with confidential values in: "
            + ", ".join(self.matched_categories)
        )


class AnonymizationUnsupportedLayerError(ValueError):
    """Raised when a workbook layer cannot be scanned safely before export."""

    def __init__(self, layer: str) -> None:
        self.layer = str(layer or "unknown")
        super().__init__(
            f"Cannot anonymize workbook safely: unsupported layer {self.layer}"
        )


@dataclass(frozen=True)
class AnonymizedWorkbook:
    content: bytes
    filename: str
    replaced_categories: tuple[str, ...]
    warnings: tuple[str, ...] = ()
    replaced_layers: tuple[str, ...] = ()


class AnonymizationSession:
    def __init__(self, session_id: str, secret: str | bytes) -> None:
        self.session_id = str(session_id or "").strip()
        if not self.session_id:
            raise ValueError("session_id is required")
        secret_bytes = secret if isinstance(secret, bytes) else str(secret or "").encode()
        if not secret_bytes:
            raise ValueError("secret is required")
        self._secret = secret_bytes
        self._replacements: dict[tuple[str, str], str] = {}

    def replace(self, category: str, value: Any):
        normalized_category = _validate_category(category)
        if value is None:
            return None
        source = str(value)
        if not source.strip():
            return source
        canonical = source.strip().casefold()
        cache_key = (normalized_category, canonical)
        if cache_key not in self._replacements:
            digest = hmac.new(
                self._secret,
                f"{self.session_id}\0{normalized_category}\0{canonical}".encode(
                    "utf-8"
                ),
                hashlib.sha256,
            ).digest()
            self._replacements[cache_key] = _replacement_for(
                normalized_category,
                source,
                digest,
            )
        return self._replacements[cache_key]

    anonymize = replace


def scan_confidential_values(
    payload: Any,
    confidential_values: Mapping[str, Iterable[Any]],
) -> tuple[str, ...]:
    searchable_values = tuple(
        value.casefold() for value in _iter_text_values(payload) if value.strip()
    )
    matches: list[str] = []
    for category in ANONYMIZATION_CATEGORIES:
        if category not in confidential_values:
            continue
        _validate_category(category)
        originals = (
            str(value).strip().casefold()
            for value in _confidential_value_items(confidential_values[category])
            if value is not None and str(value).strip()
        )
        if any(
            _contains_confidential(candidate, original)
            for original in originals
            for candidate in searchable_values
        ):
            matches.append(category)
    return tuple(matches)


def discover_pii_values(payload: Any) -> dict[str, tuple[str, ...]]:
    """Discover high-confidence PII tokens without relying on column names."""
    discovered: dict[str, list[str]] = {}
    seen: dict[str, set[str]] = {}
    patterns = (
        *tuple(_PII_PATTERNS.items()),
        ("counterparty", _VIETNAMESE_NAME_PATTERN),
        ("address", _VIETNAMESE_ADDRESS_PATTERN),
        ("document_number", _CCCD_PATTERN),
    )
    for candidate in _iter_text_values(payload):
        text = str(candidate or "")
        for category, pattern in patterns:
            for match in pattern.finditer(text):
                value = next(
                    (group for group in match.groups() if group is not None),
                    match.group(0),
                )
                value = value.strip().rstrip(".,;:")
                if not value or _is_anonymized_token(text, match.start(), category, value):
                    continue
                if category == "phone":
                    digit_count = len(re.sub(r"\D", "", value))
                    if digit_count < 9 or digit_count > 11:
                        continue
                canonical = value.casefold()
                category_seen = seen.setdefault(category, set())
                if canonical in category_seen:
                    continue
                category_seen.add(canonical)
                discovered.setdefault(category, []).append(value)
    return {category: tuple(values) for category, values in discovered.items()}


def _scan_export_pii_independently(payload: Any) -> tuple[str, ...]:
    matches: set[str] = set()
    for candidate in _iter_text_values(payload):
        text = str(candidate or "")
        for category, pattern in _POST_SCAN_PII_PATTERNS:
            if category in matches:
                continue
            for match in pattern.finditer(text):
                value = match.group(0)
                if category == "email" and value.casefold().endswith("@example.invalid"):
                    continue
                prefix = text[max(0, match.start() - 16) : match.start()].upper()
                if category == "address" and "ADDRESS-" in value.upper():
                    continue
                generated_prefixes = {
                    "phone": ("PHONE-", "TAX-", "BANK-", "DOC-"),
                    "document_number": ("DOC-",),
                }.get(category, ())
                if any(prefix.endswith(marker) for marker in generated_prefixes):
                    continue
                matches.add(category)
                break
    return tuple(
        category for category in ANONYMIZATION_CATEGORIES if category in matches
    )


def _is_anonymized_token(
    text: str, start: int, category: str, value: str
) -> bool:
    if category == "email" and value.casefold().endswith("@example.invalid"):
        return True
    if re.fullmatch(
        rf"(?:{'|'.join(map(re.escape, _TEXT_PREFIXES.values()))})-[A-F0-9]{{12}}",
        value,
        flags=re.IGNORECASE,
    ):
        return True
    prefix = text[max(0, start - 16) : start].upper()
    if any(prefix.endswith(marker) for marker in ("PHONE-", "TAX-", "BANK-", "DOC-")):
        return True
    expected = {
        "phone": "PHONE-",
        "tax_code": "TAX-",
        "bank_account": "BANK-",
    }.get(category)
    return bool(expected and prefix.endswith(expected))


def _merge_confidential_values(
    supplied: Mapping[str, Iterable[Any]], discovered: Mapping[str, Iterable[Any]]
) -> dict[str, tuple[Any, ...]]:
    merged: dict[str, list[Any]] = {}
    seen: dict[str, set[str]] = {}
    for source in (supplied, discovered):
        for category, values in source.items():
            normalized_category = _validate_category(category)
            for value in _confidential_value_items(values):
                if value is None or not str(value).strip():
                    continue
                canonical = str(value).strip().casefold()
                category_seen = seen.setdefault(normalized_category, set())
                if canonical in category_seen:
                    continue
                category_seen.add(canonical)
                merged.setdefault(normalized_category, []).append(value)
    return {category: tuple(values) for category, values in merged.items()}


def anonymize_workbook_bytes(
    *,
    filename: str,
    content: bytes,
    session: AnonymizationSession,
    confidential_values: Mapping[str, Iterable[Any]],
    full_document_numbers: bool = False,
    analyzed_sheet_name: str | None = None,
    analyzed_header_row_index: int | None = None,
    analyzed_headers: Iterable[Any] | None = None,
) -> AnonymizedWorkbook:
    """Return a newly serialized, scanner-gated workbook without writing the source."""
    normalized_filename = str(filename or "").strip()
    extension = Path(normalized_filename).suffix.lower()
    if extension not in {".xlsx", ".xls"}:
        raise ValueError("Only .xlsx and .xls workbooks can be anonymized")
    if not isinstance(content, bytes) or not content:
        raise ValueError("Workbook content is required")

    if extension == ".xlsx":
        _validate_xlsx_archive(content)
    source_values = _workbook_values(content, extension)
    discovered_values = discover_pii_values(source_values)
    all_confidential_values = _merge_confidential_values(
        confidential_values,
        discovered_values,
    )
    active_values = _active_confidential_values(
        all_confidential_values,
        full_document_numbers=full_document_numbers,
    )
    if discovered_values.get("document_number"):
        active_values["document_number"] = discovered_values["document_number"]
    if extension == ".xlsx":
        exported, replaced_categories, replaced_layers = _anonymize_xlsx(
            content,
            session,
            active_values,
            analyzed_sheet_name=analyzed_sheet_name,
            analyzed_header_row_index=analyzed_header_row_index,
            analyzed_headers=analyzed_headers,
        )
        warnings: tuple[str, ...] = ()
    else:
        raise AnonymizationUnsupportedLayerError("xls_metadata_layers")

    exported_values = _workbook_values(exported, extension)
    matches = scan_confidential_values(exported_values, active_values)
    try:
        independent_matches = _scan_export_pii_independently(exported_values)
    except Exception as exc:
        raise AnonymizationUnsupportedLayerError("pii_post_scan") from exc
    failed_categories = tuple(dict.fromkeys((*matches, *independent_matches)))
    if failed_categories:
        raise AnonymizationExportError(failed_categories)
    return AnonymizedWorkbook(
        content=exported,
        filename=_export_filename(normalized_filename),
        replaced_categories=tuple(
            category for category in ANONYMIZATION_CATEGORIES if category in replaced_categories
        ),
        warnings=warnings,
        replaced_layers=tuple(sorted(replaced_layers)),
    )


def _active_confidential_values(
    confidential_values: Mapping[str, Iterable[Any]],
    *,
    full_document_numbers: bool,
) -> dict[str, tuple[Any, ...]]:
    active: dict[str, tuple[Any, ...]] = {}
    for category, values in confidential_values.items():
        normalized_category = _validate_category(category)
        if normalized_category == "document_number" and not full_document_numbers:
            continue
        active[normalized_category] = tuple(_confidential_value_items(values))
    return active


def _confidential_value_items(values: Iterable[Any] | Any) -> Iterable[Any]:
    if isinstance(values, (str, bytes, bytearray)):
        return (values,)
    return values


def _anonymize_xlsx(
    content: bytes,
    session: AnonymizationSession,
    confidential_values: Mapping[str, Iterable[Any]],
    *,
    analyzed_sheet_name: str | None,
    analyzed_header_row_index: int | None,
    analyzed_headers: Iterable[Any] | None,
) -> tuple[bytes, set[str], set[str]]:
    _validate_xlsx_archive(content)
    source_workbook = load_workbook(BytesIO(content), data_only=False)
    if list(getattr(source_workbook, "_external_links", ())):
        raise AnonymizationUnsupportedLayerError("external_links")
    _reject_unsupported_xlsx_binary_layers(content)

    visible_sheets = [
        worksheet
        for worksheet in source_workbook.worksheets
        if worksheet.sheet_state == "visible"
    ]
    if not visible_sheets:
        raise AnonymizationUnsupportedLayerError("no_visible_worksheets")

    source_worksheet = _analyzed_xlsx_worksheet(
        source_workbook,
        visible_sheets,
        analyzed_sheet_name=analyzed_sheet_name,
        analyzed_header_row_index=analyzed_header_row_index,
        analyzed_headers=analyzed_headers,
    )

    # Export only the analyzed sheet. Other visible sheets are not part of the
    # Student Assistant output and could otherwise bypass its value inventory.
    workbook = Workbook()
    replaced_categories: set[str] = set()
    replaced_layers = _removed_xlsx_layers(source_workbook)
    if len(visible_sheets) > 1:
        replaced_layers.add("non_analyzed_visible_sheets_removed")
    worksheet = workbook.worksheets[0]
    worksheet.title = source_worksheet.title
    for row in source_worksheet.iter_rows():
        for source_cell in row:
            value = source_cell.value
            replacement = _replacement_for_cell(
                value, session, confidential_values
            )
            if replacement is not None:
                value, categories = replacement
                replaced_categories.update(categories)
                replaced_layers.add("cell_values")
            worksheet.cell(
                row=source_cell.row,
                column=source_cell.column,
                value=value,
            )
    replaced_layers.update(_sanitize_xlsx_metadata(workbook, session, confidential_values))
    stream = BytesIO()
    workbook.save(stream)
    exported, archive_categories = _sanitize_xlsx_archive(
        stream.getvalue(), session, confidential_values
    )
    if archive_categories:
        replaced_categories.update(archive_categories)
        replaced_layers.add("ooxml_text_parts")
    return exported, replaced_categories, replaced_layers


def _analyzed_xlsx_worksheet(
    workbook: Any,
    visible_sheets: list[Any],
    *,
    analyzed_sheet_name: str | None,
    analyzed_header_row_index: int | None,
    analyzed_headers: Iterable[Any] | None,
) -> Any:
    context_values = (
        analyzed_sheet_name,
        analyzed_header_row_index,
        analyzed_headers,
    )
    if all(value is None for value in context_values):
        if len(visible_sheets) == 1:
            return visible_sheets[0]
        raise AnonymizationUnsupportedLayerError("analyzed_sheet_context")
    if any(value is None for value in context_values):
        raise AnonymizationUnsupportedLayerError("analyzed_sheet_context")

    sheet_name = str(analyzed_sheet_name).strip()
    if not sheet_name or sheet_name not in workbook.sheetnames:
        raise AnonymizationUnsupportedLayerError("analyzed_sheet_not_found")
    if isinstance(analyzed_header_row_index, bool) or not isinstance(
        analyzed_header_row_index, int
    ) or analyzed_header_row_index < 0:
        raise AnonymizationUnsupportedLayerError("analyzed_sheet_header_context")
    if isinstance(analyzed_headers, (str, bytes, bytearray)):
        raise AnonymizationUnsupportedLayerError("analyzed_sheet_header_context")

    worksheet = workbook[sheet_name]
    header_row = next(
        worksheet.iter_rows(
            min_row=analyzed_header_row_index + 1,
            max_row=analyzed_header_row_index + 1,
            values_only=True,
        ),
        (),
    )
    actual_headers = tuple(
        "" if value is None else str(value).strip() for value in header_row
    )
    expected_headers = tuple(
        "" if value is None else str(value).strip() for value in analyzed_headers
    )
    if actual_headers != expected_headers:
        raise AnonymizationUnsupportedLayerError("analyzed_sheet_header_mismatch")
    return worksheet


def _reject_unsupported_xlsx_binary_layers(content: bytes) -> None:
    with ZipFile(BytesIO(content), "r") as archive:
        for name in archive.namelist():
            if name.casefold().startswith(("xl/media/", "xl/embeddings/", "xl/activex/")):
                raise AnonymizationUnsupportedLayerError("embedded_binary_objects")


def _removed_xlsx_layers(workbook: Any) -> set[str]:
    layers = {"workbook_properties", "defined_names"}
    if any(sheet.sheet_state != "visible" for sheet in workbook.worksheets):
        layers.update({"hidden_sheets_removed", "hidden_sheet_cells"})
    if any(
        cell.comment is not None
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    ):
        layers.add("comments")
    if any(
        cell.hyperlink is not None
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    ):
        layers.add("hyperlinks")
    custom_properties = getattr(workbook, "custom_doc_props", None)
    if custom_properties and len(custom_properties):
        layers.add("custom_document_properties")
    return layers


def _validate_xlsx_archive(content: bytes) -> None:
    try:
        with ZipFile(BytesIO(content), "r") as archive:
            entries = archive.infolist()
    except (BadZipFile, OSError) as exc:
        raise AnonymizationUnsupportedLayerError("malformed_xlsx_archive") from exc

    if len(entries) > MAX_XLSX_ARCHIVE_ENTRIES:
        raise AnonymizationUnsupportedLayerError("archive_entry_count")

    total_uncompressed = 0
    total_compressed = 0
    seen_names: set[str] = set()
    for info in entries:
        normalized_name = info.filename.replace("\\", "/")
        parts = tuple(part for part in normalized_name.split("/") if part)
        if (
            not normalized_name
            or normalized_name.startswith("/")
            or ".." in parts
            or normalized_name in seen_names
        ):
            raise AnonymizationUnsupportedLayerError("unsafe_archive_entry")
        seen_names.add(normalized_name)
        if info.flag_bits & 0x1:
            raise AnonymizationUnsupportedLayerError("encrypted_archive_entry")
        if info.is_dir():
            continue
        if info.file_size > MAX_XLSX_ENTRY_UNCOMPRESSED_BYTES:
            raise AnonymizationUnsupportedLayerError("archive_entry_size")
        total_uncompressed += info.file_size
        total_compressed += info.compress_size

    if total_uncompressed > MAX_XLSX_UNCOMPRESSED_BYTES:
        raise AnonymizationUnsupportedLayerError("archive_uncompressed_size")
    if total_uncompressed and total_uncompressed / max(total_compressed, 1) > MAX_XLSX_COMPRESSION_RATIO:
        raise AnonymizationUnsupportedLayerError("archive_compression_ratio")


def _anonymize_xls(
    content: bytes,
    session: AnonymizationSession,
    confidential_values: Mapping[str, Iterable[Any]],
) -> tuple[bytes, set[str], set[str]]:
    source = open_workbook(file_contents=content, formatting_info=True)
    workbook = copy_xls_workbook(source)
    replaced_categories: set[str] = set()
    replaced_layers: set[str] = set()
    for sheet_index, source_sheet in enumerate(source.sheets()):
        target_sheet = workbook.get_sheet(sheet_index)
        for row_index in range(source_sheet.nrows):
            for column_index in range(source_sheet.ncols):
                replacement = _replacement_for_cell(
                    source_sheet.cell_value(row_index, column_index),
                    session,
                    confidential_values,
                )
                if replacement is not None:
                    target_sheet.write(row_index, column_index, replacement[0])
                    # xlutils copies the original XF table; restore the source cell's entry after writing.
                    target_sheet._Worksheet__rows[row_index]._Row__cells[
                        column_index
                    ].xf_idx = source_sheet.cell(row_index, column_index).xf_index
                    replaced_categories.add(replacement[1])
                    replaced_layers.add("cell_values")
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue(), replaced_categories, replaced_layers


def _sanitize_xlsx_metadata(
    workbook: Any,
    session: AnonymizationSession,
    confidential_values: Mapping[str, Iterable[Any]],
) -> set[str]:
    """Remove metadata channels that are not safe to preserve for student files."""
    replaced_layers: set[str] = set()

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.comment is not None:
                    comment_text, text_category = _replace_known_text(
                        cell.comment.text, session, confidential_values
                    )
                    comment_author, author_category = _replace_known_text(
                        cell.comment.author, session, confidential_values
                    )
                    if text_category or author_category:
                        comment = Comment(comment_text, comment_author)
                        comment.width = cell.comment.width
                        comment.height = cell.comment.height
                        cell.comment = comment
                        replaced_layers.add("comments")
                if cell.hyperlink is not None:
                    cell.hyperlink = None
                    replaced_layers.add("hyperlinks")

    properties = getattr(workbook, "properties", None)
    if properties is not None:
        property_names = (
            "creator",
            "lastModifiedBy",
            "title",
            "subject",
            "description",
            "keywords",
            "category",
            "contentStatus",
            "identifier",
            "language",
            "version",
            "revision",
        )
        if any(getattr(properties, name, None) for name in property_names):
            replaced_layers.add("workbook_properties")
        for name in property_names:
            if hasattr(properties, name):
                setattr(properties, name, "")
        properties.created = datetime(1980, 1, 1)
        properties.modified = datetime(1980, 1, 1)
        properties.lastPrinted = None

    custom_properties = getattr(workbook, "custom_doc_props", None)
    if custom_properties and len(custom_properties):
        custom_properties.props.clear()
        replaced_layers.add("custom_document_properties")

    defined_names = getattr(workbook, "defined_names", None)
    if defined_names:
        # Named ranges can preserve sensitive text even when no worksheet cell does.
        for name in list(defined_names):
            try:
                del defined_names[name]
            except (KeyError, TypeError):
                continue
        replaced_layers.add("defined_names")

    return replaced_layers


def _sanitize_xlsx_archive(
    content: bytes,
    session: AnonymizationSession,
    confidential_values: Mapping[str, Iterable[Any]],
) -> tuple[bytes, set[str]]:
    """Sanitize OOXML text parts; reject binary layers that cannot be inspected."""
    output = BytesIO()
    replaced_categories: set[str] = set()
    with ZipFile(BytesIO(content), "r") as source, ZipFile(output, "w") as target:
        for info in source.infolist():
            data = source.read(info.filename)
            lower_name = info.filename.casefold()
            if lower_name.startswith(("xl/media/", "xl/embeddings/", "xl/activex/")):
                raise AnonymizationUnsupportedLayerError("embedded_binary_objects")
            if lower_name.endswith((".xml", ".rels", ".txt")):
                try:
                    text = data.decode("utf-8")
                except UnicodeDecodeError as exc:
                    raise AnonymizationUnsupportedLayerError(
                        "non_utf8_ooxml_text"
                    ) from exc
                text, categories = _replace_known_ooxml_text(
                    text, session, confidential_values
                )
                if categories:
                    data = text.encode("utf-8")
                    replaced_categories.update(categories)
            target.writestr(info, data)
    return output.getvalue(), replaced_categories


def _replace_known_ooxml_text(
    text: str,
    session: AnonymizationSession,
    confidential_values: Mapping[str, Iterable[Any]],
) -> tuple[str, set[str]]:
    encoded = text.encode("utf-8")
    try:
        for _, namespace in ElementTree.iterparse(
            BytesIO(encoded), events=("start-ns",)
        ):
            prefix, uri = namespace
            try:
                ElementTree.register_namespace(prefix or "", uri)
            except ValueError:
                continue
        root = ElementTree.fromstring(encoded)
    except ElementTree.ParseError as exc:
        raise AnonymizationUnsupportedLayerError("malformed_ooxml_text") from exc

    replaced_categories: set[str] = set()
    for element in root.iter():
        element_name = _xml_local_name(element.tag)
        if element_name not in _OOXML_FORMULA_ELEMENTS and element.text:
            element.text, categories = _replace_all_known_text(
                element.text, session, confidential_values
            )
            replaced_categories.update(categories)
        replaceable_attributes = _OOXML_REPLACEABLE_ATTRIBUTES.get(element_name, set())
        for attribute_name, attribute_value in list(element.attrib.items()):
            if _xml_local_name(attribute_name) not in replaceable_attributes:
                continue
            element.attrib[attribute_name], categories = _replace_all_known_text(
                attribute_value, session, confidential_values
            )
            replaced_categories.update(categories)
    return ElementTree.tostring(root, encoding="unicode"), replaced_categories


def _xml_local_name(value: Any) -> str:
    return str(value).rsplit("}", 1)[-1].rsplit(":", 1)[-1]


def _ooxml_text_values(text: str) -> list[str]:
    try:
        root = ElementTree.fromstring(text.encode("utf-8"))
    except ElementTree.ParseError as exc:
        raise AnonymizationUnsupportedLayerError("malformed_ooxml_text") from exc
    values: list[str] = []
    for element in root.iter():
        element_name = _xml_local_name(element.tag)
        if element.text and element.text.strip():
            values.append(element.text)
        if element_name in _OOXML_RICH_TEXT_CONTAINERS:
            rich_text_parts = [
                str(descendant.text or "")
                for descendant in element.iter()
                if _xml_local_name(descendant.tag) == "t" and descendant.text
            ]
            if len(rich_text_parts) > 1:
                values.append("".join(rich_text_parts))
        values.extend(element.attrib.values())
    return values


def _replace_known_text(
    value: Any,
    session: AnonymizationSession,
    confidential_values: Mapping[str, Iterable[Any]],
) -> tuple[str, str | None]:
    text, categories = _replace_all_known_text(value, session, confidential_values)
    return text, next(iter(categories), None)


def _replace_all_known_text(
    value: Any,
    session: AnonymizationSession,
    confidential_values: Mapping[str, Iterable[Any]],
) -> tuple[str, set[str]]:
    text = str(value or "")
    matched_categories: set[str] = set()
    candidates = sorted(
        (
            (category, str(original or "").strip())
            for category in ANONYMIZATION_CATEGORIES
            for original in confidential_values.get(category, ())
            if str(original or "").strip()
        ),
        key=lambda item: len(item[1]),
        reverse=True,
    )
    for category, source in candidates:
        if _contains_confidential(text, source):
            text = _replace_confidential(
                text, source, session.replace(category, source)
            )
            matched_categories.add(category)
    return text, matched_categories


def _contains_confidential(candidate: str, source: str) -> bool:
    normalized_source = str(source or "").strip()
    if not normalized_source:
        return False
    if len(normalized_source) <= 2:
        return bool(
            re.search(
                rf"(?<!\w){re.escape(normalized_source)}(?!\w)",
                str(candidate or ""),
                flags=re.IGNORECASE,
            )
        )
    return normalized_source.casefold() in str(candidate or "").casefold()


def _replace_confidential(text: str, source: str, replacement: str) -> str:
    pattern = re.escape(source)
    if len(source) <= 2:
        pattern = rf"(?<!\w){pattern}(?!\w)"
    return re.sub(pattern, replacement, text, flags=re.IGNORECASE)


def _replacement_for_cell(
    value: Any,
    session: AnonymizationSession,
    confidential_values: Mapping[str, Iterable[Any]],
) -> tuple[str, set[str]] | None:
    if not isinstance(value, str) or value.startswith("="):
        return None
    replacement, categories = _replace_all_known_text(
        value, session, confidential_values
    )
    return (replacement, categories) if categories else None


def _workbook_values(content: bytes, extension: str) -> list[Any]:
    if extension == ".xlsx":
        workbook = load_workbook(BytesIO(content), data_only=False)
        payload: list[Any] = []
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    payload.append(cell.value)
                    if cell.comment is not None:
                        payload.extend((cell.comment.text, cell.comment.author))
                    if cell.hyperlink is not None:
                        payload.append(cell.hyperlink.target)
        properties = getattr(workbook, "properties", None)
        if properties is not None:
            payload.extend(
                getattr(properties, name, None)
                for name in (
                    "creator",
                    "lastModifiedBy",
                    "title",
                    "subject",
                    "description",
                    "keywords",
                    "category",
                    "contentStatus",
                    "identifier",
                    "language",
                    "version",
                    "revision",
                )
            )
        custom_properties = getattr(workbook, "custom_doc_props", None)
        if custom_properties:
            payload.extend(
                (getattr(item, "name", None), getattr(item, "value", None))
                for item in custom_properties
            )
        defined_names = getattr(workbook, "defined_names", None)
        if defined_names:
            payload.extend(
                (getattr(item, "name", None), getattr(item, "attr_text", None))
                for item in defined_names.values()
            )
        with ZipFile(BytesIO(content), "r") as archive:
            for name in archive.namelist():
                lower_name = name.casefold()
                if not lower_name.endswith((".xml", ".rels", ".txt")):
                    continue
                archive_text = archive.read(name).decode("utf-8")
                if lower_name.endswith((".xml", ".rels")):
                    payload.extend(_ooxml_text_values(archive_text))
                else:
                    payload.append(archive_text)
        return payload
    workbook = open_workbook(file_contents=content)
    return [
        [
            [sheet.cell_value(row_index, column_index) for column_index in range(sheet.ncols)]
            for row_index in range(sheet.nrows)
        ]
        for sheet in workbook.sheets()
    ]


def _export_filename(filename: str) -> str:
    source = Path(filename)
    return f"{source.stem}-anonymized{source.suffix.lower()}"


def _replacement_for(category: str, source: str, digest: bytes) -> str:
    token = digest.hex()[:12].upper()
    if category in _TEXT_PREFIXES:
        return f"{_TEXT_PREFIXES[category]}-{token}"
    if category == "email":
        return f"student-{digest.hex()[:12]}@example.invalid"
    if category in _NUMERIC_PREFIXES:
        return f"{_NUMERIC_PREFIXES[category]}-{_numeric_token(source, digest)}"
    raise ValueError(f"Unsupported anonymization category: {category}")


def _numeric_token(source: str, digest: bytes) -> str:
    source_digits = "".join(character for character in source if character.isdigit())
    length = max(8, len(source_digits))
    leading_zeroes = len(source_digits) - len(source_digits.lstrip("0"))
    leading_zeroes = min(leading_zeroes, max(0, length - 1))
    generated = "".join(str(byte % 10) for byte in digest)
    while len(generated) < length:
        generated += generated
    return ("0" * leading_zeroes + generated)[:length]


def _validate_category(category: str) -> str:
    normalized = str(category or "").strip().lower()
    if normalized not in ANONYMIZATION_CATEGORIES:
        raise ValueError(f"Unsupported anonymization category: {category}")
    return normalized


def _iter_text_values(value: Any):
    if isinstance(value, str):
        yield value
    elif isinstance(value, Mapping):
        for item in value.values():
            yield from _iter_text_values(item)
    elif isinstance(value, Iterable) and not isinstance(value, (bytes, bytearray)):
        for item in value:
            yield from _iter_text_values(item)
    elif value is not None and not isinstance(value, (bytes, bytearray)):
        yield str(value)
