from __future__ import annotations

from dataclasses import dataclass
from copy import copy
from datetime import date, datetime
from decimal import Decimal
from itertools import islice
from pathlib import Path
import re
from struct import pack_into, unpack_from
from typing import Any

import openpyxl
import xlrd
import xlwt
from openpyxl.styles.numbers import is_date_format
from xlrd.compdoc import CompDoc
from xlwt.CompoundDoc import XlsDoc
from xlutils.filter import XLRDReader, XLWTWriter, process

from app.normalization import is_blank, normalize_header
from app.parsing import parse_date


SMART_PURCHASE_HEADERS = frozenset(
    {
        "sr_hd",
        "soct",
        "ngayct",
        "so_hd",
        "tendm",
        "mathang",
        "phan_loai",
        "donvi",
        "luong",
        "dgvnd",
        "ttvnd",
        "tkthue",
        "ts_gtgt",
        "thuevnd",
        "makh",
        "tenkh",
    }
)
EXCEL_SAFE_INTEGER_LIMIT = 2**53

_ZERO_ONLY_NUMBER_FORMAT = re.compile(r"0+")
_BIFF_BOUNDSHEET = 0x0085
_BIFF_CONTINUE = 0x003C
_BIFF_EOF = 0x000A
_BIFF_EXTSST = 0x00FF
_BIFF_SST = 0x00FC
_BIFF_NAME_DEPENDENCIES = {0x0017, 0x0018, 0x0023, 0x01AE}


class InputReadError(Exception):
    """Structured failure when reading Excel input."""

    def __init__(self, code: str, message: str) -> None:
        self.code = code
        self.message = message
        super().__init__(message)


@dataclass(frozen=True)
class InputTable:
    headers: list[str]
    rows: list[dict[str, Any]]
    sheet_name: str | None = None
    header_row_index: int = 0


@dataclass(frozen=True)
class TemplateWorkbook:
    path: Path
    sheet_name: str
    header_row_index: int
    headers: list[str]
    preamble_rows: list[list[Any]]


def find_header_row(sheet: xlrd.sheet.Sheet) -> int:
    best_score: tuple[int, int, int] | None = None
    best_row = 0
    max_scan_rows = min(sheet.nrows, 30)

    for row_idx in range(max_scan_rows):
        values = [str(sheet.cell_value(row_idx, col)).strip() for col in range(sheet.ncols)]
        non_empty = sum(1 for value in values if value)
        keyword_hits = sum(
            1
            for value in values
            if any(
                keyword in value.lower()
                for keyword in (
                    "ngày",
                    "mã",
                    "số",
                    "hình thức",
                    "phương thức",
                    "đơn giá",
                    "thành tiền",
                )
            )
        )
        score = (keyword_hits, non_empty, -row_idx)
        if best_score is None or score > best_score:
            best_score = score
            best_row = row_idx

    return best_row


def read_template(path: Path) -> TemplateWorkbook:
    book = xlrd.open_workbook(str(path), formatting_info=False)
    sheet = book.sheet_by_index(0)
    header_row_index = find_header_row(sheet)
    headers = [str(sheet.cell_value(header_row_index, col)).strip() for col in range(sheet.ncols)]
    preamble_rows = [
        [sheet.cell_value(row, col) for col in range(sheet.ncols)] for row in range(header_row_index)
    ]
    return TemplateWorkbook(
        path=path,
        sheet_name=sheet.name,
        header_row_index=header_row_index,
        headers=headers,
        preamble_rows=preamble_rows,
    )


def read_input_table(path: Path) -> InputTable:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _read_xlsx(path)
    if suffix == ".xls":
        return _read_xls(path)
    raise InputReadError("unsupported_format", "Only .xls and .xlsx files are supported.")


def read_purchase_adjustment_context(
    path: Path,
    detail_table: InputTable,
) -> list[dict[str, Any]]:
    """Return review-only invoice adjustment metadata from supported partner workbooks."""
    if path.suffix.lower() != ".xlsx" or detail_table.sheet_name != "Smart_KTSC_OK":
        return []
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []
    try:
        if "HoaDon_TongQuat" not in workbook.sheetnames:
            return []
        summary = workbook["HoaDon_TongQuat"]
        header_row_index = None
        header_map: dict[str, int] = {}
        for row_index, row in enumerate(
            summary.iter_rows(min_row=1, max_row=min(summary.max_row, 30), values_only=True),
            start=1,
        ):
            candidate = {
                normalize_header(value): column_index
                for column_index, value in enumerate(row)
                if not is_blank(value)
            }
            if {"mst_nguoi_ban", "ky_hieu", "so", "trang_thai_hd"}.issubset(candidate):
                header_row_index = row_index
                header_map = candidate
                break
        if header_row_index is None:
            return []

        detail_keys: dict[tuple[str, str, str], int] = {}
        for row in detail_table.rows:
            key = (
                _review_key_value(row.get("MADTPNCO") or row.get("MAKH")),
                _review_key_value(row.get("SR_HD")),
                _review_key_value(row.get("SO_HD") or row.get("SOCT")),
            )
            detail_keys[key] = detail_keys.get(key, 0) + 1

        contexts: list[dict[str, Any]] = []
        for values in summary.iter_rows(min_row=header_row_index + 1, values_only=True):
            status = str(values[header_map["trang_thai_hd"]] or "").strip()
            if "dieu_chinh" not in normalize_header(status):
                continue
            key = (
                _review_key_value(values[header_map["mst_nguoi_ban"]]),
                _review_key_value(values[header_map["ky_hieu"]]),
                _review_key_value(values[header_map["so"]]),
            )
            if key not in detail_keys:
                continue
            reference = _adjustment_reference(status)
            contexts.append(
                {
                    "supplier_tax_code": key[0],
                    "invoice_symbol": key[1],
                    "invoice_number": key[2],
                    "status": status,
                    "adjusts_invoice": reference,
                    "detail_row_count": detail_keys[key],
                    "requires_user_review": True,
                }
            )
        return contexts
    finally:
        workbook.close()


def _review_key_value(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value or "").strip()


def _adjustment_reference(status: str) -> dict[str, str] | None:
    normalized = normalize_header(status)
    match = re.search(
        r"ky_hieu_hoa_don_(.+?)_so_(.+?)_ngay_lap_(\d{1,2})_(\d{1,2})_(\d{4})",
        normalized,
        flags=re.IGNORECASE,
    )
    if not match:
        return None
    return {
        "invoice_symbol": match.group(1).strip().upper(),
        "invoice_number": match.group(2).strip(),
        "invoice_date": f"{match.group(3)}/{match.group(4)}/{match.group(5)}",
    }


def score_header_rows(rows: list[Any]) -> tuple[int, int]:
    best_score: tuple[int, int] | None = None
    best_idx = 0
    for idx, row in enumerate(rows[:30]):
        values = ["" if value is None else str(value).strip() for value in row]
        non_empty = sum(1 for value in values if value)
        keyword_hits = sum(
            1
            for value in values
            if any(
                keyword in value.lower()
                for keyword in ("mã", "ngày", "thời gian", "tên", "số lượng", "đơn giá", "thành tiền")
            )
        )
        score = (keyword_hits, non_empty)
        if best_score is None or score > best_score:
            best_score = score
            best_idx = idx
    return best_idx, best_score or (0, 0)


def is_smart_purchase_header(values: list[Any] | tuple[Any, ...]) -> bool:
    normalized = {normalize_header(value) for value in values if not is_blank(value)}
    return SMART_PURCHASE_HEADERS.issubset(normalized)


def _smart_purchase_header_index(rows: list[Any]) -> int | None:
    for idx, row in enumerate(rows[:30]):
        if is_smart_purchase_header(row):
            return idx
    return None


def _read_xlsx(path: Path) -> InputTable:
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise InputReadError("corrupt_xlsx", f"Cannot read Excel workbook: {exc}") from exc
    try:
        preferred_sheet = _known_purchase_detail_sheet(workbook)
        if preferred_sheet is not None:
            header_row = next(preferred_sheet.iter_rows(min_row=1, max_row=1), ())
            headers = _validated_headers([cell.value for cell in header_row])
            return InputTable(
                headers=headers,
                rows=_rows_to_records(
                    headers,
                    (
                        [_xlsx_input_cell_value(cell) for cell in row]
                        for row in preferred_sheet.iter_rows(min_row=2)
                    ),
                ),
                sheet_name=preferred_sheet.title,
                header_row_index=0,
            )

        best_name: str | None = None
        best_score = (-1, -1)
        best_header_idx = 0
        for name in workbook.sheetnames:
            sheet = workbook[name]
            rows = list(islice(sheet.iter_rows(values_only=True), 30))
            if not rows:
                continue
            smart_header_idx = _smart_purchase_header_index(rows)
            if smart_header_idx is not None:
                header_idx = smart_header_idx
                score = (1000, len([value for value in rows[header_idx] if not is_blank(value)]))
            else:
                header_idx, score = score_header_rows(rows)
            if score > best_score:
                best_score = score
                best_name = name
                best_header_idx = header_idx

        if not best_name:
            return InputTable(headers=[], rows=[], sheet_name=None, header_row_index=0)

        sheet = workbook[best_name]
        header_row = next(
            sheet.iter_rows(
                min_row=best_header_idx + 1,
                max_row=best_header_idx + 1,
            ),
            (),
        )
        headers = _validated_headers([cell.value for cell in header_row])
        records = _rows_to_records(
            headers,
            (
                [_xlsx_input_cell_value(cell) for cell in row]
                for row in sheet.iter_rows(min_row=best_header_idx + 2)
            ),
        )
        return InputTable(
            headers=headers,
            rows=records,
            sheet_name=best_name,
            header_row_index=best_header_idx,
        )
    finally:
        workbook.close()


def _known_purchase_detail_sheet(workbook: openpyxl.Workbook):
    """Select the partner's line-detail sheet only when its schema markers match."""
    if "Smart_KTSC_OK" not in workbook.sheetnames:
        return None
    sheet = workbook["Smart_KTSC_OK"]
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = {str(value).strip() for value in first_row if value is not None}
    required_markers = {"SR_HD", "SOCT", "MATHANG", "Phân loại", "TTVND"}
    return sheet if required_markers.issubset(headers) else None


def _read_xls(path: Path) -> InputTable:
    try:
        book = xlrd.open_workbook(str(path), formatting_info=True)
    except xlrd.XLRDError as exc:
        raise InputReadError("corrupt_xls", f"Corrupt or unsupported .xls file: {exc}") from exc
    except Exception as exc:
        raise InputReadError("corrupt_xls", f"Cannot read .xls file: {exc}") from exc

    best_name: str | None = None
    best_sheet: xlrd.sheet.Sheet | None = None
    best_score = (-1, -1)
    best_header_idx = 0

    for sheet in book.sheets():
        rows = [
            [sheet.cell_value(row, col) for col in range(sheet.ncols)]
            for row in range(min(sheet.nrows, 30))
        ]
        if not rows:
            continue
        smart_header_idx = _smart_purchase_header_index(rows)
        if smart_header_idx is not None:
            header_idx = smart_header_idx
            score = (1000, len([value for value in rows[header_idx] if not is_blank(value)]))
        else:
            header_idx, score = score_header_rows(rows)
        if score > best_score:
            best_score = score
            best_name = sheet.name
            best_sheet = sheet
            best_header_idx = header_idx

    if best_sheet is None:
        return InputTable(headers=[], rows=[], sheet_name=None, header_row_index=0)

    headers = _validated_headers(best_sheet.row_values(best_header_idx))
    records = _rows_to_records(
        headers,
        (
            [
                _xls_input_cell_value(book, best_sheet.cell(row, col))
                for col in range(best_sheet.ncols)
            ]
            for row in range(best_header_idx + 1, best_sheet.nrows)
        ),
    )
    return InputTable(
        headers=headers,
        rows=records,
        sheet_name=best_name,
        header_row_index=best_header_idx,
    )


def _validated_headers(values: list[Any]) -> list[str]:
    headers = ["" if value is None else str(value).strip() for value in values]
    first_column_by_normalized_header: dict[str, int] = {}
    for column_idx, header in enumerate(headers, start=1):
        normalized = normalize_header(header)
        if not normalized:
            continue
        first_column = first_column_by_normalized_header.get(normalized)
        if first_column is not None:
            raise InputReadError(
                "duplicate_headers",
                f"Duplicate normalized header '{normalized}' at columns "
                f"{first_column} and {column_idx}. Rename one column before importing.",
            )
        first_column_by_normalized_header[normalized] = column_idx
    return headers


def _xlsx_input_cell_value(cell: Any) -> Any:
    return _numeric_display_value(cell.value, cell.number_format)


def _xls_input_cell_value(book: xlrd.book.Book, cell: xlrd.sheet.Cell) -> Any:
    if cell.ctype != xlrd.XL_CELL_NUMBER:
        return cell.value
    xf = book.xf_list[cell.xf_index]
    number_format = book.format_map[xf.format_key].format_str
    return _numeric_display_value(cell.value, number_format)


def _numeric_display_value(value: Any, number_format: str) -> Any:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return value
    if abs(value) >= EXCEL_SAFE_INTEGER_LIMIT:
        raise InputReadError(
            "unsafe_numeric_precision",
            "Numeric cell exceeds Excel's safe precision. Format the source "
            "cell as text and upload again to prevent silent rounding.",
        )
    format_code = str(number_format or "").strip()
    if not _ZERO_ONLY_NUMBER_FORMAT.fullmatch(format_code):
        return value
    if isinstance(value, float) and not value.is_integer():
        return value
    integer_value = int(value)
    if integer_value < 0:
        return value
    return str(integer_value).zfill(len(format_code))



def _sanitize_output_sheet_name(name: str | None, fallback: str) -> str:
    raw = str(name or "").strip() or fallback
    cleaned = "".join("_" if ch in "\\/?*[]" else ch for ch in raw).strip()
    cleaned = cleaned[:31] or fallback[:31] or "Sheet1"
    return cleaned


def _rows_to_records(headers: list[str], rows: list[Any]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for row in rows:
        values = list(row)
        if all(is_blank(value) for value in values):
            continue
        record: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            record[header] = values[idx] if idx < len(values) else None
        if any(not is_blank(value) for value in record.values()):
            records.append(record)
    return records


def write_xls_from_template(
    template: TemplateWorkbook,
    output_rows: list[dict[str, Any]],
    output_path: Path,
    *,
    output_sheet_name: str | None = None,
) -> None:
    source_book = xlrd.open_workbook(str(template.path), formatting_info=True)
    source_sheet = source_book.sheet_by_name(template.sheet_name)
    writer = XLWTWriter()
    process(XLRDReader(source_book, str(template.path)), writer)
    workbook = writer.output[0][1]
    source_sheet_index = source_book.sheet_names().index(template.sheet_name)
    sheet = workbook.get_sheet(source_sheet_index)
    sheet.name = _sanitize_output_sheet_name(output_sheet_name, template.sheet_name)
    styles = list(writer.style_list)
    data_start_row = template.header_row_index + 1
    output_end_row = data_start_row + len(output_rows)

    if output_end_row > 65536:
        raise ValueError("Output exceeds the .xls row limit of 65,536 rows.")

    _trim_copied_template_data_rows(sheet, data_start_row)
    _restore_merged_covered_styles(
        source_sheet,
        sheet,
        styles,
        data_start_row=data_start_row,
    )
    for row_idx in range(data_start_row, max(output_end_row, 12)):
        _copy_row_layout(
            source_sheet,
            sheet,
            row_idx,
            fallback_row=data_start_row,
        )
    for record_idx, record in enumerate(output_rows, start=data_start_row):
        _copy_row_layout(
            source_sheet,
            sheet,
            record_idx,
            fallback_row=data_start_row,
        )
        for col_idx, header in enumerate(template.headers):
            if not header:
                continue
            value = record.get(header, "")
            template_style = _style_for_cell(
                source_sheet,
                styles,
                record_idx,
                col_idx,
                fallback_row=data_start_row,
            )
            cell_style = _xls_cell_style(value, template_style, header)
            sheet.write(
                record_idx,
                col_idx,
                _xls_cell_value(value, cell_style, header),
                cell_style,
            )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    _save_filled_template_workbook(template.path, workbook, output_path)


def _trim_copied_template_data_rows(sheet: xlwt.Worksheet.Worksheet, start_row: int) -> None:
    rows = sheet._Worksheet__rows
    for row_idx in [row_idx for row_idx in rows if row_idx >= start_row]:
        del rows[row_idx]
    sheet._Worksheet__merged_ranges = [
        merged for merged in sheet._Worksheet__merged_ranges if merged[0] < start_row
    ]
    sheet.first_used_row = min(rows, default=0)
    sheet.last_used_row = max(rows, default=0)


def _restore_merged_covered_styles(
    source_sheet: xlrd.sheet.Sheet,
    output_sheet: xlwt.Worksheet.Worksheet,
    styles: list[xlwt.Style.XFStyle],
    *,
    data_start_row: int,
) -> None:
    output_sheet._cell_overwrite_ok = True
    for row_low, row_high, col_low, col_high in source_sheet.merged_cells:
        if row_low >= data_start_row:
            continue
        for row_index in range(row_low, row_high):
            for col_index in range(col_low, col_high):
                if (row_index, col_index) == (row_low, col_low):
                    continue
                style = _style_for_cell(
                    source_sheet,
                    styles,
                    row_index,
                    col_index,
                    fallback_row=row_low,
                )
                output_sheet.write(row_index, col_index, "", style)


def _save_filled_template_workbook(
    template_path: Path,
    workbook: xlwt.Workbook,
    output_path: Path,
) -> None:
    template_bytes = template_path.read_bytes()
    compdoc = CompDoc(template_bytes)
    stream_memory, stream_offset, stream_length = compdoc.locate_named_stream("Workbook")
    if stream_memory is None:
        stream_memory, stream_offset, stream_length = compdoc.locate_named_stream("Book")
    if stream_memory is None:
        raise ValueError("The .xls template does not contain a Workbook stream.")

    source_stream = bytes(stream_memory[stream_offset : stream_offset + stream_length])
    generated_stream = workbook.get_biff_data()
    filled_stream = _merge_template_globals_with_generated_sheets(
        source_stream,
        generated_stream,
    )
    if stream_memory is not template_bytes or stream_offset <= 0:
        XlsDoc().save(str(output_path), filled_stream)
        return
    if len(filled_stream) > stream_length:
        # The original OLE stream capacity is not a safe limit for a larger
        # export. Rebuild the envelope while retaining the merged BIFF globals.
        XlsDoc().save(str(output_path), filled_stream)
        return

    output_bytes = bytearray(template_bytes)
    output_bytes[stream_offset : stream_offset + stream_length] = filled_stream.ljust(
        stream_length,
        b"\x00",
    )
    output_path.write_bytes(output_bytes)


def _merge_template_globals_with_generated_sheets(
    source_stream: bytes,
    generated_stream: bytes,
) -> bytes:
    source_globals, _ = _split_biff_globals(source_stream)
    generated_globals, generated_sheets = _split_biff_globals(generated_stream)
    generated_boundsheets = [
        raw for record_id, raw in generated_globals if record_id == _BIFF_BOUNDSHEET
    ]
    source_boundsheet_count = sum(
        record_id == _BIFF_BOUNDSHEET for record_id, _ in source_globals
    )
    if source_boundsheet_count != len(generated_boundsheets):
        raise ValueError("The filled workbook changed the .xls template sheet structure.")

    # Sheet records reference the generated workbook's XF/font/format indices.
    # Keep generated globals, then restore template-defined names and refs.
    name_records = [
        raw
        for record_id, raw in source_globals
        if record_id in _BIFF_NAME_DEPENDENCIES
    ]
    merged_records: list[bytes] = []
    for record_id, raw in generated_globals:
        if record_id == _BIFF_EOF:
            merged_records.extend(name_records)
        merged_records.append(raw)

    generated_global_bytes = b"".join(raw for _, raw in generated_globals)
    generated_sheet_offsets = [
        unpack_from("<I", raw, 4)[0] - len(generated_global_bytes)
        for raw in generated_boundsheets
    ]
    merged_globals = bytearray(b"".join(merged_records))
    boundsheet_idx = 0
    for record_offset, record_id, _ in _iter_biff_records(bytes(merged_globals)):
        if record_id == _BIFF_BOUNDSHEET:
            pack_into(
                "<I",
                merged_globals,
                record_offset + 4,
                len(merged_globals) + generated_sheet_offsets[boundsheet_idx],
            )
            boundsheet_idx += 1
    return bytes(merged_globals) + generated_sheets


def _sst_records(records: list[tuple[int, bytes]]) -> list[bytes]:
    for idx, (record_id, raw) in enumerate(records):
        if record_id != _BIFF_SST:
            continue
        result = [raw]
        for next_record_id, next_raw in records[idx + 1 :]:
            if next_record_id != _BIFF_CONTINUE:
                break
            result.append(next_raw)
        return result
    return []


def _split_biff_globals(stream: bytes) -> tuple[list[tuple[int, bytes]], bytes]:
    records: list[tuple[int, bytes]] = []
    for record_offset, record_id, raw in _iter_biff_records(stream):
        records.append((record_id, raw))
        if record_id == _BIFF_EOF:
            return records, stream[record_offset + len(raw) :]
    raise ValueError("The .xls Workbook stream has no global EOF record.")


def _iter_biff_records(stream: bytes):
    offset = 0
    while offset + 4 <= len(stream):
        record_id, data_length = unpack_from("<HH", stream, offset)
        end = offset + 4 + data_length
        if end > len(stream):
            raise ValueError("The .xls Workbook stream contains a truncated BIFF record.")
        yield offset, record_id, stream[offset:end]
        offset = end


def _copy_column_layout(
    source_sheet: xlrd.sheet.Sheet,
    output_sheet: xlwt.Worksheet.Worksheet,
    max_col_count: int,
) -> None:
    for col_idx in range(max_col_count):
        source_col = source_sheet.colinfo_map.get(col_idx)
        if source_col is None:
            continue
        output_col = output_sheet.col(col_idx)
        output_col.width = source_col.width
        output_col.hidden = source_col.hidden
        output_col.level = source_col.outline_level
        output_col.collapse = source_col.collapsed


def _copy_static_template_rows(
    *,
    source_sheet: xlrd.sheet.Sheet,
    output_sheet: xlwt.Worksheet.Worksheet,
    styles: list[xlwt.Style.XFStyle],
    max_col_count: int,
    data_start_row: int,
) -> None:
    merged_top_left: dict[tuple[int, int], tuple[int, int, int, int]] = {}
    merged_covered: set[tuple[int, int]] = set()
    for rlo, rhi, clo, chi in source_sheet.merged_cells:
        if rlo >= data_start_row:
            continue
        merged_top_left[(rlo, clo)] = (rlo, rhi, clo, chi)
        for row_idx in range(rlo, rhi):
            for col_idx in range(clo, chi):
                if (row_idx, col_idx) != (rlo, clo):
                    merged_covered.add((row_idx, col_idx))

    for row_idx in range(data_start_row):
        _copy_row_layout(source_sheet, output_sheet, row_idx, fallback_row=row_idx)
        for col_idx in range(max_col_count):
            if (row_idx, col_idx) in merged_covered:
                continue
            value = (
                source_sheet.cell_value(row_idx, col_idx)
                if row_idx < source_sheet.nrows and col_idx < source_sheet.ncols
                else ""
            )
            style = _style_for_cell(
                source_sheet,
                styles,
                row_idx,
                col_idx,
                fallback_row=min(row_idx, max(source_sheet.nrows - 1, 0)),
            )
            merged = merged_top_left.get((row_idx, col_idx))
            if merged:
                rlo, rhi, clo, chi = merged
                output_sheet.write_merge(rlo, rhi - 1, clo, chi - 1, value, style)
            else:
                output_sheet.write(row_idx, col_idx, value, style)


def _copy_row_layout(
    source_sheet: xlrd.sheet.Sheet,
    output_sheet: xlwt.Worksheet.Worksheet,
    row_idx: int,
    *,
    fallback_row: int,
) -> None:
    source_row = source_sheet.rowinfo_map.get(row_idx) or source_sheet.rowinfo_map.get(fallback_row)
    if source_row is None:
        return
    output_row = output_sheet.row(row_idx)
    output_row.height = source_row.height
    output_row.height_mismatch = True


def _xlwt_styles_for(book: xlrd.book.Book) -> list[xlwt.Style.XFStyle]:
    writer = XLWTWriter()
    writer.start()
    writer.workbook(book, str(book))
    return list(writer.style_list)


def _style_for_cell(
    sheet: xlrd.sheet.Sheet,
    styles: list[xlwt.Style.XFStyle],
    row_idx: int,
    col_idx: int,
    fallback_row: int,
) -> xlwt.Style.XFStyle:
    source_row = row_idx if row_idx < sheet.nrows else min(fallback_row, sheet.nrows - 1)
    try:
        xf_index = sheet.cell_xf_index(source_row, col_idx)
    except IndexError:
        xf_index = sheet.cell_xf_index(min(fallback_row, sheet.nrows - 1), col_idx)
    if xf_index is None or xf_index >= len(styles):
        return xlwt.Style.default_style
    return styles[xf_index]


def _xls_cell_style(
    value: Any,
    style: xlwt.Style.XFStyle,
    header: str,
) -> xlwt.Style.XFStyle:
    if not _is_date_output(value, style, header) or is_date_format(style.num_format_str):
        return style
    date_style = copy(style)
    date_style.num_format_str = "dd/mm/yyyy hh:mm:ss"
    return date_style


def _xls_cell_value(
    value: Any,
    style: xlwt.Style.XFStyle,
    header: str = "",
) -> Any:
    if isinstance(value, Decimal):
        as_float = float(value)
        if abs(value) >= EXCEL_SAFE_INTEGER_LIMIT or Decimal(str(as_float)) != value:
            raise ValueError(
                "Numeric output exceeds Excel's safe precision; correct the "
                "source value or store an identifier as text before export."
            )
        return int(value) if value == value.to_integral_value() else as_float
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if abs(value) >= EXCEL_SAFE_INTEGER_LIMIT:
            raise ValueError(
                "Numeric output exceeds Excel's safe precision; correct the "
                "source value or store an identifier as text before export."
            )
    if isinstance(value, (datetime, date)):
        return _excel_serial(value)
    if value is None:
        return ""
    if isinstance(value, str) and (
        is_date_format(style.num_format_str) or _is_date_header(header)
    ):
        parsed = parse_date(value)
        if isinstance(parsed, (datetime, date)):
            return _excel_serial(parsed)
    return value


def _is_date_output(value: Any, style: xlwt.Style.XFStyle, header: str) -> bool:
    if isinstance(value, (datetime, date)):
        return True
    if not isinstance(value, str):
        return False
    if not (is_date_format(style.num_format_str) or _is_date_header(header)):
        return False
    return isinstance(parse_date(value), (datetime, date))


def _is_date_header(header: str) -> bool:
    normalized = normalize_header(header)
    return normalized.startswith("ngay_") or normalized in {
        "ngay",
        "han_su_dung",
        "han_thanh_toan",
    }


def _clear_stale_template_rows(
    source_sheet: xlrd.sheet.Sheet,
    output_sheet: xlwt.Worksheet.Worksheet,
    styles: list[xlwt.Style.XFStyle],
    start_row: int,
    max_col_count: int,
    fallback_row: int,
) -> None:
    last_value_row = _last_nonblank_row(source_sheet, start_row, max_col_count)
    for row_idx in range(start_row, last_value_row + 1):
        for col_idx in range(max_col_count):
            if col_idx >= source_sheet.ncols:
                continue
            if is_blank(source_sheet.cell_value(row_idx, col_idx)):
                continue
            style = _style_for_cell(
                source_sheet,
                styles,
                row_idx,
                col_idx,
                fallback_row=fallback_row,
            )
            output_sheet.write(row_idx, col_idx, "", style)


def _last_nonblank_row(sheet: xlrd.sheet.Sheet, start_row: int, max_col_count: int) -> int:
    for row_idx in range(sheet.nrows - 1, start_row - 1, -1):
        values = sheet.row_values(row_idx, 0, min(max_col_count, sheet.ncols))
        if any(not is_blank(value) for value in values):
            return row_idx
    return start_row - 1


def _excel_serial(value: datetime | date) -> float:
    if isinstance(value, datetime):
        dt = value
    else:
        dt = datetime(value.year, value.month, value.day)
    epoch = datetime(1899, 12, 30)
    delta = dt - epoch
    return delta.days + delta.seconds / 86400 + delta.microseconds / 86400 / 1_000_000
