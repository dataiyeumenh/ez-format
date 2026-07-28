from __future__ import annotations

import os
from pathlib import Path
from typing import Any

import openpyxl
import xlrd
from openpyxl.utils.cell import get_column_letter


def inspect_workbook_structure(path: Path) -> dict[str, Any]:
    if path.suffix.lower() == ".xlsx":
        return _inspect_xlsx(path)
    if path.suffix.lower() == ".xls":
        return _inspect_xls(path)
    raise ValueError("Chỉ hỗ trợ workbook .xls hoặc .xlsx")


def enforce_workbook_limits(
    *,
    content_size: int,
    row_count: int,
    column_count: int,
) -> None:
    max_bytes = int(os.getenv("RECONSTRUCTION_MAX_FILE_BYTES", str(30 * 1024 * 1024)))
    max_rows = int(os.getenv("RECONSTRUCTION_MAX_ROWS", "50000"))
    max_cells = int(os.getenv("RECONSTRUCTION_MAX_CELLS", "3000000"))
    if content_size > max_bytes:
        raise ValueError(f"File vượt giới hạn {max_bytes} bytes")
    if row_count > max_rows:
        raise ValueError(f"File vượt giới hạn {max_rows} dòng")
    if row_count * max(1, column_count) > max_cells:
        raise ValueError(f"Workbook vượt giới hạn {max_cells} ô dữ liệu")


def validate_excel_magic(filename: str, content: bytes) -> None:
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".xlsx" and not content.startswith(b"PK\x03\x04"):
        raise ValueError("Nội dung file .xlsx không đúng định dạng ZIP/OpenXML")
    if suffix == ".xls" and not content.startswith(bytes.fromhex("D0CF11E0A1B11AE1")):
        raise ValueError("Nội dung file .xls không đúng định dạng OLE2")


def _inspect_xlsx(path: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(
        path,
        read_only=False,
        data_only=False,
        keep_links=False,
    )
    try:
        sheets = []
        total_formulas = 0
        total_hidden_rows = 0
        total_hidden_columns = 0
        total_merged_ranges = 0
        for sheet in workbook.worksheets:
            hidden_rows = [
                index for index, dimension in sheet.row_dimensions.items() if dimension.hidden
            ]
            hidden_columns = [
                key for key, dimension in sheet.column_dimensions.items() if dimension.hidden
            ]
            formula_cells = [
                cell.coordinate
                for row in sheet.iter_rows()
                for cell in row
                if cell.data_type == "f"
            ]
            merged_ranges = [str(item) for item in sheet.merged_cells.ranges]
            total_formulas += len(formula_cells)
            total_hidden_rows += len(hidden_rows)
            total_hidden_columns += len(hidden_columns)
            total_merged_ranges += len(merged_ranges)
            sheets.append(
                {
                    "name": sheet.title,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "hidden_rows": hidden_rows[:1000],
                    "hidden_columns": hidden_columns[:1000],
                    "formula_cells": formula_cells[:1000],
                    "merged_ranges": merged_ranges[:1000],
                }
            )
        return {
            "format": "xlsx",
            "formula_detection": "available",
            "sheet_count": len(sheets),
            "sheets": sheets,
            "formula_cell_count": total_formulas,
            "hidden_row_count": total_hidden_rows,
            "hidden_column_count": total_hidden_columns,
            "merged_range_count": total_merged_ranges,
            "has_external_links": bool(getattr(workbook, "_external_links", [])),
            "warnings": _warnings(
                total_formulas,
                total_hidden_rows,
                total_hidden_columns,
                total_merged_ranges,
            ),
        }
    finally:
        workbook.close()


def _inspect_xls(path: Path) -> dict[str, Any]:
    workbook = xlrd.open_workbook(str(path), formatting_info=True)
    sheets = []
    hidden_rows = 0
    hidden_columns = 0
    merged_ranges = 0
    for sheet in workbook.sheets():
        sheet_hidden_rows = [
            index + 1
            for index, info in sheet.rowinfo_map.items()
            if getattr(info, "hidden", 0)
        ]
        sheet_hidden_columns = [
            get_column_letter(index + 1)
            for index, info in sheet.colinfo_map.items()
            if getattr(info, "hidden", 0)
        ]
        hidden_rows += len(sheet_hidden_rows)
        hidden_columns += len(sheet_hidden_columns)
        merged_ranges += len(sheet.merged_cells)
        sheets.append(
            {
                "name": sheet.name,
                "max_row": sheet.nrows,
                "max_column": sheet.ncols,
                "hidden_rows": sheet_hidden_rows[:1000],
                "hidden_columns": sheet_hidden_columns[:1000],
                "formula_cells": [],
                "merged_ranges": [list(item) for item in sheet.merged_cells[:1000]],
            }
        )
    return {
        "format": "xls",
        # xlrd exposes formula results but not reliable formula cell metadata for XLS.
        "formula_detection": "unavailable",
        "sheet_count": len(sheets),
        "sheets": sheets,
        "formula_cell_count": 0,
        "hidden_row_count": hidden_rows,
        "hidden_column_count": hidden_columns,
        "merged_range_count": merged_ranges,
        "has_external_links": False,
        "warnings": _warnings(
            0,
            hidden_rows,
            hidden_columns,
            merged_ranges,
            formula_detection_unavailable=True,
        ),
    }


def _warnings(
    formula_count: int,
    hidden_rows: int,
    hidden_columns: int,
    merged_ranges: int,
    *,
    formula_detection_unavailable: bool = False,
) -> list[dict[str, str]]:
    warnings: list[dict[str, str]] = []
    if formula_detection_unavailable:
        warnings.append(
            {
                "code": "formula_detection_unavailable",
                "message": "Không thể xác định tin cậy ô công thức trong workbook .xls.",
            }
        )
    if formula_count:
        warnings.append(
            {
                "code": "formula_cells_detected",
                "message": "Workbook có ô công thức; hệ thống dùng cached value nếu có.",
            }
        )
    if hidden_rows or hidden_columns:
        warnings.append(
            {
                "code": "hidden_rows_or_columns_detected",
                "message": "Workbook có dòng hoặc cột ẩn và cần được kiểm tra.",
            }
        )
    if merged_ranges:
        warnings.append(
            {
                "code": "merged_cells_detected",
                "message": "Workbook có merged cells; header và fill-down cần được kiểm tra.",
            }
        )
    return warnings
