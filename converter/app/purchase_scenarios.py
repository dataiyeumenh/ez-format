from __future__ import annotations

import hashlib
import io
import json
import random
import zipfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


MISA_IMPORT_SOURCE = "https://helpact.misa.vn/kb/html_10050000/"

SEMANTIC_TARGETS: dict[str, str | list[str]] = {
    "classification": ["Hình thức mua hàng", "TK kho/TK chi phí (*)"],
    "payment_method": "Phương thức thanh toán",
    "posting_date": ["Ngày hạch toán (*)", "Ngày chứng từ (*)"],
    "receipt_no": "Số phiếu nhập (*)",
    "invoice_symbol": "Ký hiệu HĐ",
    "invoice_no": "Số hóa đơn",
    "invoice_date": "Ngày hóa đơn",
    "supplier_tax_code": ["Mã nhà cung cấp", "Mã số thuế"],
    "supplier_name": "Tên nhà cung cấp",
    "supplier_address": "Địa chỉ",
    "description": "Diễn giải",
    "item_code": "Mã hàng (*)",
    "item_name": "Tên hàng",
    "unit": "ĐVT",
    "quantity": "Số lượng",
    "unit_price": "Đơn giá",
    "amount": "Thành tiền",
    "discount_rate": "Tỷ lệ CK (%)",
    "discount_amount": "Tiền chiết khấu",
    "vat_rate": "% thuế GTGT",
    "vat_amount": "Tiền thuế GTGT",
    "vat_account": "TK thuế GTGT",
    "credit_account": "TK công nợ/TK tiền (*)",
}

ALIASES: dict[str, dict[str, str]] = {
    "vietnamese": {
        "classification": "Phân loại",
        "payment_method": "Hình thức thanh toán",
        "posting_date": "Ngày chứng từ",
        "receipt_no": "Số phiếu nhập",
        "invoice_symbol": "Ký hiệu hóa đơn",
        "invoice_no": "Số hóa đơn",
        "invoice_date": "Ngày hóa đơn",
        "supplier_tax_code": "Mã số thuế NCC",
        "supplier_name": "Tên nhà cung cấp",
        "supplier_address": "Địa chỉ nhà cung cấp",
        "description": "Diễn giải",
        "item_code": "Mã hàng",
        "item_name": "Tên hàng hóa dịch vụ",
        "unit": "Đơn vị tính",
        "quantity": "Số lượng",
        "unit_price": "Đơn giá",
        "amount": "Thành tiền chưa thuế",
        "discount_rate": "Tỷ lệ chiết khấu",
        "discount_amount": "Tiền chiết khấu",
        "vat_rate": "Thuế suất GTGT",
        "vat_amount": "Tiền thuế GTGT",
        "vat_account": "Tài khoản thuế",
        "credit_account": "Tài khoản công nợ",
    },
    "no_accent": {
        "classification": "Phan loai",
        "payment_method": "Hinh thuc thanh toan",
        "posting_date": "Ngay chung tu",
        "receipt_no": "So phieu nhap",
        "invoice_symbol": "Ky hieu HD",
        "invoice_no": "So hoa don",
        "invoice_date": "Ngay hoa don",
        "supplier_tax_code": "Ma so thue NCC",
        "supplier_name": "Ten nha cung cap",
        "supplier_address": "Dia chi NCC",
        "description": "Dien giai",
        "item_code": "Ma hang",
        "item_name": "Ten hang",
        "unit": "DVT",
        "quantity": "So luong",
        "unit_price": "Don gia",
        "amount": "Thanh tien",
        "discount_rate": "Ty le CK",
        "discount_amount": "Tien CK",
        "vat_rate": "Thue suat GTGT",
        "vat_amount": "Tien thue GTGT",
        "vat_account": "TK thue",
        "credit_account": "TK cong no",
    },
    "accounting_codes": {
        "classification": "LOAI_MUA",
        "payment_method": "HTTT",
        "posting_date": "NGAYCT",
        "receipt_no": "SOCT",
        "invoice_symbol": "SR_HD",
        "invoice_no": "SO_HD",
        "invoice_date": "NGAY_HD",
        "supplier_tax_code": "MADTPNCO",
        "supplier_name": "TENKH",
        "supplier_address": "DIACHI",
        "description": "DIENGIAI",
        "item_code": "MA_VTHH",
        "item_name": "MATHANG",
        "unit": "DONVI",
        "quantity": "LUONG",
        "unit_price": "DGVND",
        "amount": "TTVND",
        "discount_rate": "PT_CK",
        "discount_amount": "CHIETKHAU",
        "vat_rate": "TS_GTGT",
        "vat_amount": "THUEVND",
        "vat_account": "TKTHUE",
        "credit_account": "TKCO",
    },
    "english": {
        "classification": "Purchase type",
        "payment_method": "Payment method",
        "posting_date": "Posting date",
        "receipt_no": "Receipt number",
        "invoice_symbol": "Invoice series",
        "invoice_no": "Invoice number",
        "invoice_date": "Invoice date",
        "supplier_tax_code": "Supplier tax code",
        "supplier_name": "Supplier name",
        "supplier_address": "Supplier address",
        "description": "Description",
        "item_code": "Item code",
        "item_name": "Item name",
        "unit": "Unit",
        "quantity": "Quantity",
        "unit_price": "Unit price",
        "amount": "Net amount",
        "discount_rate": "Discount percent",
        "discount_amount": "Discount amount",
        "vat_rate": "VAT rate",
        "vat_amount": "VAT amount",
        "vat_account": "VAT account",
        "credit_account": "Payable account",
    },
    "einvoice_export": {
        "classification": "Tính chất HHDV",
        "payment_method": "HT thanh toán",
        "posting_date": "Ngày lập CT",
        "receipt_no": "Số CT mua",
        "invoice_symbol": "Ký hiệu mẫu số/Ký hiệu HĐ",
        "invoice_no": "Số HĐ điện tử",
        "invoice_date": "Ngày lập hóa đơn",
        "supplier_tax_code": "MST người bán",
        "supplier_name": "Tên người bán",
        "supplier_address": "Địa chỉ người bán",
        "description": "Nội dung hàng hóa dịch vụ",
        "item_code": "Mã SP trên HĐ",
        "item_name": "Tên HHDV",
        "unit": "ĐVT HĐ",
        "quantity": "SL HĐ",
        "unit_price": "Đơn giá chưa thuế",
        "amount": "Tiền trước thuế",
        "discount_rate": "% CK",
        "discount_amount": "Tiền CK thương mại",
        "vat_rate": "Thuế suất",
        "vat_amount": "Tiền VAT",
        "vat_account": "TK thuế GTGT được khấu trừ",
        "credit_account": "TK phải trả người bán",
    },
}

LAYOUTS = (
    "flat",
    "title_rows",
    "summary_and_detail",
    "noisy_reordered",
    "merged_hidden_formula",
)
CATEGORIES = ("goods", "service", "mixed", "adjustment")


@dataclass(frozen=True)
class PurchaseScenario:
    id: str
    category: str
    alias_profile: str
    layout_profile: str
    expected_sheet: str
    header_row: int
    expected_row_count: int
    expected_mapping: dict[str, str | list[str]]
    warnings: tuple[str, ...]
    blockers: tuple[str, ...]
    seed: int
    source_url: str = MISA_IMPORT_SOURCE

    @property
    def schema_fingerprint(self) -> str:
        payload = f"{self.category}|{self.alias_profile}|{self.layout_profile}|{self.seed}"
        return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def build_purchase_scenarios() -> list[PurchaseScenario]:
    scenarios: list[PurchaseScenario] = []
    index = 0
    for category in CATEGORIES:
        for alias_profile in ALIASES:
            for layout_profile in LAYOUTS:
                index += 1
                sheet_name, header_row = _layout_location(layout_profile)
                aliases = ALIASES[alias_profile]
                warnings = ["account_review_required"]
                blockers: list[str] = []
                if category == "adjustment":
                    warnings.append("negative_amount_context_unclear")
                scenarios.append(
                    PurchaseScenario(
                        id=f"purchase_{index:03d}",
                        category=category,
                        alias_profile=alias_profile,
                        layout_profile=layout_profile,
                        expected_sheet=sheet_name,
                        header_row=header_row,
                        expected_row_count=2,
                        expected_mapping={
                            aliases[semantic]: target
                            for semantic, target in SEMANTIC_TARGETS.items()
                        },
                        warnings=tuple(warnings),
                        blockers=tuple(blockers),
                        seed=10_000 + index,
                    )
                )
    return scenarios


def write_purchase_scenario_workbook(scenario: PurchaseScenario, path: Path) -> Path:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    if scenario.layout_profile == "summary_and_detail":
        summary = workbook.create_sheet("TongHopHoaDon")
        summary.append(["STT", "Thông tin", "Giá trị"])
        summary.append([1, "Số hóa đơn", f"HD-{scenario.id}"])

    sheet = workbook.create_sheet(scenario.expected_sheet)
    if scenario.header_row > 1:
        for row_index in range(1, scenario.header_row):
            sheet.cell(row_index, 1, f"DỮ LIỆU MUA VÀO - {scenario.id}")
        if scenario.layout_profile == "merged_hidden_formula":
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    aliases = ALIASES[scenario.alias_profile]
    semantic_order = list(SEMANTIC_TARGETS)
    if scenario.layout_profile in {"noisy_reordered", "merged_hidden_formula"}:
        random.Random(scenario.seed).shuffle(semantic_order)
    headers = [aliases[key] for key in semantic_order]
    extra_headers = _extra_headers(scenario)
    if scenario.layout_profile in {"noisy_reordered", "merged_hidden_formula"}:
        headers = [extra_headers[0], *headers, *extra_headers[1:]]
    else:
        headers.extend(extra_headers)
    sheet.append(headers)

    rows = _scenario_rows(scenario)
    for row_number, semantic_values in enumerate(rows, start=1):
        values_by_header = {
            aliases[semantic]: _format_value(value, semantic, scenario, row_number)
            for semantic, value in semantic_values.items()
        }
        values_by_header.update(
            {
                extra_headers[0]: f"SRC-{scenario.id}-{row_number}",
                extra_headers[1]: "",
                extra_headers[2]: f"=1+{row_number}",
            }
        )
        sheet.append([values_by_header.get(header, "") for header in headers])

    if scenario.layout_profile == "merged_hidden_formula":
        sheet.row_dimensions[scenario.header_row + 2].hidden = True
    sheet.freeze_panes = f"A{scenario.header_row + 1}"
    _save_deterministic_xlsx(workbook, path)
    return path


def scenario_catalog_payload(scenarios: list[PurchaseScenario] | None = None) -> list[dict[str, Any]]:
    return [
        {
            "id": item.id,
            "category": item.category,
            "alias_profile": item.alias_profile,
            "layout_profile": item.layout_profile,
            "expected_sheet": item.expected_sheet,
            "header_row": item.header_row,
            "expected_row_count": item.expected_row_count,
            "expected_mapping": item.expected_mapping,
            "warnings": list(item.warnings),
            "blockers": list(item.blockers),
            "seed": item.seed,
            "source_url": item.source_url,
        }
        for item in (scenarios or build_purchase_scenarios())
    ]


def _layout_location(layout_profile: str) -> tuple[str, int]:
    return {
        "flat": ("MuaVao", 1),
        "title_rows": ("DuLieuMuaVao", 4),
        "summary_and_detail": ("ChiTietHoaDon", 1),
        "noisy_reordered": ("Data", 8),
        "merged_hidden_formula": ("NhapLieu", 12),
    }[layout_profile]


def _extra_headers(scenario: PurchaseScenario) -> list[str]:
    return [
        f"Nguồn dữ liệu {scenario.id}",
        f"Ghi chú nội bộ {scenario.seed}",
        f"Cột công thức {scenario.alias_profile}",
    ]


def _scenario_rows(scenario: PurchaseScenario) -> list[dict[str, Any]]:
    if scenario.category == "goods":
        classes = ("Hàng hóa", "Hàng hóa")
    elif scenario.category == "service":
        classes = ("Dịch vụ", "Dịch vụ")
    else:
        classes = ("Dịch vụ", "Hàng hóa")
    amounts = (2_905_880, 1_250_000)
    if scenario.category == "adjustment":
        amounts = (-500_000, 1_250_000)
    rows: list[dict[str, Any]] = []
    for offset, classification in enumerate(classes, start=1):
        amount = amounts[offset - 1]
        vat_rate: Any = (0, 5, 8, 10, "KCT")[scenario.seed % 5]
        numeric_vat = vat_rate if isinstance(vat_rate, int) else 0
        rows.append(
            {
                "classification": classification,
                "payment_method": ("Chưa thanh toán", "Chuyển khoản")[offset - 1],
                "posting_date": datetime(2026, 4, offset + 1),
                "receipt_no": f"PN{scenario.seed}{offset}",
                "invoice_symbol": f"1C26T{chr(64 + offset)}A",
                "invoice_no": f"HD{scenario.seed}{offset}",
                "invoice_date": datetime(2026, 4, offset + 1),
                "supplier_tax_code": f"03{scenario.seed:08d}{offset}"[:10],
                "supplier_name": f"Nhà cung cấp tổng hợp {offset}",
                "supplier_address": "Thành phố Hồ Chí Minh",
                "description": f"Mua vào synthetic {classification.lower()}",
                "item_code": f"MV-{scenario.seed}-{offset}",
                "item_name": f"Mặt hàng synthetic {offset}",
                "unit": "Cái" if classification == "Hàng hóa" else "Lần",
                "quantity": 1 if classification == "Dịch vụ" else 10,
                "unit_price": amount if classification == "Dịch vụ" else amount / 10,
                "amount": amount,
                "discount_rate": 0,
                "discount_amount": 0,
                "vat_rate": vat_rate,
                "vat_amount": round(amount * numeric_vat / 100),
                "vat_account": "1331",
                "credit_account": "331" if offset == 1 else "1121",
            }
        )
    return rows


def _format_value(value: Any, semantic: str, scenario: PurchaseScenario, row_number: int) -> Any:
    variant = scenario.seed % 5
    if isinstance(value, datetime):
        if variant == 1:
            return value.strftime("%d/%m/%Y")
        if variant == 2:
            return value.strftime("%Y-%m-%d")
        if variant == 3:
            return value.strftime("%d-%m-%Y")
        return value
    if semantic in {"quantity", "unit_price", "amount", "discount_amount", "vat_amount"}:
        if variant == 1 and isinstance(value, (int, float)):
            return f"{value:,.0f}".replace(",", ".")
        if variant == 2 and isinstance(value, (int, float)):
            return f"{value:,.2f}"
        if variant == 3 and isinstance(value, (int, float)) and value < 0:
            return f"({abs(value):,.0f})"
    return value


def _save_deterministic_xlsx(workbook: openpyxl.Workbook, path: Path) -> None:
    workbook.properties.created = datetime(2026, 1, 1)
    workbook.properties.modified = datetime(2026, 1, 1)
    buffer = io.BytesIO()
    workbook.save(buffer)
    source = zipfile.ZipFile(io.BytesIO(buffer.getvalue()), "r")
    output = io.BytesIO()
    with source, zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as target:
        for name in sorted(source.namelist()):
            info = zipfile.ZipInfo(name, date_time=(2020, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = source.getinfo(name).external_attr
            target.writestr(info, source.read(name))
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(output.getvalue())

